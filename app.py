# app.py
from flask import Flask, g, redirect, url_for, request, render_template, flash, jsonify, session, Response, stream_with_context, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from functools import wraps
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from apscheduler.schedulers.background import BackgroundScheduler
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import glob
import atexit
import json
import subprocess
import shutil
import bcrypt
import os
import sys
import uuid
import time
import threading
import pickle
from itsdangerous import TimestampSigner
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import pytz
import time
from collections import defaultdict
import psutil
import time
import platform
import concurrent.futures
from datetime import timedelta
import random

# Importar módulo de seguridad de tesorería
from modules.tesoreria_security import (
    init_security,
    csrf_protected,
    rate_limited,
    tesoreria_secured,
    CSRFProtection,
    AuditLogger
)

# ========================================================================
# ===== CONFIGURACIÓN INICIAL DE LA APLICACIÓN =====
# ========================================================================

# acá, declaro la app y ni bien declaro la app ya lo que hago es crear los decoradores
# .. y funciones necesarias, con el fin de que estén disponibles para los blueprints
# que importe luego y sea inyectable en los blueprints
# esto debe ser así ya que cargar el app.py está tardando mucho y no llega a cargar
# los decoradores antes de que los blueprints los necesiten

app = Flask(__name__)







@app.context_processor
def inject_role_level():
    try:
        lvl = get_user_level()
    except Exception:
        lvl = 1
    return {"role_level": int(lvl)}

import mysql.connector

def get_db_connection():
    db_name    = os.getenv("DB_NAME", "cajasdb")
    db_user    = os.getenv("DB_USER", "app_cajas")
    db_pass    = os.getenv("DB_PASS")
    db_charset = os.getenv("DB_CHARSET", "utf8mb4")
    db_tz      = os.getenv("DB_TIMEZONE", "America/Argentina/Buenos_Aires")
    db_host    = os.getenv("DB_HOST", "")      # puede ser '/cloudsql/PROJECT:REGION:INSTANCE'
    db_port    = int(os.getenv("DB_PORT", ""))

    if not db_pass:
        raise RuntimeError("DB_PASS no seteada")

    # kwargs comunes
    base_kwargs = dict(
        user=db_user,
        password=db_pass,
        database=db_name,
        autocommit=True,
        charset=db_charset,
        use_unicode=True,
    )

    # Si DB_HOST empieza con /cloudsql usamos Unix Socket (Cloud SQL)
    if db_host.startswith("/cloudsql/"):
        conn = mysql.connector.connect(
            unix_socket=db_host,
            **base_kwargs
        )
    else:
        # fallback a host:puerto (por ejemplo en local)
        conn = mysql.connector.connect(
            host=db_host or "127.0.0.1",
            port=db_port,
            **base_kwargs
        )

    # Setear zona horaria (opcional, si tu MySQL lo permite)
    try:
        with conn.cursor() as cur:
            cur.execute("SET time_zone = %s", (db_tz,))
    except Exception:
        pass

    return conn

def _normalize_fecha(fecha):
    if isinstance(fecha, datetime):
        return fecha.date()
    if isinstance(fecha, date):
        return fecha
    if isinstance(fecha, str):
        try:
            return datetime.strptime(fecha, "%Y-%m-%d").date()
        except ValueError:
            return None
    return None

def get_user_level() -> int:
    """
    Devuelve el nivel del rol desde sesión.
    Asume que guardaste 'role_level' en session en el login.
    Si no, mapeá role -> level acá.

    Niveles:
    1 = cajero
    2 = encargado/administrativo
    3 = auditor
    4 = anticipos (solo crea anticipos en locales asignados)
    5 = jefe_auditor
    6 = admin_anticipos (gestiona todos los anticipos)
    """
    lvl = session.get('role_level')
    if lvl is not None:
        return int(lvl)
    role = (session.get('role') or '').strip().lower()
    MAP = {
        'cajero': 1,
        'encargado': 2,
        'administrativo': 2,
        'auditor': 3,
        'anticipos': 4,  # Rol limitado: solo crea anticipos en locales asignados
        'jefe_auditor': 5,
        'admin_anticipos': 6  # Rol especial: gestiona todos los anticipos
    }
    return MAP.get(role, 0)

def get_user_allowed_locales() -> list:
    """
    Obtiene los locales a los que el usuario tiene acceso para anticipos.
    - Si el nivel es 6 (admin_anticipos), tiene acceso a TODOS los locales
    - Si el nivel es 4 (anticipos), consulta la tabla user_local_permissions
    - Caso contrario, retorna lista vacía (sin acceso a anticipos)
    """
    lvl = get_user_level()

    # Admin de anticipos (nivel 6): acceso total a todos los locales
    if lvl == 6:
        return []  # Lista vacía significa "todos los locales"

    # Usuario con rol 'anticipos' (nivel 4): consultar permisos específicos
    if lvl == 4:
        username = session.get('username')
        if not username:
            return []

        try:
            conn = get_db_connection()
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                SELECT local FROM user_local_permissions
                WHERE username = %s
                ORDER BY local ASC
            """, (username,))
            rows = cur.fetchall()
            cur.close()
            conn.close()
            return [row['local'] for row in rows]
        except Exception as e:
            print(f"⚠️  Error obteniendo permisos de locales: {e}")
            return []

    # Otros roles (cajero, encargado, auditor, jefe_auditor): sin acceso a anticipos
    return []

def can_user_access_local_for_anticipos(local: str) -> bool:
    """
    Verifica si el usuario actual puede acceder a un local específico para anticipos.
    - Si nivel == 6 (admin_anticipos): puede acceder a todos
    - Si nivel == 4 (anticipos): verifica en user_local_permissions
    - Otros niveles: False
    """
    lvl = get_user_level()

    # Admin de anticipos (nivel 6): acceso total
    if lvl == 6:
        return True

    # Usuario con rol 'anticipos' (nivel 4): verificar permiso específico
    if lvl == 4:
        allowed_locales = get_user_allowed_locales()
        return local in allowed_locales

    # Otros roles: sin acceso
    return False

def is_local_closed(conn, local:str, fecha) -> bool:
    cur = conn.cursor()
    cur.execute("""
        SELECT estado FROM cierres_locales
        WHERE local=%s AND fecha=%s
        LIMIT 1
    """, (local, _normalize_fecha(fecha)))
    row = cur.fetchone()
    cur.close()
    # Si no existe registro, lo consideramos abierto (estado=1)
    return (row is not None and int(row[0]) == 0)


def is_local_auditado(conn, local: str, fecha) -> bool:
    """
    Verifica si un local está marcado como auditado para una fecha específica.
    Retorna True si existe un registro en locales_auditados.
    """
    f = _normalize_fecha(fecha)
    if not f:
        return False

    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT COUNT(*) FROM locales_auditados
            WHERE local=%s AND DATE(fecha)=%s
        """, (local, f))
        row = cur.fetchone()
        return row and row[0] > 0
    finally:
        cur.close()

def can_edit(conn, local: str, caja: str, turno: str, fecha, user_level: int) -> bool:
    """
    Reglas:
    - Si el LOCAL está AUDITADO: NADIE puede modificar (ni siquiera nivel 3).
    - Si el LOCAL está cerrado (por día): sólo nivel 3 puede modificar.
    - Si el LOCAL está abierto:
        * Nivel 2 (o más) puede modificar aunque la CAJA esté cerrada.
        * Nivel 1 sólo puede modificar con CAJA abierta (por turno).
    """
    f = _normalize_fecha(fecha)

    # NUEVA REGLA: Si está auditado, NADIE puede editar
    if is_local_auditado(conn, local, f):
        return False

    # Cierre de local es global por fecha (ignora turno)
    if is_local_closed(conn, local, f):
        return user_level >= 3

    # Local abierto: L2+ puede modificar aunque la caja esté cerrada
    if user_level >= 2:
        return True

    # L1: sólo si la caja (para ese turno) está abierta
    return is_caja_abierta(conn, local, caja, f, turno)


def can_edit_remesa_retirada(conn, local: str, caja: str, turno: str, fecha, user_level: int) -> bool:
    """
    Permisos especiales para editar 'retirada_por' y 'fecha_retirada' de remesas:
    - Nivel 3 (auditor/admin): Puede editar SIEMPRE (es solo informativo para flujo de dinero)
    - Nivel 2 (encargado): Puede editar SIEMPRE (es solo informativo para flujo de dinero)
    - Nivel 1 (cajero): Puede editar mientras local NO esté auditado (puede estar cerrado el local y/o la caja)
    """
    # Niveles 2 y 3 (encargado/admin/auditor): SIEMPRE pueden actualizar retirada, sin importar estado
    # Ya que es solo informativo para flujo de dinero, no afecta datos contables
    if user_level >= 2:
        return True

    f = _normalize_fecha(fecha)

    # Nivel 1 (cajero): solo puede editar si NO está auditado
    if is_local_auditado(conn, local, f):
        return False

    # Nivel 1: puede editar si no está auditado
    return True

# ==== RBAC Lectura/Escritura común ====
from flask import g, request, session, jsonify
from functools import wraps

# --- Subquery: "caja cerrada" (usa tu tabla de estado de cajas) ---
# Ajustado a: tablas "cajas_estado", columna fecha_operacion y estado=0 (cerrada).
CLOSED_BOX_SUBQUERY = """
  EXISTS (
    SELECT 1
    FROM cajas_estado cc
    WHERE cc.local = {a}.local
      AND cc.caja  = {a}.caja
      AND LOWER(cc.turno) = LOWER({a}.turno)
      AND DATE(cc.fecha_operacion) = DATE({a}.fecha)
      AND cc.id = (
        SELECT MAX(cc2.id)
        FROM cajas_estado cc2
        WHERE cc2.local = cc.local
          AND cc2.caja  = cc.caja
          AND LOWER(cc2.turno) = LOWER(cc.turno)
          AND DATE(cc2.fecha_operacion) = DATE(cc.fecha_operacion)
      )
      AND cc.estado = 0  -- el ÚLTIMO estado del día/turno es "cerrada"
  )
"""


# --- Subquery: "local cerrado" (usa tu tabla de cierres de local) ---
# Ajustado a: tabla "cierres_locales", columna fecha y estado=0 (cerrado).
CLOSED_LOCAL_SUBQUERY = """
  EXISTS (
    SELECT 1
    FROM cierres_locales cl
    WHERE cl.local = {a}.local
      AND DATE(cl.fecha) = DATE({a}.fecha)
      AND cl.estado = 0  -- 0 = cerrado
  )
"""

# --- Subquery: "local auditado" (usa tabla locales_auditados) ---
AUDITED_LOCAL_SUBQUERY = """
  EXISTS (
    SELECT 1
    FROM locales_auditados la
    WHERE la.local = {a}.local
      AND DATE(la.fecha) = DATE({a}.fecha)
  )
"""

def _ctx_from_request():
    """Extrae contexto (local/caja/fecha/turno) desde args/JSON + session.local."""
    data = request.get_json(silent=True) or {}

    # Para auditores: priorizar local del request/body, sino usar get_local_param()
    # Para otros roles: usar get_local_param() que devuelve session['local']
    local_from_request = request.args.get('local') or data.get('local')
    if local_from_request:
        local_value = local_from_request
    else:
        local_value = get_local_param()

    return {
        'local': local_value,
        'caja':  request.args.get('caja')  or data.get('caja'),
        'fecha': request.args.get('fecha') or data.get('fecha'),
        'turno': request.args.get('turno') or data.get('turno'),
    }

def get_local_param():
    """
    Obtiene el local según el nivel del usuario:
    - Nivel 3 (auditor): Usa request.args.get('local') si existe y no está vacío, sino session
    - Nivel 1-2: Usa session['local'] siempre (no pueden cambiar de local)
    """
    lvl = get_user_level()
    if lvl >= 3:
        # Auditor: prioridad al parámetro si no está vacío, fallback a session
        local_param = (request.args.get('local') or '').strip()
        return local_param if local_param else session.get('local')
    else:
        # Cajero/Encargado: solo su local de sesión
        return session.get('local')

def read_scope_sql(alias: str = 't') -> str:
    """
    Qué filas puede VER cada rol en endpoints de lectura (SELECT):
      - Nivel 1 (cajero): sin filtro extra → ve todo lo suyo.
      - Nivel 2 (encargado): SOLO cajas cerradas → AND CLOSED_BOX_SUBQUERY.
      - Nivel 3 (auditor): SOLO locales cerrados → AND CLOSED_LOCAL_SUBQUERY.
    """
    lvl = get_user_level()
    if lvl >= 3:
        # Auditor: solo locales cerrados
        return f" AND {CLOSED_LOCAL_SUBQUERY.format(a=alias)}"
    if lvl >= 2:
        # Encargado/Administrativo: solo cajas cerradas
        return f" AND {CLOSED_BOX_SUBQUERY.format(a=alias)}"
    # Cajero: sin filtro adicional de lectura
    return ""

def with_read_scope(alias='t'):
    """
    Decorador para endpoints GET:
      - Expone g.read_scope con el fragmento SQL de visibilidad según el rol.
    Uso:
      @app.route('/remesas_hoy')
      @login_required
      @with_read_scope('t')
      def remesas_hoy():
          sql = f\"\"\"SELECT ... FROM remesas_trns t WHERE ... {g.read_scope} ORDER BY ...\"\"\"
          ...
    """
    def deco(fn):
        @wraps(fn)
        def wrapper(*a, **k):
            g.read_scope = read_scope_sql(alias)
            return fn(*a, **k)
        return wrapper
    return deco

def require_edit_ctx(fn):
    """
    Decorador para endpoints de ESCRITURA (POST/PUT/DELETE):
      - Ensambla contexto (local, caja, fecha, turno)
      - Valida permiso con can_edit(...)
      - Si todo ok, expone g.ctx con el contexto validado.
    Uso:
      @app.route('/remesas/<int:id>', methods=['PUT'])
      @login_required
      @require_edit_ctx
      def update_remesa(id):
          # g.ctx['local'], g.ctx['caja'], g.ctx['fecha'], g.ctx['turno']
          ...
    """
    @wraps(fn)
    def wrapper(*a, **k):
        ctx = _ctx_from_request()
        if not all([ctx['local'], ctx['caja'], ctx['fecha'], ctx['turno']]):
            return jsonify(success=False, msg='falta local/caja/fecha/turno'), 400
        conn = get_db_connection()
        try:
            # Verificar primero si está auditado para mensaje más claro
            if is_local_auditado(conn, ctx['local'], ctx['fecha']):
                return jsonify(success=False, msg='❌ El local está AUDITADO para esta fecha. No se pueden realizar más modificaciones.'), 403
            ok = can_edit(conn, ctx['local'], ctx['caja'], ctx['turno'], ctx['fecha'], get_user_level())
        finally:
            conn.close()
        if not ok:
            # 409 = conflicto por estado/rol (no permitido)
            return jsonify(success=False, msg='No permitido para tu rol/estado'), 409
        g.ctx = ctx
        return fn(*a, **k)
    return wrapper

def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get('user_id'):
            # guarda next (ruta original) solo para GET HTML
            if request.method == 'GET' and request.accept_mimetypes.accept_html:
                return redirect(url_for('login', next=request.full_path))
            return jsonify(success=False, msg='Auth requerido'), 401

        # IMPORTANTE: Forzar redirección para usuarios de anticipos
        # Si el usuario tiene rol 'anticipos' (nivel 4) o 'admin_anticipos' (nivel 6)
        # y NO está accediendo a rutas permitidas, redirigir a /gestion-anticipos
        user_level = get_user_level()
        current_endpoint = request.endpoint

        # Rutas permitidas para usuarios de anticipos
        allowed_endpoints_anticipos = [
            'gestion_anticipos_page',
            'logout',
            'api_mi_perfil_anticipos',
            'listar_anticipos_recibidos',
            'crear_anticipo_recibido',
            'editar_anticipo_recibido',
            'eliminar_anticipo_recibido',
            'api_locales',
            'api_locales_options',
            'auditoria.auditoria_locales',  # Endpoint /api/locales en blueprint auditoria
            'files_upload',
            'files_list',
            'files_download',
            'bp_files.upload',
            'bp_files.list',
            'bp_files.download',
            'static',
            'gestion_usuarios',  # Solo para admin_anticipos (nivel 6)
            'api_usuarios_anticipos_listar',
            'api_usuarios_anticipos_crear',
            'api_usuarios_anticipos_asignar_local',
            'api_usuarios_anticipos_quitar_local',
            'api_usuarios_anticipos_resetear_password',
            # Endpoints de medios de pago para anticipos
            'api_medios_anticipos_listar',  # Solo admin_anticipos (nivel 6)
            'api_medios_anticipos_crear',   # Solo admin_anticipos (nivel 6)
            'api_medios_anticipos_eliminar', # Solo admin_anticipos (nivel 6)
            'api_medios_anticipos_activos'  # Todos los usuarios de anticipos
        ]

        # Si es usuario de anticipos (nivel 4 o 6) y NO está en una ruta permitida
        if user_level in [4, 6] and current_endpoint not in allowed_endpoints_anticipos:
            # DEBUG: Imprimir endpoint actual
            print(f"🚫 ACCESO DENEGADO - Endpoint: '{current_endpoint}', Path: {request.path}, Nivel: {user_level}")

            # Detectar si es una petición API (empieza con /api/ o es AJAX)
            is_api_request = (
                request.path.startswith('/api/') or
                request.path.startswith('/files/') or
                request.headers.get('X-Requested-With') == 'XMLHttpRequest' or
                'application/json' in request.headers.get('Accept', '')
            )

            # Solo redirigir si es una petición GET de HTML (no APIs)
            if not is_api_request and request.method == 'GET':
                return redirect(url_for('gestion_anticipos_page'))
            # Si es API, retornar error 403
            return jsonify(success=False, msg='No tenés acceso a esta sección'), 403

        return view(*args, **kwargs)
    return wrapped





def _is_safe_url(target: str) -> bool:
    """Evita open-redirects: sólo permite URLs relativas o del mismo host."""
    if not target:
        return False
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return (test_url.scheme in ("http", "https") and ref_url.netloc == test_url.netloc)

# ---------- Router por nivel ----------
def route_for_current_role() -> str:
    """
    Devuelve la URL destino según nivel:
      1 -> 'index'            => '/'
      2 -> 'encargado'        => '/encargado' (o el que tengas: p.ej. 'ventas_cargadas')
      3 -> 'auditor'          => '/auditor'
      4 -> 'anticipos'        => '/gestion-anticipos'
      5+ -> 'jefe_auditor'    => '/auditor'
      6+ -> 'admin_anticipos' => '/gestion-anticipos'
    """
    try:
        lvl = int(get_user_level())
    except Exception:
        lvl = 1

    if lvl == 2:
        # Encargado
        return url_for('encargado')
    if lvl == 4:
        # Anticipos
        return url_for('gestion_anticipos_page')
    if lvl >= 7:
        # Tesorería (7) y Jefe de Tesorería (8+)
        return url_for('tesoreria_home_new')
    if lvl >= 6:
        # Admin de anticipos
        return url_for('gestion_anticipos_page')
    if lvl >= 3:
        # Auditor y jefe_auditor
        return url_for('auditor')
    return url_for('index')












@app.route('/home')
@app.route('/inicio')
@login_required
def go_home():
    """
    Atajo para entrar siempre al lugar correcto según el nivel.
    Útil como landing post-login.
    """
    return redirect(route_for_current_role())

def redirect_after_login():
    nxt = request.args.get('next') or (request.form.get('next') if hasattr(request, 'form') else None)
    lvl = get_user_level()

    # Usuario con rol 'tesoreria' (nivel 7+) siempre va a /tesoreria
    if lvl >= 7:
        return redirect(url_for('tesoreria_home_new'))

    # Usuario con rol 'anticipos' (nivel 4) siempre va a /gestion-anticipos
    if lvl == 4:
        return redirect(url_for('gestion_anticipos_page'))

    # Admin de anticipos (nivel 6) también va a /gestion-anticipos
    if lvl == 6:
        return redirect(url_for('gestion_anticipos_page'))

    if nxt and _is_safe_url(nxt):
        path_only = urlparse(nxt).path or '/'
        if lvl == 2 and path_only in ('/', url_for('index')):
            return redirect(url_for('encargado'))   # <-- endpoint real de encargado
        if lvl >= 3 and path_only in ('/', url_for('index')):
            return redirect(url_for('auditor'))
        return redirect(nxt)
    return redirect(route_for_current_role())



# Inyectamos también la URL "inicio" correcta en los templates
@app.context_processor
def inject_role_level_and_home():
    try:
        lvl = get_user_level()
    except Exception:
        lvl = 1
    try:
        home_url = route_for_current_role()
    except Exception:
        home_url = url_for('index')
    return {
        "role_level": int(lvl),
        "home_url": home_url,
    }










def page_access_required(slug):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            pages = session.get('pages') or []
            if slug not in pages:
                return jsonify(success=False, msg=f"Sin acceso a {slug}"), 403
            return view(*args, **kwargs)
        return wrapped
    return decorator

def role_min_required(min_level:int):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            lvl = int(session.get('role_level') or 0)
            if lvl < min_level:
                return jsonify(success=False, msg='Rol insuficiente'), 403
            return view(*args, **kwargs)
        return wrapped
    return decorator
# app.py
from modules import files_gcs

# ... acá definís login_required, get_db_connection, can_edit, get_user_level, _normalize_fecha ...

files_gcs.inject_dependencies(
    login_required=login_required,
    get_db_connection=get_db_connection,
    can_edit=can_edit,
    get_user_level=get_user_level,
    _normalize_fecha_fn=_normalize_fecha,
    get_local_param=get_local_param,
)

from modules.files_gcs import bp_files     # ← importa el blueprint
app.register_blueprint(bp_files, url_prefix="/files")
from modules.auditoria import auditoria_bp
app.register_blueprint(auditoria_bp)
from modules.terminales import terminales_bp
app.register_blueprint(terminales_bp)
from modules.tabla_auditoria import tabla_auditoria_bp, registrar_auditoria, obtener_registro_anterior
app.register_blueprint(tabla_auditoria_bp)


app.secret_key = '8V#n*aQHYUt@7MdGBY0wE8f'  # Cambiar en producción
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///registros.db'
app.config['SESSION_COOKIE_SECURE'] = False  # Cambiar a True solo en producción con HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Permite cookies en navegación normal
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=3)  # Aumentar a 3 días
app.config['DATA_FOLDER'] = 'c:\\Users\\PROPIETARIO\\Downloads\\01.Proyectos\\form-project\\data'
db = SQLAlchemy(app)
from dotenv import load_dotenv
load_dotenv()

# Inicializar sistema de seguridad de tesorería
init_security(app)



# ________________
# 
# 
# _________________________________________________________________________________ #


@app.route('/', endpoint='index')
@login_required
@page_access_required('index')
def nuevo_registro():
    lvl = get_user_level()
    if lvl == 2:
        return redirect(url_for('encargado'))   # <-- endpoint real de encargado
    if lvl >= 3:
        return redirect(url_for('auditor'))

    # Nivel 1: render normal
    local = session.get('local')
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT cantidad_cajas FROM locales WHERE local = %s LIMIT 1", (local,))
    row = cur.fetchone()
    cantidad_cajas = row[0] if row else 1

    cur.execute("SELECT turnos FROM locales WHERE local = %s", (local,))
    turnos = [r[0] for r in cur.fetchall()] or ['UNI']

    cur.close(); conn.close()
    return render_template('index.html', cantidad_cajas=cantidad_cajas, turnos=turnos)


@app.route('/encargado', endpoint='encargado')
@login_required
@role_min_required(2)  # o el slug que corresponda
def encargado():
    # Defensa por nivel (opcional, pero recomendado)
    if get_user_level() < 2:
        return redirect(route_for_current_role())

    local = session.get('local')

    # Default seguros si por algún motivo no hay local en sesión
    cantidad_cajas = 1
    turnos = ['UNI']

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # cantidad de cajas
        cur.execute("SELECT cantidad_cajas FROM locales WHERE local = %s LIMIT 1", (local,))
        row = cur.fetchone()
        if row and row[0]:
            cantidad_cajas = int(row[0])

        # turnos habilitados
        cur.execute("SELECT turnos FROM locales WHERE local = %s", (local,))
        rows = cur.fetchall()
        turnos = [r[0] for r in rows] or ['UNI']

        cur.close(); conn.close()
    except Exception:
        # En caso de fallo DB, mantenemos defaults para no romper la vista
        pass

    # IMPORTANTE: pasar las variables que la plantilla espera
    return render_template(
        'index_encargado.html',
        cantidad_cajas=cantidad_cajas,
        turnos=turnos
    )


@app.route('/auditor', endpoint='auditor')
@login_required
@role_min_required(3)  # o el slug que corresponda
def encargado():
    # Defensa por nivel (opcional, pero recomendado)
    if get_user_level() < 3:
        return redirect(route_for_current_role())

    local = session.get('local')

    # Default seguros si por algún motivo no hay local en sesión
    cantidad_cajas = 1
    turnos = ['UNI']

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # cantidad de cajas
        cur.execute("SELECT cantidad_cajas FROM locales WHERE local = %s LIMIT 1", (local,))
        row = cur.fetchone()
        if row and row[0]:
            cantidad_cajas = int(row[0])

        # turnos habilitados
        cur.execute("SELECT turnos FROM locales WHERE local = %s", (local,))
        rows = cur.fetchall()
        turnos = [r[0] for r in rows] or ['UNI']

        cur.close(); conn.close()
    except Exception:
        # En caso de fallo DB, mantenemos defaults para no romper la vista
        pass

    # IMPORTANTE: pasar las variables que la plantilla espera
    return render_template(
        'index_auditor.html',
        cantidad_cajas=cantidad_cajas,
        turnos=turnos
    )

@app.route('/tesoreria')
@login_required
@role_min_required(7)  # Solo tesorería (nivel 7+)
def tesoreria_old_redirect():
    """
    Redirección de la ruta antigua /tesoreria a /tesoreria/home
    """
    return redirect(url_for('tesoreria_home_new'))


# __________________________________REMESAS__________________________________________#
# =========================== REMESAS (turno + usuario) ============================

# =========================== REMESAS (turno + usuario) ============================
# ===============================
# REMESAS – LECTURAS (GET)
# ===============================
# Quita cualquier @with_read_scope acá
@app.route('/remesas_no_retiradas')
@login_required
def remesas_no_retiradas():
    caja   = request.args.get("caja")
    local  = get_local_param()
    fecha  = request.args.get("fecha")  # Ahora es requerida
    turno  = request.args.get("turno")  # opcional (si lo querés usar)
    lvl    = get_user_level()

    if not (caja and local and fecha):
        return jsonify([])

    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        # CAMBIO: Ahora solo muestra remesas NO retiradas de la FECHA ACTUAL
        # Las remesas no retiradas se gestionan desde /remesas-no-retiradas (nuevo botón sidebar)
        # Por lo tanto, NO deben "arrastrarse" a fechas futuras

        extra_sql = ""
        params = [caja, local, _normalize_fecha(fecha)]

        # (opcional) si querés además filtrar por turno de selección actual:
        # if turno:
        #     extra_sql += " AND LOWER(t.turno) = LOWER(%s)"
        #     params.append(turno)

        sql = f"""
          SELECT t.id, t.caja, t.nro_remesa, t.precinto, t.monto, t.retirada, t.retirada_por, t.fecha, t.turno
          FROM remesas_trns t
          WHERE t.retirada='No'
            AND t.caja=%s
            AND t.local=%s
            AND DATE(t.fecha)=%s
            {extra_sql}
          ORDER BY t.id ASC
        """
        cur.execute(sql, tuple(params))
        return jsonify(cur.fetchall())
    except Exception as e:
        print("❌ remesas_no_retiradas:", e)
        return jsonify([])
    finally:
        try: cur.close()
        except: ...
        conn.close()


@app.route('/remesas_hoy')
@login_required
@with_read_scope('t')
def remesas_hoy():
    caja  = request.args.get("caja")
    local = get_local_param()
    fecha = request.args.get("fecha")
    turno = request.args.get("turno")
    if not (caja and local and fecha and turno):
        return jsonify([])

    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        sql = f"""
          SELECT t.id, t.fecha, t.nro_remesa, t.precinto, t.monto, t.retirada, t.retirada_por, t.ult_mod, t.turno, t.local, t.caja
          FROM remesas_trns t
          WHERE t.caja=%s AND DATE(t.fecha)=%s AND t.local=%s AND t.turno=%s
            AND t.retirada='Sí'
            {g.read_scope}   -- L2: cajas cerradas (match tolerante). L3: locales cerrados.
          ORDER BY t.id ASC
        """
        cur.execute(sql, (caja, _normalize_fecha(fecha), local, turno))
        return jsonify(cur.fetchall())
    finally:
        cur.close(); conn.close()


# ===============================
# REMESAS – ALTAS (POST)
# ===============================

@app.route('/guardar_remesas_lote', methods=['POST'])
@login_required
@require_edit_ctx  # valida can_edit con (local,caja,fecha,turno) del body + session
def remesas_guardar_lote():
    """
    Body:
    {
      "caja": "Caja 1",
      "fecha": "2025-07-21",
      "turno": "UNI",
      "remesas": [
        {
          "nro_remesa": "...",
          "precinto": "...",
          "monto": "12.345,67" | 12345.67,
          "retirada": "Sí"|"No",
          "retirada_por": "Nombre",
          "fecha_retirada": "YYYY-MM-DD"   # opcional
        }, ...
      ]
    }
    """
    data    = request.get_json() or {}
    remesas = data.get('remesas') or []
    if not remesas:
        return jsonify(success=False, msg="No se recibieron remesas."), 400

    ctx     = g.ctx   # ← lo armó require_edit_ctx
    local   = ctx['local']
    caja    = ctx['caja']
    fecha   = _normalize_fecha(ctx['fecha'])
    turno   = ctx['turno']
    usuario = session.get('username')

    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        sql = """
          INSERT INTO remesas_trns
          (usuario, local, caja, turno, fecha, nro_remesa, precinto, monto, retirada, retirada_por, fecha_retirada, ult_mod, estado)
          VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),'revision')
        """
        inserted = 0
        for t in remesas:
            nro_remesa   = (t.get('nro_remesa') or "").strip()
            precinto     = (t.get('precinto') or "").strip()

            # monto saneado
            monto_str = str(t.get('monto', '0')).replace('.', '').replace(',', '.')
            try:   monto = float(monto_str or 0)
            except: monto = 0.0

            retirada     = (t.get('retirada') or "No").strip()
            retirada_por = (t.get('retirada_por') or "").strip()
            f_ret        = t.get('fecha_retirada')
            fecha_ret    = _normalize_fecha(f_ret) if f_ret else None

            cur.execute(sql, (usuario, local, caja, turno, fecha,
                              nro_remesa, precinto, monto, retirada, retirada_por, fecha_ret))
            inserted += cur.rowcount

        conn.commit()
        return jsonify(success=True, inserted=inserted, msg="Remesas guardadas correctamente.")
    except Exception as e:
        conn.rollback()
        print("❌ ERROR al guardar remesas:", e)
        return jsonify(success=False, msg=str(e)), 500
    finally:
        try: cur.close()
        except: ...
        conn.close()


# ===============================
# REMESAS – UPDATE puntuales
# ===============================

@app.route("/actualizar_retirada", methods=["POST"])
@login_required
def remesas_actualizar_retirada():
    """
    Body: { "id": 123, "retirada": "Sí"|"No", "retirada_por": "...", "fecha_retirada": "YYYY-MM-DD" (opcional) }

    Permisos especiales: Este endpoint permite editar 'retirada_por' y 'fecha_retirada'
    incluso con caja cerrada para nivel 1 (cajero), siempre que el local esté abierto.
    """
    try:
        data = request.get_json() or {}
        print(f"📨 actualizar_retirada recibió: {data}")

        remesa_id        = data.get('id')
        nueva_retirada   = data.get('retirada')
        retirada_por     = (data.get('retirada_por') or '').strip()
        fecha_retirada_s = data.get('fecha_retirada')

        if not remesa_id:
            return jsonify(success=False, msg="Falta id"), 400

        # Validar nueva_retirada si se proporciona
        if nueva_retirada and nueva_retirada not in ('Sí', 'No'):
            return jsonify(success=False, msg=f"Valor de 'retirada' inválido: {nueva_retirada}"), 400

        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)

        cur.execute("SELECT id, local, caja, fecha, turno, retirada FROM remesas_trns WHERE id=%s", (remesa_id,))
        row = cur.fetchone()
        if not row:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="No existe la remesa"), 404

        # Si no se proporciona nueva_retirada, usar la actual
        if not nueva_retirada:
            nueva_retirada = row['retirada']

        # Usar permisos especiales para editar campos de retirada
        lvl = get_user_level()
        f_norm = _normalize_fecha(row['fecha'])

        # Debug: verificar estados
        is_auditado = is_local_auditado(conn, row['local'], f_norm)
        is_cerrado = is_local_closed(conn, row['local'], f_norm)

        print(f"🔍 Verificando permisos: nivel={lvl}, auditado={is_auditado}, cerrado={is_cerrado}, local={row['local']}, fecha={f_norm}")

        if not can_edit_remesa_retirada(conn, row['local'], row['caja'], row['turno'], row['fecha'], lvl):
            msg = f"No permitido (nivel={lvl}, auditado={is_auditado}, cerrado={is_cerrado})"
            print(f"❌ Permisos rechazados: {msg}")
            cur.close()
            conn.close()
            return jsonify(success=False, msg=msg), 403

        fecha_retirada = _normalize_fecha(fecha_retirada_s) if fecha_retirada_s else None

        cur2 = conn.cursor()
        cur2.execute("""
            UPDATE remesas_trns
               SET retirada=%s,
                   retirada_por=%s,
                   fecha_retirada = COALESCE(%s, fecha_retirada),
                   ult_mod=NOW()
             WHERE id=%s
        """, (nueva_retirada, retirada_por, fecha_retirada, remesa_id))
        conn.commit()

        print(f"✅ Remesa {remesa_id} actualizada correctamente")

        cur.close()
        cur2.close()
        conn.close()
        return jsonify(success=True, updated=cur2.rowcount)

    except Exception as e:
        print(f"❌ ERROR actualizar_retirada: {e}")
        import traceback
        traceback.print_exc()
        try:
            conn.rollback()
            conn.close()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


@app.route('/remesas/<int:remesa_id>', methods=['PUT'])
@login_required
def remesas_update(remesa_id):
    """
    Body (opcional por campo): { nro_remesa, precinto, monto, retirada, retirada_por, fecha_retirada }
    """
    data = request.get_json() or {}

    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        # 1. Obtener registro anterior para auditoría
        datos_anteriores = obtener_registro_anterior(conn, 'remesas_trns', remesa_id)

        # contexto real → para can_edit
        cur.execute("SELECT id, local, caja, fecha, turno FROM remesas_trns WHERE id=%s", (remesa_id,))
        row = cur.fetchone()
        if not row:
            return jsonify(success=False, msg="No existe la remesa"), 404

        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], get_user_level()):
            return jsonify(success=False, msg="No permitido (caja/local cerrados para tu rol)"), 409

        sets, params = [], []
        datos_nuevos = {}

        def add(col, val):
            sets.append(f"{col}=%s")
            params.append(val)
            datos_nuevos[col] = val

        if 'nro_remesa' in data:
            add("nro_remesa", (data.get('nro_remesa') or "").strip())
        if 'precinto' in data:
            add("precinto", (data.get('precinto') or "").strip())
        if 'monto' in data:
            m = str(data.get('monto', '0')).replace('.', '').replace(',', '.')
            try: add("monto", float(m or 0))
            except: return jsonify(success=False, msg="Monto inválido"), 400
        if 'retirada' in data:
            r = (data.get('retirada') or "").strip()
            if r not in ('Sí', 'No'): return jsonify(success=False, msg="retirada inválida"), 400
            add("retirada", r)
        if 'retirada_por' in data:
            add("retirada_por", (data.get('retirada_por') or "").strip())
        if 'fecha_retirada' in data:
            f = data.get('fecha_retirada')
            add("fecha_retirada", _normalize_fecha(f) if f else None)

        if not sets:
            return jsonify(success=False, msg="Sin cambios"), 400

        sets.append("ult_mod=NOW()")
        sql = f"UPDATE remesas_trns SET {', '.join(sets)} WHERE id=%s"
        params.append(remesa_id)

        cur2 = conn.cursor()
        cur2.execute(sql, tuple(params))
        conn.commit()

        # 2. Registrar auditoría
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='remesas_trns',
            registro_id=remesa_id,
            datos_anteriores=datos_anteriores,
            datos_nuevos=datos_nuevos,
            contexto_override={'local': row['local'], 'caja': row['caja'], 'fecha_operacion': row['fecha'], 'turno': row['turno']}
        )

        return jsonify(success=True, updated=cur2.rowcount)
    except Exception as e:
        conn.rollback()
        print("❌ remesas_update:", e)
        return jsonify(success=False, msg=str(e)), 500
    finally:
        try: cur.close()
        except: ...
        conn.close()


@app.route('/remesas/<int:remesa_id>', methods=['DELETE'])
@login_required
def remesas_delete(remesa_id):
    """
    Eliminar remesa:
    - Nivel 1 (cajero): Puede eliminar solo si es su caja Y está abierta
    - Nivel 2 (encargado): Puede eliminar mientras local esté abierto
    - Nivel 3 (auditor): Puede eliminar mientras local NO esté auditado
    """
    lvl = get_user_level()

    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        # 1. Obtener registro anterior para auditoría
        datos_anteriores = obtener_registro_anterior(conn, 'remesas_trns', remesa_id)

        cur.execute("SELECT id, local, caja, fecha, turno FROM remesas_trns WHERE id=%s", (remesa_id,))
        row = cur.fetchone()
        if not row:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="No existe la remesa"), 404

        # Verificar permisos según nivel
        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], lvl):
            cur.close()
            conn.close()
            return jsonify(success=False, msg="No permitido (local cerrado/auditado para tu rol)"), 403

        cur2 = conn.cursor()
        cur2.execute("DELETE FROM remesas_trns WHERE id=%s", (remesa_id,))
        conn.commit()

        # 2. Registrar auditoría
        registrar_auditoria(
            conn=conn,
            accion='DELETE',
            tabla='remesas_trns',
            registro_id=remesa_id,
            datos_anteriores=datos_anteriores,
            contexto_override={'local': row['local'], 'caja': row['caja'], 'fecha_operacion': row['fecha'], 'turno': row['turno']}
        )

        cur.close()
        cur2.close()
        conn.close()
        return jsonify(success=True, deleted=cur2.rowcount)
    except Exception as e:
        conn.rollback()
        print("❌ remesas_delete:", e)
        try:
            cur.close()
            conn.close()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


    # _________________________________________ tarjetas ____________________-
# =========================
#  TARJETAS (helpers + API)
# =========================
# ----------------------------------------------
# TARJETAS (con TURNO + USUARIO en todas las ops)
# ----------------------------------------------
import json
from flask import request, jsonify

def _parse_float(value):
    """
    Reemplazo seguro para json.loads(..., parse_float=_parse_float)
    Limpia comas y puntos según formato argentino antes de convertir.
    """
    try:
        if value is None or str(value).strip() == "":
            return 0.0
        s = str(value).strip()

        # Elimina espacios y símbolos monetarios
        s = s.replace("$", "").replace(" ", "")

        # Si tiene formato argentino (1.234,56) -> 1234.56
        if "," in s and "." in s and s.find(".") < s.find(","):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")  # por si solo usa coma

        return float(s)
    except Exception:
        return 0.0

# Lista completa de tarjetas que siempre se deben insertar
TODAS_LAS_TARJETAS = [
    "VISA",
    "VISA DÉBITO",
    "VISA PREPAGO",
    "MASTERCARD",
    "MASTERCARD DÉBITO",
    "MASTERCARD PREPAGO",
    "CABAL",
    "CABAL DÉBITO",
    "AMEX",
    "MAESTRO",
    "NARANJA",
    "MAS DELIVERY",
    "DINERS",
    "PAGOS INMEDIATOS"
]

def _ensure_full_brand_set(conn, local, caja, turno, fecha, terminal, lote, usuario, tarjetas_cargadas):
    """
    Garantiza que TODAS las tarjetas del conjunto estándar existan para el lote.
    Si una tarjeta no vino en tarjetas_cargadas, se inserta con monto=0 y monto_tip=0.
    """
    cursor = conn.cursor()

    sql_upsert = """
        INSERT INTO tarjetas_trns
        (usuario, local, caja, turno, tarjeta, terminal, lote, monto, monto_tip, fecha, estado)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'revision')
        ON DUPLICATE KEY UPDATE
          usuario=VALUES(usuario),
          estado='revision'
    """

    # Crear set de tarjetas ya cargadas
    tarjetas_existentes = {t.upper() for t in tarjetas_cargadas}

    # Insertar las que faltan en 0
    for tarjeta in TODAS_LAS_TARJETAS:
        if tarjeta.upper() not in tarjetas_existentes:
            cursor.execute(sql_upsert, (
                usuario, local, caja, turno, tarjeta, terminal, lote, 0.0, 0.0, fecha
            ))

    cursor.close()


@app.route('/guardar_tarjetas_lote', methods=['POST'])
@login_required
@require_edit_ctx
def guardar_tarjetas_lote():
    data     = request.get_json() or {}
    tarjetas = data.get('tarjetas', [])
    if not tarjetas:
        return jsonify(success=False, msg="No se recibieron tarjetas"), 400

    ctx     = g.ctx
    local   = ctx['local']; caja = ctx['caja']; fecha = ctx['fecha']; turno = ctx['turno']
    usuario = session.get('username')

    # agrupamos por (terminal,lote) para luego completar el set de marcas
    grupos: Dict[Tuple[str,str], List[dict]] = {}
    for t in tarjetas:
        terminal = (t.get('terminal') or "").strip()
        lote     = (t.get('lote') or "").strip()
        if not terminal or not lote:
            # si falta, no procesamos esta fila
            continue
        grupos.setdefault((terminal, lote), []).append(t)

    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)

        # VALIDACIÓN: Verificar que ningún lote ya exista para este contexto
        for (terminal, lote) in grupos.keys():
            cur.execute("""
                SELECT COUNT(*) as cnt
                FROM tarjetas_trns
                WHERE local=%s AND caja=%s AND DATE(fecha)=%s AND turno=%s
                  AND terminal=%s AND lote=%s
            """, (local, caja, fecha, turno, terminal, lote))
            resultado = cur.fetchone()
            if resultado and resultado['cnt'] > 0:
                cur.close(); conn.close()
                return jsonify(
                    success=False,
                    msg=f"El lote '{lote}' para la terminal '{terminal}' ya existe en esta caja/fecha/turno. No se puede duplicar."
                ), 409

        # Si pasó la validación, procedemos con el insert
        sql_insert = """
            INSERT INTO tarjetas_trns
            (usuario, local, caja, turno, tarjeta, terminal, lote, monto, monto_tip, fecha, estado)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'revision')
        """

        inserted = 0
        for (terminal, lote), filas in grupos.items():
            tarjetas_cargadas = []

            # Primero: insert las que vinieron con datos
            for t in filas:
                tarjeta   = (t.get('tarjeta') or "").strip()
                if not tarjeta:
                    continue
                tarjetas_cargadas.append(tarjeta)
                monto     = _parse_float(t.get('monto', 0))
                monto_tip = _parse_float(t.get('tip', 0))
                cur.execute(sql_insert, (
                    usuario, local, caja, turno, tarjeta, terminal, lote, monto, monto_tip, fecha
                ))
                inserted += cur.rowcount

            # Luego: garantizar set completo de marcas con 0 (las que faltan)
            _ensure_full_brand_set(
                conn, local=local, caja=caja, turno=turno, fecha=fecha,
                terminal=terminal, lote=lote, usuario=usuario,
                tarjetas_cargadas=tarjetas_cargadas
            )

        conn.commit()
        cur.close(); conn.close()
        return jsonify(success=True, msg="Tarjetas guardadas correctamente.", inserted=inserted)
    except Exception as e:
        print("❌ ERROR al guardar tarjetas:", e)
        return jsonify(success=False, msg=f"Error al guardar tarjetas: {e}"), 500

# ------------------------------------------------------------------------------------
#  Listado de tarjetas cargadas del día (única tabla: incluye monto_tip)
#  Protegido por with_read_scope (L2/L3 ven según cierre de caja/local)
# ------------------------------------------------------------------------------------
@app.route("/tarjetas_cargadas_hoy")
@login_required
@with_read_scope('t')
def tarjetas_cargadas_hoy():
    caja      = request.args.get("caja")
    fecha_str = request.args.get("fecha")
    turno     = request.args.get("turno")
    local     = get_local_param()

    if not (caja and fecha_str and turno and local):
        return jsonify([])

    fecha = _normalize_fecha(fecha_str)
    if not fecha:
        return jsonify([])

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        sql = f"""
            SELECT t.id, t.tarjeta, t.terminal, t.lote, t.monto, t.monto_tip, t.estado
              FROM tarjetas_trns t
             WHERE t.local=%s AND t.caja=%s AND DATE(t.fecha)=%s AND t.turno=%s
               {g.read_scope}
             ORDER BY t.terminal, t.lote, t.tarjeta, t.id
        """
        cursor.execute(sql, (local, caja, fecha, turno))
        resultados = cursor.fetchall()
        cursor.close(); conn.close()
        return jsonify(resultados)
    except Exception as e:
        print("❌ ERROR al consultar tarjetas cargadas:", e)
        return jsonify([])

# ------------------------------------------------------------------------------------
#  Update puntual de una tarjeta (monto y/o monto_tip)
#  can_edit por fila
# ------------------------------------------------------------------------------------
@app.route("/tarjetas/<int:tarjeta_id>", methods=["PUT"])
@login_required
def actualizar_tarjeta(tarjeta_id: int):
    data = request.get_json() or {}
    monto_raw   = data.get("monto")
    tip_raw     = data.get("monto_tip")
    estado_raw  = data.get("estado")  # opcional, por si luego querés marcar 'ok'/'cargado'

    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)

        # 1. Obtener registro anterior para auditoría
        datos_anteriores = obtener_registro_anterior(conn, 'tarjetas_trns', tarjeta_id)

        cur.execute("SELECT id, local, caja, fecha, turno FROM tarjetas_trns WHERE id=%s", (tarjeta_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify(success=False, msg="Registro no encontrado"), 404

        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], get_user_level()):
            cur.close(); conn.close()
            return jsonify(success=False, msg="No permitido"), 409

        sets = []
        vals = []
        datos_nuevos = {}
        if monto_raw is not None:
            monto = _parse_float(monto_raw)
            sets.append("monto=%s"); vals.append(monto)
            datos_nuevos['monto'] = monto
        if tip_raw is not None:
            monto_tip = _parse_float(tip_raw)
            sets.append("monto_tip=%s"); vals.append(monto_tip)
            datos_nuevos['monto_tip'] = monto_tip
        if estado_raw is not None:
            sets.append("estado=%s"); vals.append(str(estado_raw))
            datos_nuevos['estado'] = str(estado_raw)

        if not sets:
            cur.close(); conn.close()
            return jsonify(success=False, msg="Sin cambios"), 400

        vals.append(tarjeta_id)
        cur2 = conn.cursor()
        cur2.execute(f"UPDATE tarjetas_trns SET {', '.join(sets)} WHERE id=%s", tuple(vals))
        conn.commit()

        # 2. Registrar auditoría
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='tarjetas_trns',
            registro_id=tarjeta_id,
            datos_anteriores=datos_anteriores,
            datos_nuevos=datos_nuevos,
            contexto_override={'local': row['local'], 'caja': row['caja'], 'fecha_operacion': row['fecha'], 'turno': row['turno']}
        )

        cur2.close(); cur.close(); conn.close()
        return jsonify(success=True)
    except Exception as e:
        print("❌ ERROR al actualizar tarjeta:", e)
        return jsonify(success=False, msg=f"Error al actualizar tarjeta: {e}"), 500

# ------------------------------------------------------------------------------------
#  Delete en bloque por id (borra todo el grupo terminal/lote del día/turno/caja/local)
#  can_edit por fila
# ------------------------------------------------------------------------------------
@app.route("/tarjetas/<int:tarjeta_id>", methods=["DELETE"])
@login_required
def borrar_tarjeta(tarjeta_id: int):
    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)

        # 1. Obtener todos los registros del grupo para auditoría
        cur.execute("""
            SELECT id, local, caja, fecha, turno, terminal, lote
              FROM tarjetas_trns
             WHERE id=%s
        """, (tarjeta_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify(success=False, msg="Registro no encontrado"), 404

        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], get_user_level()):
            cur.close(); conn.close()
            return jsonify(success=False, msg="No permitido"), 409

        # 2. Obtener todos los registros del lote que se van a eliminar
        cur.execute("""
            SELECT * FROM tarjetas_trns
             WHERE local=%s AND caja=%s AND DATE(fecha)=%s AND turno=%s
               AND terminal=%s AND lote=%s
        """, (row['local'], row['caja'], _normalize_fecha(row['fecha']), row['turno'], row['terminal'], row['lote']))
        registros_a_eliminar = cur.fetchall()

        cur2 = conn.cursor()
        cur2.execute("""
            DELETE FROM tarjetas_trns
             WHERE local=%s AND caja=%s AND DATE(fecha)=%s AND turno=%s
               AND terminal=%s AND lote=%s
        """, (row['local'], row['caja'], _normalize_fecha(row['fecha']), row['turno'], row['terminal'], row['lote']))
        eliminadas = cur2.rowcount
        conn.commit()

        # 3. Registrar auditoría para cada registro eliminado
        for reg in registros_a_eliminar:
            registrar_auditoria(
                conn=conn,
                accion='DELETE',
                tabla='tarjetas_trns',
                registro_id=reg['id'],
                datos_anteriores=dict(reg),
                descripcion=f"Eliminación de lote completo: {row['terminal']} / {row['lote']}",
                contexto_override={'local': row['local'], 'caja': row['caja'], 'fecha_operacion': row['fecha'], 'turno': row['turno']}
            )

        cur2.close(); cur.close(); conn.close()
        return jsonify(success=True, deleted=eliminadas)
    except Exception as e:
        print("❌ ERROR al borrar tarjetas (bloque):", e)
        return jsonify(success=False, msg=f"Error al borrar: {e}"), 500


# ------------------------------------------------------------------------------------
#  LOTES AUDITADOS - Solo para auditores (nivel 3)
# ------------------------------------------------------------------------------------
@app.route("/lotes_auditados", methods=["GET"])
@login_required
def obtener_lotes_auditados():
    """
    Obtiene los lotes auditados para una fecha/caja/turno/local específico.
    Query params: local, caja, fecha, turno
    """
    local = request.args.get("local")
    caja  = request.args.get("caja")
    fecha = request.args.get("fecha")
    turno = request.args.get("turno")

    if not all([local, caja, fecha, turno]):
        return jsonify(success=False, msg="Faltan parámetros"), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT terminal, lote, auditado, auditado_por, fecha_auditoria
            FROM lotes_auditados
            WHERE local=%s AND caja=%s AND DATE(fecha)=%s AND turno=%s
        """, (local, caja, fecha, turno))
        resultados = cur.fetchall()
        cur.close()
        conn.close()

        # Convertir a dict para fácil acceso desde JS
        lotes_dict = {}
        for r in resultados:
            key = f"{r['terminal']}||{r['lote']}"
            lotes_dict[key] = {
                'auditado': bool(r['auditado']),
                'auditado_por': r['auditado_por'],
                'fecha_auditoria': str(r['fecha_auditoria']) if r['fecha_auditoria'] else None
            }

        return jsonify(success=True, lotes=lotes_dict)
    except Exception as e:
        print("❌ ERROR al obtener lotes auditados:", e)
        return jsonify(success=False, msg=str(e)), 500


@app.route("/lotes_auditados/marcar", methods=["POST"])
@login_required
def marcar_lote_auditado():
    """
    Marca o desmarca un lote como auditado.
    Body: { local, caja, fecha, turno, terminal, lote, auditado: true/false }
    Solo para nivel 3 (auditores)
    """
    if get_user_level() < 3:
        return jsonify(success=False, msg="Solo auditores pueden marcar lotes"), 403

    data = request.get_json() or {}
    local    = data.get("local")
    caja     = data.get("caja")
    fecha    = data.get("fecha")
    turno    = data.get("turno")
    terminal = data.get("terminal")
    lote     = data.get("lote")
    auditado = data.get("auditado", True)
    usuario  = session.get("username")

    if not all([local, caja, fecha, turno, terminal, lote]):
        return jsonify(success=False, msg="Faltan parámetros"), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        if auditado:
            # Marcar como auditado
            cur.execute("""
                INSERT INTO lotes_auditados
                (local, caja, fecha, turno, terminal, lote, auditado, auditado_por, fecha_auditoria)
                VALUES (%s, %s, %s, %s, %s, %s, TRUE, %s, NOW())
                ON DUPLICATE KEY UPDATE
                  auditado=TRUE,
                  auditado_por=%s,
                  fecha_auditoria=NOW()
            """, (local, caja, fecha, turno, terminal, lote, usuario, usuario))
        else:
            # Desmarcar
            cur.execute("""
                INSERT INTO lotes_auditados
                (local, caja, fecha, turno, terminal, lote, auditado, auditado_por, fecha_auditoria)
                VALUES (%s, %s, %s, %s, %s, %s, FALSE, NULL, NULL)
                ON DUPLICATE KEY UPDATE
                  auditado=FALSE,
                  auditado_por=NULL,
                  fecha_auditoria=NULL
            """, (local, caja, fecha, turno, terminal, lote))

        conn.commit()
        cur.close()
        conn.close()
        return jsonify(success=True)
    except Exception as e:
        print("❌ ERROR al marcar lote auditado:", e)
        return jsonify(success=False, msg=str(e)), 500




    #____________________________________RAPPI_______________________________

# =================== R A P P I  (API con TURNOS) ===================
# ===== Rappi =====

@app.route('/guardar_rappi_lote', methods=['POST'])
@login_required
@require_edit_ctx  # usa local/caja/fecha/turno del body y valida can_edit
def guardar_rappi_lote():
    data = request.get_json() or {}
    transacciones = data.get('transacciones', []) or []

    if not transacciones:
        return jsonify(success=False, msg="No se recibieron transacciones"), 400

    # Contexto validado por el decorador
    ctx     = g.ctx
    local   = ctx['local']
    caja    = ctx['caja']
    fecha   = _normalize_fecha(ctx['fecha'])
    turno   = ctx['turno']
    usuario = session.get('username') or 'sistema'

    try:
        conn = get_db_connection()
        cur  = conn.cursor()

        sql = """
            INSERT INTO rappi_trns (usuario, caja, transaccion, monto, fecha, local, turno, estado)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'revision')
        """

        inserted = 0
        for t in transacciones:
            transaccion = (t.get('transaccion') or "").strip()
            m = t.get('monto', 0)
            try:
                if isinstance(m, str):
                    m = float(m.replace('.', '').replace(',', '.'))
                monto = float(m or 0)
            except Exception:
                monto = 0.0

            cur.execute(sql, (usuario, caja, transaccion, monto, fecha, local, turno))
            inserted += cur.rowcount

        conn.commit()
        cur.close(); conn.close()
        return jsonify(success=True, inserted=inserted, msg="Rappi guardado correctamente.")
    except Exception as e:
        print("❌ ERROR guardar_rappi_lote:", e)
        return jsonify(success=False, msg=str(e)), 500


# ===============================
# Rappi – LECTURA (GET)
# ===============================
@app.route('/rappi_cargadas')
@login_required
@with_read_scope('t')  # agrega g.read_scope acorde al nivel (L2: cajas cerradas; L3: locales cerrados)
def rappi_cargadas():
    local = get_local_param()
    caja  = request.args.get('caja')
    fecha = request.args.get('fecha')
    turno = request.args.get('turno')

    if not (local and caja and fecha and turno):
        return jsonify(success=True, datos=[])

    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)
        sql = f"""
            SELECT t.id, t.transaccion, t.monto, t.fecha, t.turno
              FROM rappi_trns t
             WHERE t.local=%s
               AND t.caja=%s
               AND t.turno=%s
               AND DATE(t.fecha)=%s
               {g.read_scope}   -- L2/L3 pueden ver aun con cierres
             ORDER BY t.id ASC
        """
        cur.execute(sql, (local, caja, turno, _normalize_fecha(fecha)))
        datos = cur.fetchall()
        cur.close(); conn.close()
        return jsonify(success=True, datos=datos)
    except Exception as e:
        print("❌ ERROR rappi_cargadas:", e)
        return jsonify(success=False, msg=str(e)), 500


# ===============================
# Rappi – UPDATE (PUT)
# ===============================
@app.route('/rappi/<int:rappi_id>', methods=['PUT'])
@login_required
def actualizar_rappi(rappi_id):
    data = request.get_json() or {}
    transaccion = (data.get('transaccion') or "").strip()

    try:
        imp = data.get('monto', 0)
        if isinstance(imp, str):
            imp = float(imp.replace('.', '').replace(',', '.'))
        importe = float(imp or 0)
    except Exception:
        return jsonify(success=False, msg="Monto inválido"), 400

    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)

        # 1. Obtener registro anterior para auditoría
        datos_anteriores = obtener_registro_anterior(conn, 'rappi_trns', rappi_id)

        cur.execute("SELECT id, local, caja, fecha, turno FROM rappi_trns WHERE id=%s", (rappi_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify(success=False, msg="Registro no encontrado"), 404

        # Permisos por nivel/estado (L1/L2/L3)
        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], get_user_level()):
            cur.close(); conn.close()
            return jsonify(success=False, msg="No permitido (caja/local cerrados para tu rol)"), 409

        cur2 = conn.cursor()
        cur2.execute("""
            UPDATE rappi_trns
               SET transaccion=%s, monto=%s
             WHERE id=%s
        """, (transaccion, importe, rappi_id))
        conn.commit()

        # 2. Registrar auditoría
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='rappi_trns',
            registro_id=rappi_id,
            datos_anteriores=datos_anteriores,
            datos_nuevos={'transaccion': transaccion, 'monto': importe},
            contexto_override={'local': row['local'], 'caja': row['caja'], 'fecha_operacion': row['fecha'], 'turno': row['turno']}
        )

        cur2.close(); cur.close(); conn.close()
        return jsonify(success=True)
    except Exception as e:
        print("❌ ERROR actualizar_rappi:", e)
        return jsonify(success=False, msg=str(e)), 500


# ===============================
# Rappi – DELETE
# ===============================
@app.route('/rappi/<int:rappi_id>', methods=['DELETE'])
@login_required
def borrar_rappi(rappi_id):
    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)

        # 1. Obtener registro anterior para auditoría
        datos_anteriores = obtener_registro_anterior(conn, 'rappi_trns', rappi_id)

        cur.execute("SELECT id, local, caja, fecha, turno FROM rappi_trns WHERE id=%s", (rappi_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify(success=False, msg="Registro no encontrado"), 404

        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], get_user_level()):
            cur.close(); conn.close()
            return jsonify(success=False, msg="No permitido (caja/local cerrados para tu rol)"), 409

        cur2 = conn.cursor()
        cur2.execute("DELETE FROM rappi_trns WHERE id=%s", (rappi_id,))
        conn.commit()
        deleted = cur2.rowcount

        # 2. Registrar auditoría
        registrar_auditoria(
            conn=conn,
            accion='DELETE',
            tabla='rappi_trns',
            registro_id=rappi_id,
            datos_anteriores=datos_anteriores,
            contexto_override={'local': row['local'], 'caja': row['caja'], 'fecha_operacion': row['fecha'], 'turno': row['turno']}
        )

        cur2.close(); cur.close(); conn.close()
        return jsonify(success=True, deleted=deleted)
    except Exception as e:
        print("❌ ERROR borrar_rappi:", e)
        return jsonify(success=False, msg=str(e)), 500

## __________________________________ PEDIDOS YA _______________________________________
# ---------------------------
#  PEDIDOSYA (API)
# ---------------------------
# ---------------------------
#  PEDIDOSYA (API) con TURNO
# ---------------------------

# Usa tu _normalize_fecha ya definido en tu app.
# Aquí lo invocamos tal como lo tenés en otras secciones.

def is_caja_abierta_turno(conn, local, caja, fecha, turno="UNI"):
    """
    True si la caja/turno está ABIERTA para (local, caja, fecha, turno).
    Si no hay fila en cajas_estado => se considera ABIERTA (True).
    """
    turno = (turno or "UNI").upper()
    f = _normalize_fecha(fecha)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT estado
              FROM cajas_estado
             WHERE local=%s AND caja=%s AND fecha_operacion=%s AND turno=%s
             LIMIT 1
        """, (local, caja, f, turno))
        row = cur.fetchone()
        # abierta si no hay fila o estado=1; cerrada si estado=0
        return (row is None) or (row[0] is None) or (int(row[0]) == 1)
    finally:
        cur.close()

# ===== PedidosYa =====
# Requiere utilidades ya presentes en tu app:
# - get_db_connection()
# - _normalize_fecha()
# - can_edit(conn, local, caja, turno, fecha, user_level) -> bool
# - get_user_level() -> int
# - require_edit_ctx: valida permisos de escritura y expone g.ctx = {local,caja,fecha,turno}
# - with_read_scope(alias): expone g.read_scope para filtrar según rol (L1/L2/L3)
# - login_required

# ===============================
# PedidosYa – CREAR (lote)
# ===============================
@app.route('/guardar_pedidosya_lote', methods=['POST'])
@login_required
@require_edit_ctx  # valida permisos y deja g.ctx con {local,caja,fecha,turno}
def guardar_pedidosya_lote():
    """
    Body:
    {
      "caja": "Caja 1",
      "fecha": "YYYY-MM-DD",
      "turno": "DIA" | "NOCHE" | "UNI",
      "transacciones": [{ "transaccion":"abc", "monto":"1.234,56" }, ...]
    }
    """
    data = request.get_json() or {}
    transacciones = data.get('transacciones', []) or []

    if not transacciones:
        return jsonify(success=False, msg="No se recibieron transacciones"), 400

    ctx     = g.ctx
    local   = ctx['local']
    caja    = ctx['caja']
    fecha   = _normalize_fecha(ctx['fecha'])
    turno   = (ctx['turno'] or 'UNI').upper()
    usuario = session.get('username') or 'sistema'

    try:
        conn = get_db_connection()
        cur  = conn.cursor()

        sql = """
            INSERT INTO pedidosya_trns (usuario, local, caja, turno, transaccion, monto, fecha, estado)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'revision')
        """

        inserted = 0
        for t in transacciones:
            transaccion = (t.get('transaccion') or "").strip()
            m = t.get('monto', 0)
            try:
                if isinstance(m, str):
                    m = float(m.replace('.', '').replace(',', '.'))
                monto = float(m or 0)
            except Exception:
                monto = 0.0

            cur.execute(sql, (usuario, local, caja, turno, transaccion, monto, fecha))
            inserted += cur.rowcount

        conn.commit()
        cur.close(); conn.close()
        return jsonify(success=True, inserted=inserted, msg="PedidosYa guardado correctamente.")
    except Exception as e:
        print("❌ ERROR guardar_pedidosya_lote:", e)
        return jsonify(success=False, msg=str(e)), 500


# ===============================
# PedidosYa – LECTURA (GET)
# ===============================
@app.route('/pedidosya_cargadas')
@login_required
@with_read_scope('t')  # agrega g.read_scope acorde a L1/L2/L3
def pedidosya_cargadas():
    """
    Query params: ?caja=Caja%201&fecha=YYYY-MM-DD&turno=DIA|NOCHE|UNI
    Respuesta: { success:true, datos:[{id, transaccion, monto, fecha, turno}, ...] }
    """
    local = get_local_param()
    caja  = request.args.get('caja')
    fecha = request.args.get('fecha')
    turno = (request.args.get('turno') or 'UNI').upper()

    if not (local and caja and fecha and turno):
        return jsonify(success=True, datos=[])

    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)
        sql = f"""
            SELECT t.id, t.transaccion, t.monto, t.fecha, t.turno
              FROM pedidosya_trns t
             WHERE t.local=%s
               AND t.caja=%s
               AND t.turno=%s
               AND DATE(t.fecha)=%s
               {g.read_scope}   -- L2/L3 pueden ver aun con cierres
             ORDER BY t.id ASC
        """
        cur.execute(sql, (local, caja, turno, _normalize_fecha(fecha)))
        datos = cur.fetchall()
        cur.close(); conn.close()
        return jsonify(success=True, datos=datos)
    except Exception as e:
        print("❌ ERROR pedidosya_cargadas:", e)
        return jsonify(success=False, msg=str(e)), 500


# ===============================
# PedidosYa – UPDATE (PUT)
# ===============================
@app.route('/pedidosya/<int:py_id>', methods=['PUT'])
@login_required
def actualizar_pedidosya(py_id):
    data = request.get_json() or {}
    transaccion = (data.get('transaccion') or "").strip()

    try:
        imp = data.get('monto', 0)
        if isinstance(imp, str):
            imp = float(imp.replace('.', '').replace(',', '.'))
        importe = float(imp or 0)
    except Exception:
        return jsonify(success=False, msg="Monto inválido"), 400

    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)

        # 1. Obtener registro anterior para auditoría
        datos_anteriores = obtener_registro_anterior(conn, 'pedidosya_trns', py_id)

        cur.execute("SELECT id, local, caja, fecha, turno FROM pedidosya_trns WHERE id=%s", (py_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify(success=False, msg="Registro no encontrado"), 404

        # Permisos por nivel/estado (L1/L2/L3)
        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], get_user_level()):
            cur.close(); conn.close()
            return jsonify(success=False, msg="No permitido (caja/local cerrados para tu rol)"), 409

        cur2 = conn.cursor()
        cur2.execute("""
            UPDATE pedidosya_trns
               SET transaccion=%s, monto=%s
             WHERE id=%s
        """, (transaccion, importe, py_id))
        conn.commit()

        # 2. Registrar auditoría
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='pedidosya_trns',
            registro_id=py_id,
            datos_anteriores=datos_anteriores,
            datos_nuevos={'transaccion': transaccion, 'monto': importe},
            contexto_override={'local': row['local'], 'caja': row['caja'], 'fecha_operacion': row['fecha'], 'turno': row['turno']}
        )

        cur2.close(); cur.close(); conn.close()
        return jsonify(success=True)
    except Exception as e:
        print("❌ ERROR actualizar_pedidosya:", e)
        return jsonify(success=False, msg=str(e)), 500


# ===============================
# PedidosYa – DELETE
# ===============================
@app.route('/pedidosya/<int:py_id>', methods=['DELETE'])
@login_required
def borrar_pedidosya(py_id):
    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)

        # 1. Obtener registro anterior para auditoría
        datos_anteriores = obtener_registro_anterior(conn, 'pedidosya_trns', py_id)

        cur.execute("SELECT id, local, caja, fecha, turno FROM pedidosya_trns WHERE id=%s", (py_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify(success=False, msg="Registro no encontrado"), 404

        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], get_user_level()):
            cur.close(); conn.close()
            return jsonify(success=False, msg="No permitido (caja/local cerrados para tu rol)"), 409

        cur2 = conn.cursor()
        cur2.execute("DELETE FROM pedidosya_trns WHERE id=%s", (py_id,))
        conn.commit()
        deleted = cur2.rowcount

        # 2. Registrar auditoría
        registrar_auditoria(
            conn=conn,
            accion='DELETE',
            tabla='pedidosya_trns',
            registro_id=py_id,
            datos_anteriores=datos_anteriores,
            contexto_override={'local': row['local'], 'caja': row['caja'], 'fecha_operacion': row['fecha'], 'turno': row['turno']}
        )

        cur2.close(); cur.close(); conn.close()
        return jsonify(success=True, deleted=deleted)
    except Exception as e:
        print("❌ ERROR borrar_pedidosya:", e)
        return jsonify(success=False, msg=str(e)), 500


# _________________________________________ MERCADO PAGO _____________________________________
# ---------- MERCADO PAGO ----------


# --- Helpers si hiciera falta (si ya existen globales, podés omitirlos) ---
def _normalize_fecha(fecha):
    if isinstance(fecha, datetime):
        return fecha.date()
    if isinstance(fecha, date):
        return fecha
    if isinstance(fecha, str):
        try:
            return datetime.strptime(fecha, "%Y-%m-%d").date()
        except ValueError:
            return None
    return None

# Versión con TURNO (si ya tenés una global, usá esa):
# def is_caja_abierta(conn, local, caja, fecha, turno):
#     fecha = _normalize_fecha(fecha)
#     cur = conn.cursor()
#     cur.execute("""
#         SELECT estado FROM cajas_estado
#          WHERE local=%s AND caja=%s AND fecha_operacion=%s AND turno=%s
#     """, (local, caja, fecha, turno))
#     row = cur.fetchone()
#     cur.close()
#     return (row is None) or (row[0] is None) or (int(row[0]) == 1)

# --------------------------------------------------------------------
# Guardar lote (con TURNO + USUARIO)
# --------------------------------------------------------------------
# MERCADO PAGO (con rol/estado unificado como en REMESAS)
# --------------------------------------------------------------------
# ===============================
# MERCADOPAGO – ALTAS (POST)
# ===============================
@app.route('/guardar_mercadopago_lote', methods=['POST'])
@login_required
@require_edit_ctx  # valida can_edit con (local,caja,fecha,turno) del body + session
def guardar_mercadopago_lote():
    data      = request.get_json() or {}
    registros = data.get('registros', [])

    if not registros:
        return jsonify(success=False, msg="No se recibieron registros"), 400

    ctx     = g.ctx
    local   = ctx['local']
    caja    = ctx['caja']
    fecha   = _normalize_fecha(ctx['fecha'])
    turno   = ctx['turno']
    usuario = session.get('username')

    try:
        conn = get_db_connection()
        cur  = conn.cursor()
        sql = """
            INSERT INTO mercadopago_trns
                (usuario, local, caja, turno, tipo, terminal, comprobante, importe, fecha, estado)
            VALUES
                (%s, %s, %s, %s, %s, %s, %s, %s, %s,'revision')
        """
        inserted = 0
        for r in registros:
            tipo = (r.get("tipo") or "").upper()
            if tipo not in ("NORMAL", "TIP"):
                tipo = "NORMAL"

            terminal    = (r.get("terminal") or "").strip()
            comprobante = (r.get("comprobante") or "").strip()

            imp = r.get("importe", 0)
            try:
                if isinstance(imp, str):
                    imp = float((imp or "0").replace(".", "").replace(",", "."))
                importe = float(imp or 0)
            except Exception:
                importe = 0.0

            cur.execute(sql, (usuario, local, caja, turno, tipo, terminal, comprobante, importe, fecha))
            inserted += cur.rowcount

        conn.commit()
        cur.close(); conn.close()
        return jsonify(success=True, inserted=inserted, msg="MercadoPago guardado correctamente.")
    except Exception as e:
        print("❌ ERROR guardar_mercadopago_lote:", e)
        return jsonify(success=False, msg=str(e)), 500

# (Si preferís sin require_edit_ctx, tu versión original funciona.
#  Bastaría con reemplazar el decorador por @login_required y dejar el can_edit inline.)

# ===============================
# MERCADOPAGO – LECTURA (GET)
# ===============================
@app.route('/mercadopago_cargadas')
@login_required
@with_read_scope('t')
def mercado_pago_cargadas():
    caja  = request.args.get('caja')
    fecha = request.args.get('fecha')
    turno = request.args.get('turno')
    local = get_local_param()

    if not (caja and fecha and turno and local):
        return jsonify(success=True, datos=[])

    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)
        sql = f"""
            SELECT t.id, t.tipo, t.terminal, t.comprobante, t.importe, t.fecha, t.turno
              FROM mercadopago_trns t
             WHERE t.local=%s
               AND t.caja=%s
               AND t.turno=%s
               AND DATE(t.fecha)=%s
               {g.read_scope}   -- L2: cajas cerradas; L3: locales cerrados
             ORDER BY t.id ASC
        """
        cur.execute(sql, (local, caja, turno, _normalize_fecha(fecha)))
        datos = cur.fetchall()
        cur.close(); conn.close()
        return jsonify(success=True, datos=datos)
    except Exception as e:
        print("❌ ERROR mercado_pago_cargadas:", e)
        return jsonify(success=False, msg=str(e)), 500

# ===============================
# MERCADOPAGO – UPDATE (PUT)
# ===============================
@app.route('/mercadopago/<int:mp_id>', methods=['PUT'])
@login_required
def actualizar_mercadopago(mp_id):
    data = request.get_json() or {}

    tipo = (data.get('tipo') or '').upper()
    if tipo not in ('NORMAL', 'TIP'):
        tipo = 'NORMAL'
    terminal    = (data.get('terminal') or '').strip()
    comprobante = (data.get('comprobante') or '').strip()

    try:
        imp = data.get('importe')
        if isinstance(imp, str):
            imp = float(imp.replace('.', '').replace(',', '.'))
        importe = float(imp)
    except Exception:
        return jsonify(success=False, msg="Importe inválido"), 400

    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)
        cur.execute("SELECT id, local, caja, fecha, turno FROM mercadopago_trns WHERE id=%s", (mp_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify(success=False, msg="Registro no encontrado"), 404

        # Permiso por rol/estado (igual que Remesas)
        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], get_user_level()):
            cur.close(); conn.close()
            return jsonify(success=False, msg="No permitido (caja/local cerrados para tu rol)"), 409

        # 1. Obtener registro anterior para auditoría
        datos_anteriores = obtener_registro_anterior(conn, 'mercadopago_trns', mp_id)

        cur2 = conn.cursor()
        cur2.execute("""
            UPDATE mercadopago_trns
               SET tipo=%s, terminal=%s, comprobante=%s, importe=%s
             WHERE id=%s
        """, (tipo, terminal, comprobante, importe, mp_id))
        conn.commit()

        # 2. Registrar auditoría
        datos_nuevos = {
            'tipo': tipo,
            'terminal': terminal,
            'comprobante': comprobante,
            'importe': importe
        }
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='mercadopago_trns',
            registro_id=mp_id,
            datos_anteriores=datos_anteriores,
            datos_nuevos=datos_nuevos,
            contexto_override={
                'local': row['local'],
                'caja': row['caja'],
                'fecha_operacion': row['fecha'],
                'turno': row['turno']
            }
        )

        cur2.close(); cur.close(); conn.close()
        return jsonify(success=True)
    except Exception as e:
        print("❌ ERROR actualizar_mercadopago:", e)
        return jsonify(success=False, msg=str(e)), 500

# ===============================
# MERCADOPAGO – DELETE
# ===============================
@app.route('/mercadopago/<int:mp_id>', methods=['DELETE'])
@login_required
def borrar_mercadopago(mp_id):
    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)
        cur.execute("SELECT id, local, caja, fecha, turno FROM mercadopago_trns WHERE id=%s", (mp_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify(success=False, msg="Registro no encontrado"), 404

        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], get_user_level()):
            cur.close(); conn.close()
            return jsonify(success=False, msg="No permitido (caja/local cerrados para tu rol)"), 409

        # 1. Obtener registro anterior para auditoría
        datos_anteriores = obtener_registro_anterior(conn, 'mercadopago_trns', mp_id)

        cur2 = conn.cursor()
        cur2.execute("DELETE FROM mercadopago_trns WHERE id=%s", (mp_id,))
        conn.commit()
        deleted = cur2.rowcount

        # 2. Registrar auditoría
        registrar_auditoria(
            conn=conn,
            accion='DELETE',
            tabla='mercadopago_trns',
            registro_id=mp_id,
            datos_anteriores=datos_anteriores,
            descripcion=f"Eliminación de MercadoPago: {datos_anteriores.get('terminal', '')} / {datos_anteriores.get('comprobante', '')}",
            contexto_override={
                'local': row['local'],
                'caja': row['caja'],
                'fecha_operacion': row['fecha'],
                'turno': row['turno']
            }
        )

        cur2.close(); cur.close(); conn.close()
        return jsonify(success=True, deleted=deleted)
    except Exception as e:
        print("❌ ERROR borrar_mercadopago:", e)
        return jsonify(success=False, msg=str(e)), 500

# ______________________________________ VENTAS (BASE + Z) ____________________________________________
# Reemplaza tu sección de "index" por esto.
# Quita: /guardar_ventas y TODOS los /ventas_especiales_*



# ===============================
# ANTICIPOS CONSUMIDOS
# ===============================

# ===============================
# Anticipos – CREATE (POST)
# ===============================
@app.route('/guardar_anticipos_lote', methods=['POST'])
@login_required
@require_edit_ctx  # usa local/caja/fecha/turno del body y valida can_edit
def guardar_anticipos_lote():
    """
    DEPRECADO: Endpoint viejo para guardar anticipos.
    El nuevo sistema usa /api/anticipos/consumir

    Retorna error para evitar uso del sistema antiguo.
    """
    return jsonify(
        success=False,
        msg="Este endpoint está deprecado. Use el nuevo sistema de anticipos en /api/anticipos/consumir"
    ), 410  # 410 Gone - recurso ya no disponible


# ===============================
# Anticipos – READ (GET) - DEPRECADO
# ===============================
@app.route('/anticipos_cargados')
@login_required
@with_read_scope('t')  # agrega g.read_scope acorde al nivel (L2: cajas cerradas; L3: locales cerrados)
def anticipos_cargados():
    """
    DEPRECADO: Este endpoint es para compatibilidad con el sistema antiguo.
    El nuevo sistema usa /api/anticipos/consumidos_en_caja
    Retorna datos vacíos para evitar errores.
    """
    # Retornar vacío - el nuevo sistema de anticipos_v2.js no usa este endpoint
    return jsonify(success=True, datos=[])


# ===============================
# Anticipos – UPDATE (PUT) - DEPRECADO
# ===============================
@app.route('/anticipos/<int:anticipo_id>', methods=['PUT'])
@login_required
def actualizar_anticipo(anticipo_id):
    """
    DEPRECADO: Endpoint viejo para actualizar anticipos.
    El nuevo sistema usa /api/anticipos_recibidos/editar/<id>
    """
    return jsonify(
        success=False,
        msg="Este endpoint está deprecado. Use /api/anticipos_recibidos/editar/<id>"
    ), 410  # 410 Gone


# ===============================
# Anticipos – DELETE - DEPRECADO
# ===============================
@app.route('/anticipos/<int:anticipo_id>', methods=['DELETE'])
@login_required
def borrar_anticipo(anticipo_id):
    """
    DEPRECADO: Endpoint viejo para eliminar anticipos.
    El nuevo sistema usa /api/anticipos_recibidos/eliminar/<id> o /api/anticipos/eliminar_de_caja
    """
    return jsonify(
        success=False,
        msg="Este endpoint está deprecado. Use /api/anticipos_recibidos/eliminar/<id> o /api/anticipos/eliminar_de_caja"
    ), 410  # 410 Gone


# ═══════════════════════════════════════════════════════════════════
# NUEVO SISTEMA DE ANTICIPOS RECIBIDOS
# ═══════════════════════════════════════════════════════════════════
# Parte 1: Creación y gestión de anticipos recibidos (admin_anticipos)
# Parte 2: Consumo de anticipos en cajas (cajeros)
# ═══════════════════════════════════════════════════════════════════

# ───────────────────────────────────────────────────────────────────
# PARTE 1: GESTIÓN DE ANTICIPOS RECIBIDOS (admin_anticipos)
# ───────────────────────────────────────────────────────────────────

@app.route('/api/anticipos_recibidos/crear', methods=['POST'])
@login_required
def crear_anticipo_recibido():
    """
    Crear un nuevo anticipo recibido.
    Accesible para:
    - admin_anticipos (nivel 6): puede crear en cualquier local
    - anticipos (nivel 4): solo puede crear en los locales asignados

    Body:
    {
        "fecha_pago": "2025-12-01",
        "fecha_evento": "2025-12-15",
        "importe": 5000.00,
        "cliente": "Juan Pérez",
        "numero_transaccion": "TRX123456",
        "medio_pago": "Transferencia",
        "observaciones": "Reserva para evento...",
        "local": "Ribs Infanta"
    }
    """
    user_level = get_user_level()
    if user_level < 4:
        return jsonify(success=False, msg="No tenés permisos para crear anticipos recibidos"), 403

    data = request.get_json() or {}

    # Validar campos requeridos
    required_fields = ['fecha_pago', 'fecha_evento', 'importe', 'cliente', 'local']
    for field in required_fields:
        if not data.get(field):
            return jsonify(success=False, msg=f"Campo requerido faltante: {field}"), 400

    try:
        # Parsear y validar datos
        fecha_pago = _normalize_fecha(data['fecha_pago'])
        fecha_evento = _normalize_fecha(data['fecha_evento'])
        importe = float(data['importe'])
        cliente = data['cliente'].strip()
        local = data['local'].strip()

        # Validar permiso sobre el local específico
        if not can_user_access_local_for_anticipos(local):
            return jsonify(success=False, msg=f"No tenés permisos para crear anticipos en el local '{local}'"), 403

        # Nuevos campos: divisa y adjunto
        divisa = (data.get('divisa') or 'ARS').strip().upper()
        tipo_cambio_fecha = data.get('tipo_cambio_fecha')
        if not tipo_cambio_fecha:
            tipo_cambio_fecha = fecha_pago
        else:
            tipo_cambio_fecha = _normalize_fecha(tipo_cambio_fecha)

        numero_transaccion = data.get('numero_transaccion', '').strip() or None
        medio_pago = data.get('medio_pago', '').strip() or None
        observaciones = data.get('observaciones', '').strip() or None
        adjunto_gcs_path = data.get('adjunto_gcs_path', '').strip() or None

        if importe <= 0:
            return jsonify(success=False, msg="El importe debe ser mayor a cero"), 400

        # Validar divisa
        divisas_permitidas = ['ARS', 'USD', 'EUR', 'BRL', 'CLP', 'UYU']
        if divisa not in divisas_permitidas:
            return jsonify(success=False, msg=f"Divisa no permitida. Usar: {', '.join(divisas_permitidas)}"), 400

        usuario = session.get('username', 'sistema')

        conn = get_db_connection()
        cur = conn.cursor()

        # Obtener medio_pago_id (nuevo campo)
        medio_pago_id = data.get('medio_pago_id')

        # Si no viene medio_pago_id, asignar "Efectivo" por defecto (retrocompatibilidad)
        if not medio_pago_id:
            cur.execute("SELECT id FROM medios_anticipos WHERE nombre = 'Efectivo' LIMIT 1")
            result = cur.fetchone()
            medio_pago_id = result[0] if result else None

        # Insertar anticipo recibido con divisa y medio_pago_id
        sql = """
            INSERT INTO anticipos_recibidos
            (fecha_pago, fecha_evento, importe, divisa, tipo_cambio_fecha,
             cliente, numero_transaccion, medio_pago, observaciones,
             local, medio_pago_id, estado, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pendiente', %s)
        """
        cur.execute(sql, (fecha_pago, fecha_evento, importe, divisa, tipo_cambio_fecha,
                         cliente, numero_transaccion, medio_pago, observaciones,
                         local, medio_pago_id, usuario))

        anticipo_id = cur.lastrowid

        # Vincular adjunto si existe
        if adjunto_gcs_path:
            try:
                cur.execute("""
                    UPDATE imagenes_adjuntos
                    SET entity_type = 'anticipo_recibido',
                        entity_id = %s
                    WHERE gcs_path = %s
                      AND (entity_id IS NULL OR entity_id = 0)
                """, (anticipo_id, adjunto_gcs_path))
            except Exception as e:
                print(f"⚠️  Warning: No se pudo vincular adjunto: {e}")

        conn.commit()

        # Registrar en auditoría con nuevos campos
        from modules.tabla_auditoria import registrar_auditoria
        registrar_auditoria(
            conn=conn,
            accion='INSERT',
            tabla='anticipos_recibidos',
            registro_id=anticipo_id,
            datos_nuevos={
                'fecha_pago': str(fecha_pago),
                'fecha_evento': str(fecha_evento),
                'importe': importe,
                'divisa': divisa,
                'tipo_cambio_fecha': str(tipo_cambio_fecha),
                'cliente': cliente,
                'local': local,
                'numero_transaccion': numero_transaccion,
                'medio_pago': medio_pago,
                'tiene_adjunto': bool(adjunto_gcs_path)
            },
            descripcion=f"Anticipo recibido creado: {cliente} - {divisa} {importe}"
        )

        cur.close()
        conn.close()

        return jsonify(success=True, msg="Anticipo recibido creado correctamente", anticipo_id=anticipo_id)

    except Exception as e:
        print("❌ ERROR crear_anticipo_recibido:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/anticipos_recibidos/listar', methods=['GET'])
@login_required
def listar_anticipos_recibidos():
    """
    Listar anticipos recibidos según permisos del usuario.
    - auditores (nivel 3): ven todos los anticipos
    - anticipos (nivel 4): solo ve anticipos de sus locales asignados
    - admin_anticipos (nivel 6): ve todos los anticipos

    Params:
    - estado: (opcional) 'pendiente', 'consumido', 'eliminado_global'
    - local: (opcional) filtrar por local
    - fecha_desde, fecha_hasta: (opcional) filtrar por fecha_evento
    """
    user_level = get_user_level()
    if user_level < 3:
        return jsonify(success=False, msg="No tenés permisos para ver anticipos recibidos"), 403

    estado = request.args.get('estado', '').strip()
    local = request.args.get('local', '').strip()
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Construir query con filtros opcionales
        # CORREGIDO: Verificar el estado real consultando anticipos_estados_caja
        sql = """
            SELECT
                ar.id, ar.fecha_pago, ar.fecha_evento, ar.importe, ar.divisa,
                ar.tipo_cambio_fecha, ar.cliente,
                ar.numero_transaccion, ar.medio_pago, ar.observaciones, ar.local,
                ar.estado as estado_global,
                ar.created_by, ar.created_at, ar.updated_by, ar.updated_at,
                ar.deleted_by, ar.deleted_at,
                -- Verificar si fue consumido realmente
                (SELECT COUNT(*) FROM anticipos_estados_caja aec
                 WHERE aec.anticipo_id = ar.id AND aec.estado = 'consumido') as fue_consumido,
                -- Verificar si tiene adjunto
                CASE WHEN EXISTS (
                    SELECT 1 FROM imagenes_adjuntos ia
                    WHERE ia.entity_type = 'anticipo_recibido'
                      AND ia.entity_id = ar.id
                      AND ia.estado = 'active'
                ) THEN 1 ELSE 0 END as tiene_adjunto
            FROM anticipos_recibidos ar
            WHERE 1=1
        """
        params = []

        # Filtrar por locales permitidos si es usuario nivel 4
        if user_level == 4:
            allowed_locales = get_user_allowed_locales()
            if allowed_locales:
                placeholders = ','.join(['%s'] * len(allowed_locales))
                sql += f" AND ar.local IN ({placeholders})"
                params.extend(allowed_locales)
            else:
                # Sin permisos asignados, retornar vacío
                return jsonify(success=True, anticipos=[])

        if estado:
            sql += " AND ar.estado = %s"
            params.append(estado)

        if local:
            sql += " AND ar.local = %s"
            params.append(local)

        if fecha_desde:
            sql += " AND ar.fecha_evento >= %s"
            params.append(_normalize_fecha(fecha_desde))

        if fecha_hasta:
            sql += " AND ar.fecha_evento <= %s"
            params.append(_normalize_fecha(fecha_hasta))

        sql += " ORDER BY ar.fecha_evento DESC, ar.created_at DESC"

        cur.execute(sql, params)
        anticipos_raw = cur.fetchall() or []

        # Corregir el estado basado en la verificación real
        anticipos = []
        for a in anticipos_raw:
            anticipo = dict(a)

            # Convertir fechas a formato string YYYY-MM-DD para JSON
            if anticipo.get('fecha_pago'):
                if hasattr(anticipo['fecha_pago'], 'strftime'):
                    anticipo['fecha_pago'] = anticipo['fecha_pago'].strftime('%Y-%m-%d')
                else:
                    anticipo['fecha_pago'] = str(anticipo['fecha_pago'])

            if anticipo.get('fecha_evento'):
                if hasattr(anticipo['fecha_evento'], 'strftime'):
                    anticipo['fecha_evento'] = anticipo['fecha_evento'].strftime('%Y-%m-%d')
                else:
                    anticipo['fecha_evento'] = str(anticipo['fecha_evento'])

            # Convertir timestamps a ISO string
            for campo in ['created_at', 'updated_at', 'deleted_at']:
                if anticipo.get(campo):
                    if hasattr(anticipo[campo], 'isoformat'):
                        anticipo[campo] = anticipo[campo].isoformat()
                    else:
                        anticipo[campo] = str(anticipo[campo])

            # Si fue consumido en alguna caja, el estado debe ser 'consumido'
            if anticipo['fue_consumido'] > 0:
                anticipo['estado'] = 'consumido'
            else:
                anticipo['estado'] = anticipo['estado_global']

            # Eliminar campos auxiliares
            del anticipo['estado_global']
            del anticipo['fue_consumido']

            anticipos.append(anticipo)

        cur.close()
        conn.close()

        return jsonify(success=True, anticipos=anticipos)

    except Exception as e:
        print("❌ ERROR listar_anticipos_recibidos:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/anticipos_recibidos/eliminar/<int:anticipo_id>', methods=['DELETE'])
@login_required
def eliminar_anticipo_recibido(anticipo_id):
    """
    Eliminar (marcar como eliminado_global) un anticipo recibido.
    SOLO accesible para admin_anticipos (nivel 6) - rol 'anticipos' NO puede eliminar.
    """
    user_level = get_user_level()
    if user_level < 6:
        return jsonify(success=False, msg="No tenés permisos para eliminar anticipos recibidos"), 403

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que existe y obtener datos para auditoría
        cur.execute("SELECT * FROM anticipos_recibidos WHERE id = %s", (anticipo_id,))
        anticipo = cur.fetchone()

        if not anticipo:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Anticipo no encontrado"), 404

        if anticipo['estado'] == 'consumido':
            cur.close()
            conn.close()
            return jsonify(success=False, msg="No se puede eliminar un anticipo ya consumido"), 409

        usuario = session.get('username', 'sistema')

        # Marcar como eliminado
        from datetime import datetime
        import pytz
        tz_arg = pytz.timezone('America/Argentina/Buenos_Aires')
        ahora = datetime.now(tz_arg)

        cur.execute("""
            UPDATE anticipos_recibidos
            SET estado = 'eliminado_global',
                deleted_by = %s,
                deleted_at = %s,
                updated_by = %s,
                updated_at = %s
            WHERE id = %s
        """, (usuario, ahora, usuario, ahora, anticipo_id))

        conn.commit()

        # Registrar en auditoría
        from modules.tabla_auditoria import registrar_auditoria
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='anticipos_recibidos',
            registro_id=anticipo_id,
            datos_anteriores={'estado': anticipo['estado']},
            datos_nuevos={'estado': 'eliminado_global'},
            descripcion=f"Anticipo eliminado: {anticipo['cliente']} - ${anticipo['importe']}"
        )

        cur.close()
        conn.close()

        return jsonify(success=True, msg="Anticipo eliminado correctamente")

    except Exception as e:
        print("❌ ERROR eliminar_anticipo_recibido:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/anticipos_recibidos/editar/<int:anticipo_id>', methods=['PUT'])
@login_required
def editar_anticipo_recibido(anticipo_id):
    """
    Editar un anticipo recibido.
    Solo permite editar fecha_evento y observaciones si el anticipo está pendiente.
    SOLO accesible para admin_anticipos (nivel 6) - rol 'anticipos' NO puede editar.
    """
    user_level = get_user_level()
    if user_level < 6:
        return jsonify(success=False, msg="No tenés permisos para editar anticipos recibidos"), 403

    data = request.get_json() or {}

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que existe
        cur.execute("SELECT * FROM anticipos_recibidos WHERE id = %s", (anticipo_id,))
        anticipo = cur.fetchone()

        if not anticipo:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Anticipo no encontrado"), 404

        if anticipo['estado'] != 'pendiente':
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Solo se pueden editar anticipos pendientes"), 409

        # Solo permitir editar fecha_evento y observaciones
        nueva_fecha_evento = data.get('fecha_evento')
        nuevas_observaciones = data.get('observaciones')

        if not nueva_fecha_evento:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Fecha de evento requerida"), 400

        fecha_evento_norm = _normalize_fecha(nueva_fecha_evento)
        usuario = session.get('username', 'sistema')

        from datetime import datetime
        import pytz
        tz_arg = pytz.timezone('America/Argentina/Buenos_Aires')
        ahora = datetime.now(tz_arg)

        cur.execute("""
            UPDATE anticipos_recibidos
            SET fecha_evento = %s,
                observaciones = %s,
                updated_by = %s,
                updated_at = %s
            WHERE id = %s
        """, (fecha_evento_norm, nuevas_observaciones, usuario, ahora, anticipo_id))

        conn.commit()

        # Registrar en auditoría
        from modules.tabla_auditoria import registrar_auditoria
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='anticipos_recibidos',
            registro_id=anticipo_id,
            datos_anteriores={
                'fecha_evento': str(anticipo['fecha_evento']),
                'observaciones': anticipo['observaciones']
            },
            datos_nuevos={
                'fecha_evento': str(fecha_evento_norm),
                'observaciones': nuevas_observaciones
            },
            descripcion=f"Fecha de evento actualizada para anticipo de {anticipo['cliente']}"
        )

        cur.close()
        conn.close()

        return jsonify(success=True, msg="Anticipo actualizado correctamente")

    except Exception as e:
        print("❌ ERROR editar_anticipo_recibido:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


# ───────────────────────────────────────────────────────────────────
# PARTE 2: CONSUMO DE ANTICIPOS EN CAJAS (cajeros)
# ───────────────────────────────────────────────────────────────────

@app.route('/api/anticipos/disponibles', methods=['GET'])
@login_required
def obtener_anticipos_disponibles():
    """
    Obtener anticipos disponibles para una caja específica.

    Un anticipo aparece si:
    1. Su local coincide con el local de la caja
    2. Su fecha_evento coincide con la fecha de la caja
    3. Su estado global es 'pendiente'
    4. NO existe un registro en anticipos_estados_caja para esta caja (ni eliminado ni consumido)

    Params:
    - local: local de la caja
    - caja: caja
    - fecha: fecha de operación
    - turno: turno (informativo, no afecta disponibilidad)
    """
    local = get_local_param()
    caja = request.args.get('caja', '').strip()
    fecha = request.args.get('fecha', '').strip()
    turno = request.args.get('turno', '').strip()

    if not (local and caja and fecha):
        return jsonify(success=True, disponibles=[])

    fecha_norm = _normalize_fecha(fecha)

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Obtener anticipos disponibles
        sql = """
            SELECT
                ar.id, ar.fecha_pago, ar.fecha_evento, ar.importe, ar.divisa,
                ar.cliente, ar.numero_transaccion, ar.medio_pago,
                ar.observaciones, ar.local, ar.created_by, ar.created_at,
                ia.gcs_path as adjunto_gcs_path,
                ia.original_name as adjunto_nombre
            FROM anticipos_recibidos ar
            LEFT JOIN imagenes_adjuntos ia ON (
                ia.entity_type = 'anticipo_recibido'
                AND ia.entity_id = ar.id
                AND ia.estado = 'active'
            )
            WHERE ar.local = %s
              AND ar.fecha_evento = %s
              AND ar.estado = 'pendiente'
              AND NOT EXISTS (
                  SELECT 1 FROM anticipos_estados_caja aec
                  WHERE aec.anticipo_id = ar.id
                    AND aec.local = %s
                    AND aec.caja = %s
                    AND aec.fecha = %s
              )
            ORDER BY ar.created_at ASC
        """

        cur.execute(sql, (local, fecha_norm, local, caja, fecha_norm))
        disponibles = cur.fetchall() or []

        # Convertir fechas a string para JSON
        for anticipo in disponibles:
            if anticipo.get('fecha_pago'):
                anticipo['fecha_pago'] = anticipo['fecha_pago'].isoformat() if hasattr(anticipo['fecha_pago'], 'isoformat') else str(anticipo['fecha_pago'])
            if anticipo.get('fecha_evento'):
                anticipo['fecha_evento'] = anticipo['fecha_evento'].isoformat() if hasattr(anticipo['fecha_evento'], 'isoformat') else str(anticipo['fecha_evento'])
            if anticipo.get('created_at'):
                anticipo['created_at'] = anticipo['created_at'].isoformat() if hasattr(anticipo['created_at'], 'isoformat') else str(anticipo['created_at'])

        cur.close()
        conn.close()

        return jsonify(success=True, disponibles=disponibles)

    except Exception as e:
        print("❌ ERROR obtener_anticipos_disponibles:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/anticipos/consumir', methods=['POST'])
@login_required
@require_edit_ctx
def consumir_anticipo():
    """
    Consumir un anticipo en una caja específica.

    Body:
    {
        "local": "...",
        "caja": "...",
        "fecha": "...",
        "turno": "...",
        "anticipo_id": 123,
        "observaciones_consumo": "Cliente consumió en mesa 5"
    }
    """
    data = request.get_json() or {}
    anticipo_id = data.get('anticipo_id')

    if not anticipo_id:
        return jsonify(success=False, msg="anticipo_id requerido"), 400

    # Contexto validado por el decorador
    ctx = g.ctx
    local = ctx['local']
    caja = ctx['caja']
    fecha = _normalize_fecha(ctx['fecha'])
    turno = ctx['turno']
    usuario = session.get('username', 'sistema')

    observaciones_consumo = (data.get('observaciones_consumo') or '').strip() or None

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que el anticipo existe y está disponible
        cur.execute("""
            SELECT * FROM anticipos_recibidos
            WHERE id = %s AND estado = 'pendiente'
        """, (anticipo_id,))
        anticipo = cur.fetchone()

        if not anticipo:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Anticipo no encontrado o ya fue consumido"), 404

        # Verificar que corresponde al local y fecha
        if anticipo['local'] != local or str(anticipo['fecha_evento']) != str(fecha):
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Este anticipo no corresponde a este local/fecha"), 400

        # Verificar que no fue ya procesado en esta caja
        cur.execute("""
            SELECT * FROM anticipos_estados_caja
            WHERE anticipo_id = %s AND local = %s AND caja = %s AND fecha = %s
        """, (anticipo_id, local, caja, fecha))

        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Este anticipo ya fue procesado en esta caja"), 409

        from datetime import datetime
        import pytz
        tz_arg = pytz.timezone('America/Argentina/Buenos_Aires')
        ahora = datetime.now(tz_arg)

        # Registrar consumo en estados_caja
        cur.execute("""
            INSERT INTO anticipos_estados_caja
            (anticipo_id, local, caja, fecha, turno, estado, usuario,
             timestamp_accion, importe_consumido, observaciones_consumo)
            VALUES (%s, %s, %s, %s, %s, 'consumido', %s, %s, %s, %s)
        """, (anticipo_id, local, caja, fecha, turno, usuario, ahora,
              anticipo['importe'], observaciones_consumo))

        # Actualizar estado global del anticipo a 'consumido'
        cur.execute("""
            UPDATE anticipos_recibidos
            SET estado = 'consumido',
                updated_by = %s,
                updated_at = %s
            WHERE id = %s
        """, (usuario, ahora, anticipo_id))

        conn.commit()

        # Registrar en auditoría
        from modules.tabla_auditoria import registrar_auditoria
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='anticipos_recibidos',
            registro_id=anticipo_id,
            datos_anteriores={'estado': 'pendiente'},
            datos_nuevos={'estado': 'consumido', 'caja': caja, 'turno': turno},
            descripcion=f"Anticipo consumido en caja {caja}: {anticipo['cliente']} - ${anticipo['importe']}",
            contexto_override={'local': local, 'caja': caja, 'fecha': fecha, 'turno': turno}
        )

        cur.close()
        conn.close()

        return jsonify(success=True, msg=f"Anticipo consumido correctamente (${anticipo['importe']})")

    except Exception as e:
        import traceback
        import logging
        logger = logging.getLogger('app')
        logger.error(f"❌ ERROR consumir_anticipo: {e}")
        logger.error(traceback.format_exc())
        print("❌ ERROR consumir_anticipo:", e)
        traceback.print_exc()
        return jsonify(success=False, msg=f"Error interno: {str(e)}"), 500


# ENDPOINT ELIMINADO: Ya no se permite marcar anticipos como "no vino a esta caja"
# Los anticipos sin consumir simplemente quedarán pendientes y bloquearán el cierre del local


@app.route('/api/anticipos/desconsumir', methods=['POST'])
@login_required
@require_edit_ctx
def desconsumir_anticipo():
    """
    Desconsumir un anticipo que fue consumido por error.
    Solo permitido si la caja está abierta (cajero) o el local está abierto (encargado).

    Body:
    {
        "local": "...",
        "caja": "...",
        "fecha": "...",
        "turno": "...",
        "anticipo_id": 123
    }
    """
    data = request.get_json() or {}
    anticipo_id = data.get('anticipo_id')

    if not anticipo_id:
        return jsonify(success=False, msg="anticipo_id requerido"), 400

    # Contexto validado por el decorador
    ctx = g.ctx
    local = ctx['local']
    caja = ctx['caja']
    fecha = _normalize_fecha(ctx['fecha'])
    turno = ctx['turno']
    usuario = session.get('username', 'sistema')

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que el anticipo existe y está consumido
        cur.execute("""
            SELECT * FROM anticipos_recibidos
            WHERE id = %s AND estado = 'consumido'
        """, (anticipo_id,))
        anticipo = cur.fetchone()

        if not anticipo:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Anticipo no encontrado o no está consumido"), 404

        # Verificar que fue consumido en ESTA caja
        cur.execute("""
            SELECT * FROM anticipos_estados_caja
            WHERE anticipo_id = %s
              AND local = %s
              AND caja = %s
              AND fecha = %s
              AND estado = 'consumido'
        """, (anticipo_id, local, caja, fecha))

        registro_consumo = cur.fetchone()

        if not registro_consumo:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Este anticipo no fue consumido en esta caja"), 404

        from datetime import datetime
        import pytz
        tz_arg = pytz.timezone('America/Argentina/Buenos_Aires')
        ahora = datetime.now(tz_arg)

        # Eliminar el registro de consumo en estados_caja
        cur.execute("""
            DELETE FROM anticipos_estados_caja
            WHERE anticipo_id = %s
              AND local = %s
              AND caja = %s
              AND fecha = %s
              AND estado = 'consumido'
        """, (anticipo_id, local, caja, fecha))

        # Volver el anticipo a estado 'pendiente'
        cur.execute("""
            UPDATE anticipos_recibidos
            SET estado = 'pendiente',
                updated_by = %s,
                updated_at = %s
            WHERE id = %s
        """, (usuario, ahora, anticipo_id))

        conn.commit()

        # Registrar en auditoría
        from modules.tabla_auditoria import registrar_auditoria
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='anticipos_recibidos',
            registro_id=anticipo_id,
            datos_anteriores={'estado': 'consumido'},
            datos_nuevos={'estado': 'pendiente'},
            descripcion=f"Anticipo desconsumido de caja {caja}: {anticipo['cliente']} - ${anticipo['importe']}",
            contexto_override={'local': local, 'caja': caja, 'fecha': fecha, 'turno': turno}
        )

        cur.close()
        conn.close()

        return jsonify(success=True, msg=f"Anticipo desconsumido correctamente. Vuelve a estar disponible.")

    except Exception as e:
        import traceback
        import logging
        logger = logging.getLogger('app')
        logger.error(f"❌ ERROR desconsumir_anticipo: {e}")
        logger.error(traceback.format_exc())
        print("❌ ERROR desconsumir_anticipo:", e)
        traceback.print_exc()
        return jsonify(success=False, msg=f"Error interno: {str(e)}"), 500


@app.route('/api/anticipos/consumidos_en_caja', methods=['GET'])
@login_required
def obtener_anticipos_consumidos_en_caja():
    """
    Obtener anticipos que fueron consumidos en una caja específica.
    Útil para mostrar en el resumen de la caja.

    Params:
    - local
    - caja
    - fecha
    - turno
    """
    local = get_local_param()
    caja = request.args.get('caja', '').strip()
    fecha = request.args.get('fecha', '').strip()
    turno = request.args.get('turno', '').strip()

    if not (local and caja and fecha and turno):
        return jsonify(success=True, consumidos=[])

    fecha_norm = _normalize_fecha(fecha)

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        sql = """
            SELECT
                aec.id, aec.anticipo_id, aec.timestamp_accion,
                aec.importe_consumido, aec.observaciones_consumo,
                ar.cliente, ar.numero_transaccion, ar.medio_pago,
                ar.fecha_pago, ar.observaciones as observaciones_anticipo,
                ar.divisa, ar.created_by
            FROM anticipos_estados_caja aec
            JOIN anticipos_recibidos ar ON ar.id = aec.anticipo_id
            WHERE aec.local = %s
              AND aec.caja = %s
              AND aec.fecha = %s
              AND aec.turno = %s
              AND aec.estado = 'consumido'
            ORDER BY aec.timestamp_accion ASC
        """

        cur.execute(sql, (local, caja, fecha_norm, turno))
        consumidos = cur.fetchall() or []

        # Convertir fechas a string para JSON
        for anticipo in consumidos:
            if anticipo.get('fecha_pago'):
                anticipo['fecha_pago'] = anticipo['fecha_pago'].isoformat() if hasattr(anticipo['fecha_pago'], 'isoformat') else str(anticipo['fecha_pago'])
            if anticipo.get('timestamp_accion'):
                anticipo['timestamp_accion'] = anticipo['timestamp_accion'].isoformat() if hasattr(anticipo['timestamp_accion'], 'isoformat') else str(anticipo['timestamp_accion'])

        cur.close()
        conn.close()

        return jsonify(success=True, consumidos=consumidos)

    except Exception as e:
        print("❌ ERROR obtener_anticipos_consumidos_en_caja:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/anticipos/validar_cierre_local', methods=['GET'])
@login_required
def validar_cierre_local_anticipos():
    """
    Validar si quedan anticipos sin consumir antes de cerrar el local.

    Params:
    - local
    - fecha

    Returns:
    {
        "puede_cerrar": true/false,
        "anticipos_pendientes": [...],  # Si hay anticipos sin consumir
        "msg": "..."
    }
    """
    local = get_local_param()
    fecha = request.args.get('fecha', '').strip()

    if not (local and fecha):
        return jsonify(success=False, msg="Local y fecha requeridos"), 400

    fecha_norm = _normalize_fecha(fecha)

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Buscar anticipos que:
        # 1. Corresponden al local y fecha_evento
        # 2. Estado global es 'pendiente'
        # 3. NO fueron consumidos NI eliminados de TODAS las cajas
        sql = """
            SELECT
                ar.id, ar.cliente, ar.importe, ar.numero_transaccion,
                ar.observaciones
            FROM anticipos_recibidos ar
            WHERE ar.local = %s
              AND ar.fecha_evento = %s
              AND ar.estado = 'pendiente'
        """

        cur.execute(sql, (local, fecha_norm))
        anticipos_pendientes = cur.fetchall() or []

        cur.close()
        conn.close()

        if anticipos_pendientes:
            return jsonify(
                success=True,
                puede_cerrar=False,
                anticipos_pendientes=anticipos_pendientes,
                msg=f"Hay {len(anticipos_pendientes)} anticipo(s) sin consumir. No se puede cerrar el local."
            )
        else:
            return jsonify(
                success=True,
                puede_cerrar=True,
                anticipos_pendientes=[],
                msg="No hay anticipos pendientes. Se puede cerrar el local."
            )

    except Exception as e:
        print("❌ ERROR validar_cierre_local_anticipos:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


# ═══════════════════════════════════════════════════════════════════
# FIN NUEVO SISTEMA DE ANTICIPOS RECIBIDOS
# ═══════════════════════════════════════════════════════════════════


# ─────────────────────────────────────────
# VENTA SISTEMA (única por fecha/caja/local)
# ─────────────────────────────────────────

# ─────────────────────────────────────
# Helpers comunes (si ya existen, usá los tuyos)
# ─────────────────────────────────────
from datetime import datetime

def _normalize_fecha_str(s):
    if not s:
        return None
    if isinstance(s, datetime):
        return s.strftime("%Y-%m-%d")
    if hasattr(s, 'strftime'):
        return s.strftime("%Y-%m-%d")
    if isinstance(s, str):
        try:
            return datetime.strptime(s, "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            return s
    return None

def _get_turno_from_request(data):
    # Acepta por body o por querystring; default 'UNI' si faltara
    return (data.get('turno')
            or request.args.get('turno')
            or 'UNI')

# ─────────────────────────────────────
# VENTA SISTEMA (única por fecha/caja/turno/local)
# ─────────────────────────────────────

# ─────────────────────────────────────
# VENTAS BASE
# ─────────────────────────────────────

@app.route('/ventas_base', methods=['POST', 'PUT', 'DELETE'])
@login_required
def ventas_base():
    """
    POST   -> Crea la base si NO existe para (local,caja,fecha,turno).
              body: { fecha, caja, turno, venta_total_sistema }  (acepta 'venta_total' como alias)
    PUT    -> Actualiza la base existente (solo venta_total_sistema).
              body: { fecha, caja, turno, venta_total_sistema }  (acepta 'venta_total' como alias)
    DELETE -> Borra la base.
              body: { fecha, caja, turno }
    """
    from modules.tabla_auditoria import registrar_auditoria, obtener_registro_anterior

    data   = request.get_json() or {}
    # Para auditores: usar local del body o del request, sino usar get_local_param()
    local  = data.get('local') or get_local_param()
    user   = session.get('username')

    if not local or not user:
        return jsonify(success=False, msg="Faltan datos de sesión (usuario/local)."), 401

    fecha = data.get('fecha')
    caja  = data.get('caja')
    turno = _get_turno_from_request(data)

    if not fecha or not caja or not turno:
        return jsonify(success=False, msg="Faltan parámetros (fecha/caja/turno)."), 400

    # normalizamos fecha para todas las comparaciones
    nfecha = _normalize_fecha(fecha)

    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)

    try:
        # Permisos por rol/estado
        if not can_edit(conn, local, caja, turno, nfecha, get_user_level()):
            return jsonify(success=False, msg="No permitido (caja/local cerrados para tu rol)"), 409

        # Resolvemos monto (acepta venta_total_sistema o venta_total)
        vts = data.get('venta_total_sistema')
        if vts is None:
            vts = data.get('venta_total')
        try:
            vts = float(str(vts or 0).replace('.', '').replace(',', '.'))
        except Exception:
            vts = 0.0

        if request.method == 'POST':
            # ¿existe ya una base para ese turno?
            cur.execute("""
                SELECT id
                  FROM ventas_trns
                 WHERE local=%s AND caja=%s AND DATE(fecha)=%s AND turno=%s
                 LIMIT 1
            """, (local, caja, nfecha, turno))
            row = cur.fetchone()
            if row:
                return jsonify(success=False, msg="Ya existe una venta para esa fecha/caja/turno."), 409

            cur2 = conn.cursor()
            cur2.execute("""
                INSERT INTO ventas_trns (usuario, local, caja, turno, fecha, venta_total_sistema, estado)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (user, local, caja, turno, nfecha, vts, "revision"))
            nuevo_id = cur2.lastrowid
            cur2.close()
            conn.commit()

            # AUDITORÍA: Registrar INSERT
            registrar_auditoria(
                conn=conn,
                accion='INSERT',
                tabla='ventas_trns',
                registro_id=nuevo_id,
                datos_nuevos={
                    'usuario': user,
                    'local': local,
                    'caja': caja,
                    'turno': turno,
                    'fecha': str(nfecha),
                    'venta_total_sistema': vts,
                    'estado': 'revision'
                },
                descripcion=f"Nueva venta Z creada - Monto: ${vts:,.2f}",
                contexto_override={
                    'local': local,
                    'caja': caja,
                    'fecha_operacion': str(nfecha),
                    'turno': turno
                }
            )

            return jsonify(success=True)

        if request.method == 'PUT':
            cur.execute("""
                SELECT id
                  FROM ventas_trns
                 WHERE local=%s AND caja=%s AND DATE(fecha)=%s AND turno=%s
                 LIMIT 1
            """, (local, caja, nfecha, turno))
            row = cur.fetchone()
            if not row:
                return jsonify(success=False, msg="No existe base para esa fecha/caja/turno."), 404

            # AUDITORÍA: Capturar estado anterior
            datos_anteriores = obtener_registro_anterior(conn, 'ventas_trns', row['id'])

            cur2 = conn.cursor()
            cur2.execute("""
                UPDATE ventas_trns
                   SET venta_total_sistema=%s, usuario=%s
                 WHERE id=%s
            """, (vts, user, row['id']))
            cur2.close()
            conn.commit()

            # AUDITORÍA: Registrar UPDATE
            registrar_auditoria(
                conn=conn,
                accion='UPDATE',
                tabla='ventas_trns',
                registro_id=row['id'],
                datos_anteriores=datos_anteriores,
                datos_nuevos={
                    'venta_total_sistema': vts,
                    'usuario': user
                },
                descripcion=f"Venta Z actualizada - Monto anterior: ${datos_anteriores.get('venta_total_sistema', 0):,.2f} → Nuevo: ${vts:,.2f}",
                contexto_override={
                    'local': local,
                    'caja': caja,
                    'fecha_operacion': str(nfecha),
                    'turno': turno
                }
            )

            return jsonify(success=True)

        # DELETE
        # AUDITORÍA: Capturar estado antes de borrar
        cur.execute("""
            SELECT * FROM ventas_trns
             WHERE local=%s AND caja=%s AND DATE(fecha)=%s AND turno=%s
             LIMIT 1
        """, (local, caja, nfecha, turno))
        registro_a_borrar = cur.fetchone()

        cur.execute("""
            DELETE FROM ventas_trns
             WHERE local=%s AND caja=%s AND DATE(fecha)=%s AND turno=%s
        """, (local, caja, nfecha, turno))
        conn.commit()

        # AUDITORÍA: Registrar DELETE
        if registro_a_borrar:
            registrar_auditoria(
                conn=conn,
                accion='DELETE',
                tabla='ventas_trns',
                registro_id=registro_a_borrar.get('id'),
                datos_anteriores=registro_a_borrar,
                descripcion=f"Venta Z eliminada - Monto: ${registro_a_borrar.get('venta_total_sistema', 0):,.2f}",
                contexto_override={
                    'local': local,
                    'caja': caja,
                    'fecha_operacion': str(nfecha),
                    'turno': turno
                }
            )

        return jsonify(success=True)

    except Exception as e:
        conn.rollback()
        return jsonify(success=False, msg=str(e)), 500
    finally:
        cur.close()
        conn.close()


@app.route('/ventas_cargadas')
@login_required
@with_read_scope('t')
def ventas_cargadas():
    """
    GET -> Devuelve la BASE (única) de la fecha/caja/turno/local.
    resp: { success:true, datos:[ { id, fecha:'YYYY-MM-DD', caja, turno, venta_total:<decimal> } ] }
    """
    local = get_local_param()  # Nivel 3 puede filtrar por cualquier local
    fecha = request.args.get('fecha')
    caja  = request.args.get('caja')
    turno = request.args.get('turno') or 'UNI'

    if not local or not fecha or not caja or not turno:
        return jsonify(success=False, msg="Faltan parámetros"), 400

    try:
        nfecha = _normalize_fecha(fecha)
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)
        cur.execute(f"""
            SELECT
              t.id,
              DATE_FORMAT(t.fecha,'%Y-%m-%d') AS fecha,
              t.caja,
              t.turno,
              t.venta_total_sistema AS venta_total
            FROM ventas_trns t
            WHERE t.local=%s
              AND t.caja=%s
              AND t.turno=%s
              AND DATE(t.fecha)=%s
              {g.read_scope}  -- L2: sólo cajas cerradas; L3: sólo locales cerrados
            LIMIT 1
        """, (local, caja, turno, nfecha))
        datos = cur.fetchall()
        cur.close(); conn.close()
        return jsonify(success=True, datos=datos)
    except Exception as e:
        return jsonify(success=False, msg=str(e)), 500


# ─────────────────────────────────────
# _______________FACTURAS Z, A, B, CC__________________
# ─────────────────────────────────────
# ===================== BLOQUE FACTURAS (pegar completo en app.py) =====================
# Requisitos previos existentes en tu app.py:
# - from flask import request, jsonify, session, current_app
# - import mysql.connector
# - get_db_connection()
# - login_required, page_access_required
# - get_user_level()
# - can_edit(conn, local, caja, turno, fecha, user_level)  # firma existente en tu app
# - (opcional) is_local_closed / is_box_closed que ya usa can_edit internamente

import unicodedata

# ---------- Helpers locales del módulo Facturas (no rompen nada del resto) ----------

def _f_strip_accents(s: str) -> str:
    if not isinstance(s, str):
        return s
    return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))

def _f_norm_turno_for_can_edit(t: str) -> str:
    """
    Normaliza el turno SOLO para el chequeo de permisos:
    - Saca tildes (día -> dia)
    - Recorta espacios
    - Mapea a valores esperados por tu can_edit
    Nota: NO cambia lo que insertamos en DB (ahí guardamos el turno tal cual venga).
    """
    if not t:
        return t
    t0 = _f_strip_accents(str(t)).strip()
    tl = t0.lower()
    if tl in ("uni", "turno uni", "turno unico", "turno unico/uni"):
        return "UNI"
    if tl in ("turno dia", "dia"):
        return "Turno dia"
    if tl in ("turno noche", "noche"):
        return "Turno noche"
    # fallback: devolver sin acentos
    return t0

def _f_norm(x):
    if x is None:
        return ""
    return str(x).strip()

def _f_get_json():
    try:
        if request.is_json:
            return request.get_json(force=True, silent=True) or {}
        # fallback para forms
        return dict(request.form) if request.form else {}
    except Exception:
        return {}

def _f_parse_monto(v) -> float:
    """
    Acepta '50.000,25' o '50000.25' y devuelve float. Vacío => 0.
    """
    if v is None:
        return 0.0
    s = str(v).strip()
    if not s:
        return 0.0
    # limpiar símbolos
    s = s.replace(" ", "").replace("$", "")
    # detectar último separador decimal (coma o punto)
    last_comma = s.rfind(",")
    last_dot   = s.rfind(".")
    last_sep   = max(last_comma, last_dot)
    if last_sep >= 0:
        int_part = s[:last_sep].replace(".", "").replace(",", "")
        frac     = s[last_sep+1:].replace(".", "").replace(",", "")
        s = f"{int_part}.{frac}" if frac else int_part
    else:
        s = s.replace(".", "").replace(",", "")
    try:
        n = float(s)
        if not (n == n and n != float("inf") and n != float("-inf")):
            return 0.0
        return n
    except Exception:
        return 0.0

def _f_ctx_from_request():
    body = request.get_json(silent=True) or {}
    local = (
        request.args.get("local")
        or body.get("local")
        or (request.form.get("local") if request.form else None)
        # fallbacks de sesión habituales en tu app:
        or (session.get("local") if session else None)
        or (session.get("Local") if session else None)
        or (session.get("user_local") if session else None)
        or (session.get("local_name") if session else None)
        # fallback por header opcional (si alguna vista lo manda):
        or request.headers.get("X-Local")
        or ""
    )
    caja  = request.args.get("caja")  or body.get("caja")  or (request.form.get("caja")  if request.form else None)  or ""
    fecha = request.args.get("fecha") or body.get("fecha") or (request.form.get("fecha") if request.form else None) or ""
    turno = request.args.get("turno") or body.get("turno") or (request.form.get("turno") if request.form else None) or ""

    ctx = {
        "local": (str(local).strip() if local else ""),
        "caja":  (str(caja).strip()  if caja  else ""),
        "fecha": (str(fecha).strip() if fecha else ""),
        "turno": (str(turno).strip() if turno else ""),
    }
    ctx["turno_canedit"] = _f_norm_turno_for_can_edit(ctx["turno"])
    return ctx


def _f_level_int() -> int:
    try:
        return int(get_user_level())
    except Exception:
        return 1

def _f_is_box_closed(conn, ctx) -> bool:
    """
    Verifica si la caja está cerrada para el contexto dado.
    Retorna True si la caja está cerrada (estado=0), False en caso contrario.
    """
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT estado FROM cajas_estado
                WHERE local=%s AND caja=%s AND LOWER(turno)=LOWER(%s) AND DATE(fecha_operacion)=%s
                  AND id = (
                    SELECT MAX(id) FROM cajas_estado
                    WHERE local=%s AND caja=%s AND LOWER(turno)=LOWER(%s) AND DATE(fecha_operacion)=%s
                  )
                """,
                (ctx["local"], ctx["caja"], ctx["turno"], ctx["fecha"],
                 ctx["local"], ctx["caja"], ctx["turno"], ctx["fecha"])
            )
            row = cur.fetchone()
            return row and row[0] == 0
    except Exception:
        return False

def _f_is_local_closed(conn, ctx) -> bool:
    """
    Verifica si el local está cerrado para el contexto dado.
    Retorna True si el local está cerrado (estado=0), False en caso contrario.
    """
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT estado FROM cierres_locales
                WHERE local=%s AND DATE(fecha)=%s
                """,
                (ctx["local"], ctx["fecha"])
            )
            row = cur.fetchone()
            return row and row[0] == 0
    except Exception:
        return False

def _f_is_local_auditado(conn, ctx) -> bool:
    """
    Verifica si el local está auditado para el contexto dado.
    Retorna True si el local está auditado, False en caso contrario.
    """
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id FROM locales_auditados
                WHERE local=%s AND DATE(fecha)=%s
                """,
                (ctx["local"], ctx["fecha"])
            )
            row = cur.fetchone()
            return row is not None
    except Exception:
        return False

def _f_safe_can_edit(conn, ctx) -> bool:
    """
    Verifica si el usuario puede editar basándose en su nivel y el estado de caja/local:
    - Nivel 1 (cajero): puede editar solo si la caja NO está cerrada
    - Nivel 2 (encargado): puede editar solo si el local NO está cerrado
    - Nivel 3 (auditor): puede editar solo si el local está cerrado pero NO auditado
    - NADIE puede editar si el local está auditado
    """
    lvl = _f_level_int()

    # Si el local está auditado, NADIE puede editar (inmutable)
    if _f_is_local_auditado(conn, ctx):
        return False

    # Nivel 1 (cajero): no puede editar si la caja está cerrada
    if lvl == 1:
        return not _f_is_box_closed(conn, ctx)

    # Nivel 2 (encargado): no puede editar si el local está cerrado
    if lvl == 2:
        return not _f_is_local_closed(conn, ctx)

    # Nivel 3+ (auditor): puede editar solo si el local está cerrado
    if lvl >= 3:
        return _f_is_local_closed(conn, ctx)

    return False

def _f_check_can_edit_or_409(conn, ctx):
    """
    Devuelve (True, ()) si puede editar; si no, (False, (respuesta_json_409)).
    """
    # Verificar si está auditado primero (bloquea a todos)
    if _f_is_local_auditado(conn, ctx):
        return False, (jsonify(success=False, msg="No puede editar: el local está auditado (inmutable)"), 409)

    allowed = _f_safe_can_edit(conn, ctx)
    if not allowed:
        lvl = _f_level_int()
        if lvl == 1:
            msg = "No puede editar: la caja está cerrada"
        elif lvl == 2:
            msg = "No puede editar: el local está cerrado"
        elif lvl >= 3:
            msg = "No puede editar: el local debe estar cerrado para auditar"
        else:
            msg = "No permitido para tu rol/estado"
        return False, (jsonify(success=False, msg=msg), 409)
    return True, ()

# ---------- ENDPOINTS FACTURAS ----------

@app.post("/facturas")
@login_required
@page_access_required("index")
def facturas_create():
    """
    Crea una factura (Z/A/B/CC) en facturas_trns.
    Body JSON: { tipo, punto_venta, nro_factura, comentario?, monto, estado? }
    Requiere local/caja/fecha/turno + can_edit ok.
    """
    from modules.tabla_auditoria import registrar_auditoria

    ctx = _f_ctx_from_request()
    if not all([ctx["local"], ctx["caja"], ctx["fecha"], ctx["turno"]]):
        return jsonify(success=False, msg="falta local/caja/fecha/turno"), 400

    data = _f_get_json()
    tipo        = _f_norm(data.get("tipo")).upper()
    punto_venta = _f_norm(data.get("punto_venta"))
    nro_factura = _f_norm(data.get("nro_factura"))
    comentario  = _f_norm(data.get("comentario") or "")
    monto       = _f_parse_monto(data.get("monto"))
    estado      = _f_norm(data.get("estado") or "ok")
    usuario     = session.get("username") or "system"

    if tipo not in {"Z", "A", "B", "CC"}:
        return jsonify(success=False, msg="tipo inválido (Z|A|B|CC)"), 400
    if not punto_venta or not nro_factura or not (monto > 0):
        return jsonify(success=False, msg="Faltan/invalidos: punto_venta / nro_factura / monto"), 400

    conn = get_db_connection()
    try:
        ok, resp = _f_check_can_edit_or_409(conn, ctx)
        if not ok:
            return resp

        with conn.cursor() as cur:
            try:
                cur.execute(
                    """
                    INSERT INTO facturas_trns
                      (local, caja, turno, fecha, tipo, punto_venta, nro_factura, comentario, monto, estado, usuario, update_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        ctx["local"], ctx["caja"], ctx["turno"], ctx["fecha"],
                        tipo, punto_venta, nro_factura, comentario, monto, estado, usuario, usuario
                    ),
                )
                conn.commit()
                new_id = cur.lastrowid
            except mysql.connector.IntegrityError:
                return jsonify(success=False, msg="Factura duplicada para ese contexto"), 409

        # AUDITORÍA: Registrar INSERT
        registrar_auditoria(
            conn=conn,
            accion='INSERT',
            tabla='facturas_trns',
            registro_id=new_id,
            datos_nuevos={
                'tipo': tipo,
                'punto_venta': punto_venta,
                'nro_factura': nro_factura,
                'comentario': comentario,
                'monto': monto,
                'estado': estado,
                'usuario': usuario
            },
            descripcion=f"Nueva factura {tipo} creada - PV: {punto_venta} / Nro: {nro_factura} / Monto: ${monto:,.2f}",
            contexto_override={
                'local': ctx["local"],
                'caja': ctx["caja"],
                'fecha_operacion': str(ctx["fecha"]),
                'turno': ctx["turno"]
            }
        )

        with conn.cursor(dictionary=True) as cur2:
            cur2.execute(
                """
                SELECT id, local, caja, turno, fecha, tipo, punto_venta, nro_factura, comentario, monto, estado,
                       usuario, created_at, updated_at, update_by
                  FROM facturas_trns
                 WHERE id=%s
                """,
                (new_id,),
            )
            row = cur2.fetchone()
        return jsonify(success=True, item=row), 201
    finally:
        try:
            conn.close()
        except Exception:
            pass


@app.get("/facturas_cargadas")
@login_required
@page_access_required("index")
@with_read_scope('f')
def facturas_cargadas():
    """
    Lista facturas de la fecha/caja/turno actuales (o filtros por query).
    Query: ?fecha=YYYY-MM-DD&caja=Caja%201&turno=Turno%20d%C3%ADa&tipo=Z|A|B|CC

    Control de visibilidad por rol:
    - Cajero (lvl 1): ve todo (sin restricción de estado)
    - Encargado (lvl 2): solo cajas cerradas
    - Auditor (lvl 3): solo locales cerrados
    """
    ctx = _f_ctx_from_request()
    if not ctx["fecha"] or not ctx["caja"]:
        return jsonify(success=True, datos=[])

    tipo = _f_norm(request.args.get("tipo") or (request.get_json(silent=True) or {}).get("tipo"))

    sql = [
        """
        SELECT f.id, f.local, f.caja, f.turno, f.fecha, f.tipo, f.punto_venta, f.nro_factura, f.comentario, f.monto, f.estado,
               f.usuario, f.created_at, f.updated_at, f.update_by
          FROM facturas_trns f
         WHERE f.local=%s AND f.caja=%s AND f.fecha=%s
        """
    ]
    args = [ctx["local"], ctx["caja"], ctx["fecha"]]

    if ctx["turno"]:
        sql.append("AND f.turno=%s")
        args.append(ctx["turno"])

    if tipo:
        sql.append("AND f.tipo=%s")
        args.append(tipo.upper())

    # Agregar filtro de visibilidad por rol
    sql.append(g.read_scope)

    sql.append("ORDER BY f.created_at ASC")
    query = " ".join(sql)

    conn = get_db_connection()
    try:
        with conn.cursor(dictionary=True) as cur:
            cur.execute(query, tuple(args))
            rows = cur.fetchall()
        return jsonify(success=True, datos=rows)
    finally:
        try:
            conn.close()
        except Exception:
            pass


@app.put("/facturas/<int:fid>")
@login_required
@page_access_required("index")
def facturas_update(fid: int):
    """
    Actualiza PV / nro_factura / comentario / monto / tipo / estado.
    Body JSON: cualquier subset de {tipo, punto_venta, nro_factura, comentario, monto, estado}
    Requiere can_edit ok para el contexto de esa factura.
    """
    from modules.tabla_auditoria import registrar_auditoria, obtener_registro_anterior

    # 1) Traer la fila para conocer su contexto original
    conn = get_db_connection()
    try:
        with conn.cursor(dictionary=True) as cur:
            cur.execute(
                "SELECT id, local, caja, turno, fecha FROM facturas_trns WHERE id=%s",
                (fid,)
            )
            base = cur.fetchone()
            if not base:
                return jsonify(success=False, msg="No encontrado"), 404

        # AUDITORÍA: Capturar estado anterior
        datos_anteriores = obtener_registro_anterior(conn, 'facturas_trns', fid)

        # 2) chequear permisos con el contexto de la fila (usamos turno normalizado solo para permiso)
        ctx = {
            "local": base["local"],
            "caja":  base["caja"],
            "fecha": str(base["fecha"]),
            "turno": base["turno"],
            "turno_canedit": _f_norm_turno_for_can_edit(base["turno"]),
        }
        ok, resp = _f_check_can_edit_or_409(conn, ctx)
        if not ok:
            return resp

        # 3) aplicar cambios
        data = _f_get_json()
        sets = []
        args = []
        datos_nuevos = {}

        if "tipo" in data and _f_norm(data.get("tipo")):
            sets.append("tipo=%s")
            val = _f_norm(data["tipo"]).upper()
            args.append(val)
            datos_nuevos['tipo'] = val
        if "punto_venta" in data and _f_norm(data.get("punto_venta")):
            sets.append("punto_venta=%s")
            val = _f_norm(data["punto_venta"])
            args.append(val)
            datos_nuevos['punto_venta'] = val
        if "nro_factura" in data and _f_norm(data.get("nro_factura")):
            sets.append("nro_factura=%s")
            val = _f_norm(data["nro_factura"])
            args.append(val)
            datos_nuevos['nro_factura'] = val
        if "comentario" in data:
            sets.append("comentario=%s")
            val = _f_norm(data.get("comentario") or "")
            args.append(val)
            datos_nuevos['comentario'] = val
        if "monto" in data:
            sets.append("monto=%s")
            val = _f_parse_monto(data.get("monto"))
            args.append(val)
            datos_nuevos['monto'] = val
        if "estado" in data and _f_norm(data.get("estado")):
            sets.append("estado=%s")
            val = _f_norm(data["estado"])
            args.append(val)
            datos_nuevos['estado'] = val

        usuario = session.get("username") or "system"
        sets.append("update_by=%s")
        args.append(usuario)
        datos_nuevos['update_by'] = usuario

        if len(sets) == 1:  # solo update_by => sin cambios de negocio
            return jsonify(success=True, msg="Sin cambios")

        args.append(fid)

        with conn.cursor() as cur2:
            cur2.execute(f"UPDATE facturas_trns SET {', '.join(sets)} WHERE id=%s", tuple(args))
            conn.commit()

        # AUDITORÍA: Registrar UPDATE
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='facturas_trns',
            registro_id=fid,
            datos_anteriores=datos_anteriores,
            datos_nuevos=datos_nuevos,
            descripcion=f"Factura {datos_anteriores.get('tipo')} actualizada - PV: {datos_anteriores.get('punto_venta')} / Nro: {datos_anteriores.get('nro_factura')}",
            contexto_override={
                'local': ctx["local"],
                'caja': ctx["caja"],
                'fecha_operacion': str(ctx["fecha"]),
                'turno': ctx["turno"]
            }
        )

        with conn.cursor(dictionary=True) as cur3:
            cur3.execute(
                """
                SELECT id, local, caja, turno, fecha, tipo, punto_venta, nro_factura, comentario, monto, estado,
                       usuario, created_at, updated_at, update_by
                  FROM facturas_trns WHERE id=%s
                """,
                (fid,)
            )
            row = cur3.fetchone()
        return jsonify(success=True, item=row)
    finally:
        try:
            conn.close()
        except Exception:
            pass


@app.delete("/facturas/<int:fid>")
@login_required
@page_access_required("index")
def facturas_delete(fid: int):
    """
    Borra una factura. Requiere can_edit ok con el contexto original de la fila.
    """
    from modules.tabla_auditoria import registrar_auditoria, obtener_registro_anterior

    conn = get_db_connection()
    try:
        with conn.cursor(dictionary=True) as cur:
            cur.execute(
                "SELECT id, local, caja, turno, fecha FROM facturas_trns WHERE id=%s",
                (fid,)
            )
            base = cur.fetchone()
            if not base:
                return jsonify(success=False, msg="No encontrado"), 404

        # AUDITORÍA: Capturar estado completo antes de borrar
        datos_anteriores = obtener_registro_anterior(conn, 'facturas_trns', fid)

        ctx = {
            "local": base["local"],
            "caja":  base["caja"],
            "fecha": str(base["fecha"]),
            "turno": base["turno"],
            "turno_canedit": _f_norm_turno_for_can_edit(base["turno"]),
        }
        ok, resp = _f_check_can_edit_or_409(conn, ctx)
        if not ok:
            return resp

        with conn.cursor() as cur2:
            cur2.execute("DELETE FROM facturas_trns WHERE id=%s", (fid,))
            deleted = cur2.rowcount
            conn.commit()

        # AUDITORÍA: Registrar DELETE
        if datos_anteriores:
            registrar_auditoria(
                conn=conn,
                accion='DELETE',
                tabla='facturas_trns',
                registro_id=fid,
                datos_anteriores=datos_anteriores,
                descripcion=f"Factura {datos_anteriores.get('tipo')} eliminada - PV: {datos_anteriores.get('punto_venta')} / Nro: {datos_anteriores.get('nro_factura')} / Monto: ${datos_anteriores.get('monto', 0):,.2f}",
                contexto_override={
                    'local': ctx["local"],
                    'caja': ctx["caja"],
                    'fecha_operacion': str(ctx["fecha"]),
                    'turno': ctx["turno"]
                }
            )

        if deleted == 0:
            return jsonify(success=False, msg="No encontrado"), 404
        return jsonify(success=True, deleted=deleted)
    finally:
        try:
            conn.close()
        except Exception:
            pass

# =================== FIN BLOQUE FACTURAS ===================




## _________________________________________________ CIERRE CAJA INDIVIDUAL _______________________________________
# ───────────────────────────────────────────
# /api/cierre/resumen  (caja/fecha únicos)
# ───────────────────────────────────────────
@app.route('/api/cierre/resumen')
@login_required
def cierre_resumen():
    local = get_local_param()
    caja  = request.args.get('caja')
    fecha = request.args.get('fecha')
    turno = request.args.get('turno') or 'UNI'  # <-- TURNO

    if not local or not caja or not fecha or not turno:
        return jsonify(error="Parámetros insuficientes"), 400

    # Control de visibilidad por rol (igual que facturas)
    lvl = get_user_level()
    conn = get_db_connection()

    # Nivel 2 (encargado): solo puede ver si la caja está cerrada
    if lvl == 2:
        cur_check = conn.cursor()
        cur_check.execute("""
            SELECT COUNT(*) FROM cajas_estado
            WHERE local=%s AND caja=%s AND fecha_operacion=%s AND turno=%s AND estado=0
        """, (local, caja, _normalize_fecha(fecha), turno))
        row = cur_check.fetchone()
        cur_check.close()
        if not row or row[0] == 0:
            # Caja no cerrada, retornar resumen vacío
            conn.close()
            return jsonify({k: 0.0 for k in ['venta_total','venta_z','facturas_a','facturas_b',
                'efectivo','tarjeta','mercadopago','rappi','pedidosya','gastos','cuenta_cte','tips','anticipos','discovery','total_cobrado']})

    # Nivel 3 (auditor): solo puede ver si el local está cerrado
    if lvl >= 3:
        cur_check = conn.cursor()
        cur_check.execute("""
            SELECT COUNT(*) FROM cierres_locales
            WHERE local=%s AND DATE(fecha)=%s AND estado=0
        """, (local, _normalize_fecha(fecha)))
        row = cur_check.fetchone()
        cur_check.close()
        if not row or row[0] == 0:
            # Local no cerrado, retornar resumen vacío
            conn.close()
            return jsonify({k: 0.0 for k in ['venta_total','venta_z','facturas_a','facturas_b',
                'efectivo','tarjeta','mercadopago','rappi','pedidosya','gastos','cuenta_cte','tips','anticipos','discovery','total_cobrado']})

    cur  = conn.cursor()

    resumen = {}

    try:
        # ===== DETERMINAR SI USAR SNAP TABLES =====
        # Si el local está cerrado, leer desde snap_* (igual que resumen_local)
        cur.execute("""
            SELECT COALESCE(MIN(estado), 1) AS min_estado
            FROM cierres_locales
            WHERE local=%s AND fecha=%s
        """, (local, _normalize_fecha(fecha)))
        row = cur.fetchone()
        local_cerrado = (row is not None and row[0] is not None and int(row[0]) == 0)

        # Determinar tablas a usar
        if local_cerrado:
            T_VENTAS = "snap_ventas"
            T_REMESAS = "snap_remesas"
            T_TARJETAS = "snap_tarjetas"
            T_MP = "snap_mercadopago"
            T_RAPPI = "snap_rappi"
            T_PEDIDOSYA = "snap_pedidosya"
            T_GASTOS = "snap_gastos"
            T_FACTURAS = "facturas_trns"  # Siempre facturas_trns (auditores editan post-cierre)
        else:
            T_VENTAS = "ventas_trns"
            T_REMESAS = "remesas_trns"
            T_TARJETAS = "tarjetas_trns"
            T_MP = "mercadopago_trns"
            T_RAPPI = "rappi_trns"
            T_PEDIDOSYA = "pedidosya_trns"
            T_GASTOS = "gastos_trns"
            T_FACTURAS = "facturas_trns"

        # ===== Ventas base (una fila por caja/fecha/turno) =====
        # venta_total
        cur.execute(f"""
            SELECT COALESCE(SUM(venta_total_sistema),0)
              FROM {T_VENTAS}
             WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        resumen['venta_total'] = float(row[0]) if row and row[0] is not None else 0.0

        # venta_z  (ahora desde facturas_trns con tipo='Z')
        cur.execute(f"""
            SELECT COALESCE(SUM(monto),0)
              FROM {T_FACTURAS}
             WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s AND tipo='Z'
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        resumen['venta_z'] = float(row[0]) if row and row[0] is not None else 0.0

        # facturas_a (informativo)
        cur.execute(f"""
            SELECT COALESCE(SUM(monto),0)
              FROM {T_FACTURAS}
             WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s AND tipo='A'
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        resumen['facturas_a'] = float(row[0]) if row and row[0] is not None else 0.0

        # facturas_b (informativo)
        cur.execute(f"""
            SELECT COALESCE(SUM(monto),0)
              FROM {T_FACTURAS}
             WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s AND tipo='B'
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        resumen['facturas_b'] = float(row[0]) if row and row[0] is not None else 0.0

        # DISCOVERY = venta_total - (Z + A + B)
        # Es la plata en negro (ventas no facturadas)
        total_facturas_zab = resumen['venta_z'] + resumen['facturas_a'] + resumen['facturas_b']
        resumen['discovery'] = float(resumen['venta_total'] - total_facturas_zab)

        # ===== Ingresos / medios de cobro =====
        # efectivo (remesas)
        cur.execute(f"""
            SELECT COALESCE(SUM(monto),0)
              FROM {T_REMESAS}
             WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        efectivo_base = float(row[0]) if row and row[0] is not None else 0.0

        # COMPENSACIÓN: Restar anticipos en efectivo consumidos en esta caja
        # (para evitar duplicación al crear la remesa)
        cur.execute("""
            SELECT COALESCE(SUM(aec.importe_consumido), 0)
            FROM anticipos_estados_caja aec
            JOIN anticipos_recibidos ar ON ar.id = aec.anticipo_id
            JOIN medios_anticipos ma ON ma.id = ar.medio_pago_id
            WHERE aec.fecha = %s
              AND aec.local = %s
              AND aec.caja = %s
              AND aec.turno = %s
              AND aec.estado = 'consumido'
              AND ma.es_efectivo = 1
        """, (fecha, local, caja, turno))
        row_compensacion = cur.fetchone()
        compensacion_efectivo = float(row_compensacion[0]) if row_compensacion and row_compensacion[0] is not None else 0.0

        # Efectivo final = efectivo_base - compensación
        resumen['efectivo'] = efectivo_base - compensacion_efectivo

        # tarjeta (ventas con tarjeta NO TIPS)
        cur.execute(f"""
            SELECT COALESCE(SUM(monto),0)
              FROM {T_TARJETAS}
             WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        resumen['tarjeta'] = float(row[0]) if row and row[0] is not None else 0.0

        # mercadopago normal
        cur.execute(f"""
            SELECT COALESCE(SUM(importe),0)
              FROM {T_MP}
             WHERE DATE(fecha)=%s AND local=%s AND tipo='NORMAL' AND caja=%s AND turno=%s
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        resumen['mercadopago'] = float(row[0]) if row and row[0] is not None else 0.0

        # cuenta corriente (facturas CC - suma como medio de cobro)
        cur.execute(f"""
            SELECT COALESCE(SUM(monto),0)
              FROM {T_FACTURAS}
             WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s AND tipo='CC'
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        resumen['cuenta_cte'] = float(row[0]) if row and row[0] is not None else 0.0

        # rappi
        cur.execute(f"""
            SELECT COALESCE(SUM(monto),0)
              FROM {T_RAPPI}
             WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        resumen['rappi'] = float(row[0]) if row and row[0] is not None else 0.0


        # GASTOS
        cur.execute(f"""
            SELECT COALESCE(SUM(monto),0)
              FROM {T_GASTOS}
             WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        resumen['gastos'] = float(row[0]) if row and row[0] is not None else 0.0


        # pedidosya
        cur.execute(f"""
            SELECT COALESCE(SUM(monto),0)
              FROM {T_PEDIDOSYA}
             WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        resumen['pedidosya'] = float(row[0]) if row and row[0] is not None else 0.0

        # ===== TIPS =====
        # Tips de tarjetas: ahora desde tarjetas_trns.monto_tip
        cur.execute(f"""
            SELECT COALESCE(SUM(monto_tip),0)
              FROM {T_TARJETAS}
             WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        tips_tarjeta = float(row[0]) if row and row[0] is not None else 0.0

        # Tips de MercadoPago (tipo TIP)
        cur.execute(f"""
            SELECT COALESCE(SUM(importe),0)
              FROM {T_MP}
             WHERE DATE(fecha)=%s AND local=%s AND tipo='TIP' AND caja=%s AND turno=%s
        """, (fecha, local, caja, turno))
        row = cur.fetchone()
        tips_mp = float(row[0]) if row and row[0] is not None else 0.0

        resumen['tips'] = tips_tarjeta + tips_mp

        # ===== ANTICIPOS CONSUMIDOS =====
        # Anticipos consumidos (señas usadas para justificar faltantes)
        # Nuevo sistema: usar anticipos_recibidos + anticipos_estados_caja
        cur.execute("""
            SELECT COALESCE(SUM(ar.importe),0)
              FROM anticipos_recibidos ar
              JOIN anticipos_estados_caja aec ON aec.anticipo_id = ar.id
             WHERE aec.local=%s AND aec.caja=%s AND DATE(aec.fecha)=%s AND aec.turno=%s
               AND aec.estado = 'consumido'
        """, (local, caja, fecha, turno))
        row = cur.fetchone()
        resumen['anticipos'] = float(row[0]) if row and row[0] is not None else 0.0

        # ===== Total cobrado (medios de cobro + gastos)
        # Las facturas A, B, Z son informativas y NO suman al cobrado
        # Las facturas CC sí suman porque son cuenta corriente (medio de cobro)
        # Los gastos SÍ suman al total cobrado (justifican la venta)
        # Los anticipos consumidos SÍ suman al total cobrado (justifican faltantes)
        resumen['total_cobrado'] = sum([
            resumen.get('efectivo',0.0),
            resumen.get('tarjeta',0.0),
            resumen.get('mercadopago',0.0),
            resumen.get('rappi',0.0),
            resumen.get('pedidosya',0.0),
            resumen.get('cuenta_cte',0.0),  # Cuenta corriente (facturas CC)
            resumen.get('gastos',0.0),      # Gastos justifican la venta
            resumen.get('anticipos',0.0),   # Anticipos consumidos justifican faltantes
        ])

        # ===== Estado de caja (por turno)
        try:
            cur.execute("""
                SELECT estado
                  FROM cajas_estado
                 WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s
                 ORDER BY id DESC
                 LIMIT 1
            """, (fecha, local, caja, turno))
            row = cur.fetchone()
            resumen['estado_caja'] = int(row[0]) if row and row[0] is not None else 1  # 1=abierta por defecto
        except Exception:
            resumen['estado_caja'] = 1

        return jsonify(resumen)

    except Exception as e:
        return jsonify(error=str(e)), 500
    finally:
        cur.close()
        conn.close()


#________________ ====== UTILIDADES DE CAJA ABIERTA/CERRADA ====== ________________________________________________________________________
from datetime import datetime
def ensure_estado_row(conn, local: str, caja: str, fecha: str, turno: str = "UNI"):
    """
    Garantiza que exista una fila (local,caja,fecha,turno). Si no existe, la crea como abierta (estado=1).
    """
    turno = (turno or "UNI").upper()
    cur = conn.cursor()
    cur.execute("""
        SELECT id FROM cajas_estado
        WHERE local=%s AND caja=%s AND fecha_operacion=%s AND turno=%s
    """, (local, caja, fecha, turno))
    row = cur.fetchone()
    if not row:
        cur.execute("""
            INSERT INTO cajas_estado (local, caja, fecha_operacion, turno, estado)
            VALUES (%s, %s, %s, %s, 1)
        """, (local, caja, fecha, turno))
        conn.commit()
    cur.close()

def caja_esta_cerrada(conn, local: str, caja: str, fecha: str, turno: str = "UNI") -> bool:
    """
    True si la caja está cerrada para (local,caja,fecha,turno).
    Si no hay fila -> se considera ABIERTA (False).
    """
    turno = (turno or "UNI").upper()
    cur = conn.cursor()
    cur.execute("""
        SELECT estado FROM cajas_estado
        WHERE local=%s AND caja=%s AND fecha_operacion=%s AND turno=%s
    """, (local, caja, fecha, turno))
    row = cur.fetchone()
    cur.close()
    if not row:
        return False
    return int(row[0]) == 0  # 0 = cerrada

def extract_context_from_request():
    """
    Extrae (local, caja, fecha, turno) de la request/sesión.
    """
    local = get_local_param()
    caja  = request.args.get('caja')
    fecha = request.args.get('fecha')
    turno = request.args.get('turno') or "UNI"

    if request.is_json:
        j = request.get_json(silent=True) or {}
        caja  = j.get('caja', caja)
        fecha = j.get('fecha', fecha)
        turno = j.get('turno', turno)
        local = j.get('local', local)

    if turno:
        turno = turno.upper()
    return local, caja, fecha, (turno or "UNI")

def abort_if_cerrada(conn, local, caja, fecha, turno="UNI"):
    """
    Lanza 409 si la caja está cerrada.
    """
    if not (local and caja and fecha):
        abort(400, description="Faltan parámetros de contexto (local/caja/fecha).")
    if caja_esta_cerrada(conn, local, caja, fecha, turno):
        abort(409, description="Caja cerrada: no se puede modificar.")

# ====== ENDPOINTS ======
@app.get('/estado_caja')
@login_required
def estado_caja():
    local, caja, fecha, turno = extract_context_from_request()
    if not (local and caja and fecha):
        return jsonify({"error": "Faltan parámetros local/caja/fecha"}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT estado FROM cajas_estado
        WHERE local=%s AND caja=%s AND fecha_operacion=%s AND turno=%s
    """, (local, caja, fecha, turno))
    row = cur.fetchone()
    cur.close(); conn.close()

    estado = 1 if (not row) else int(row[0])  # abierta por defecto
    return jsonify({"estado": estado})

## meto helpers a la func de cierre_caja
def _normalize_fecha(fecha):
    # tu impl. actual (la dejo como está)
    return fecha

def ensure_estado_row(conn, local, caja, fecha, turno):
    """Crea la fila (abierta) si no existe para ese local/caja/fecha/turno."""
    cur = conn.cursor()
    cur.execute("""
        INSERT IGNORE INTO cajas_estado (local, caja, turno, fecha_operacion, estado)
        VALUES (%s, %s, %s, %s, 1)
    """, (local, caja, turno, _normalize_fecha(fecha)))
    conn.commit()
    cur.close()

def is_caja_abierta(conn, local, caja, fecha, turno):
    """True si no hay fila o si estado=1 (abierta)."""
    cur = conn.cursor()
    cur.execute("""
        SELECT estado
        FROM cajas_estado
        WHERE local=%s AND caja=%s AND fecha_operacion=%s AND turno=%s
        LIMIT 1
    """, (local, caja, _normalize_fecha(fecha), turno))
    row = cur.fetchone()
    cur.close()
    if row is None:
        return True
    return int(row[0]) == 1
## ahora si la func
@app.post('/cerrar_caja')
@login_required
def cerrar_caja():
    data = request.get_json() or {}
    local   = data.get('local')  or session.get('local')
    caja    = data.get('caja')
    fecha   = data.get('fecha')
    turno   = data.get('turno')            # ← obligatorio
    observacion = data.get('observacion', '')  # ← observación/descargo opcional
    usuario = session.get('username') or data.get('usuario') or 'sistema'

    if not (local and caja and fecha and turno):
        return jsonify(ok=False, msg="Faltan parámetros local/caja/fecha/turno"), 400

    conn = get_db_connection()
    try:
        # asegurar fila abierta
        ensure_estado_row(conn, local, caja, fecha, turno)

        cur = conn.cursor()
        cur.execute("""
            SELECT id, estado
            FROM cajas_estado
            WHERE local=%s AND caja=%s AND fecha_operacion=%s AND turno=%s
            FOR UPDATE
        """, (local, caja, _normalize_fecha(fecha), turno))
        row = cur.fetchone()
        if not row:
            conn.rollback(); cur.close(); conn.close()
            return jsonify(ok=False, msg="No se encontró estado"), 404

        _id, estado = row
        if int(estado) == 0:
            cur.close(); conn.close()
            return jsonify(ok=True, already_closed=True)

        # Actualizar estado con observación
        cur.execute("""
            UPDATE cajas_estado
            SET estado=0, cerrada_en=NOW(), cerrada_por=%s, observacion=%s
            WHERE id=%s
        """, (usuario, observacion, _id))
        conn.commit()

        # Crear snapshot de la caja cerrada
        try:
            create_snapshot_for_local(conn, local, _normalize_fecha(fecha), turno, made_by=usuario)
            conn.commit()
        except Exception as snap_err:
            print(f"⚠️ Error creando snapshot: {snap_err}")
            # No rollback del cierre, solo advertencia

        cur.close(); conn.close()
        return jsonify(ok=True, closed=True)

    except Exception as e:
        try: conn.rollback()
        except: pass
        try: conn.close()
        except: pass
        # loggear e
        return jsonify(ok=False, msg=str(e)), 500


##__________________________ GASTOS PESTAÑA _____________________________-#
# __________________________________ GASTOS _______________________________________
# CRUD + lote, con soporte de turno y bloqueo por estado de caja

@app.route('/api/tipos_gastos')
@login_required
def api_tipos_gastos():
    """Retorna los tipos de gastos normalizados desde la tabla tipos_gastos"""
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT codigo, descripcion
            FROM tipos_gastos
            WHERE activo = 1
            ORDER BY orden, descripcion
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        # Convertir a diccionario {codigo: descripcion}
        tipos = {row['codigo']: row['descripcion'] for row in rows}
        return jsonify(success=True, tipos=tipos)
    except Exception as e:
        print(f"❌ ERROR api_tipos_gastos: {e}")
        return jsonify(success=False, msg=str(e)), 500


@app.route('/gastos_cargadas')
@login_required
def gastos_cargadas():
    local = get_local_param()
    caja  = request.args.get('caja')
    fecha = request.args.get('fecha')
    turno = request.args.get('turno')

    if not (local and caja and fecha and turno):
        # Por compatibilidad con el front, devolvemos flags de estado
        return jsonify(success=True, datos=[], estado_caja=1, estado_local=1)

    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)

        # ===== Estados para la UI =====
        try:
            # abierta => 1; cerrada => 0
            estado_caja = 1 if is_caja_abierta(conn, local, caja, _normalize_fecha(fecha), turno) else 0
        except Exception:
            estado_caja = 1

        try:
            # usamos tu is_local_closed; si está cerrado -> 0; abierto -> 1
            estado_local = 0 if is_local_closed(conn, local, _normalize_fecha(fecha)) else 1
        except Exception:
            estado_local = 1

        cur.execute("""
            SELECT id, fecha, local, caja, turno, tipo, monto, observaciones, created_at
            FROM gastos_trns
            WHERE local=%s AND caja=%s AND DATE(fecha)=%s AND turno=%s
            ORDER BY id ASC
        """, (local, caja, _normalize_fecha(fecha), turno))
        rows = cur.fetchall()

        cur.close(); conn.close()
        return jsonify(success=True, datos=rows, estado_caja=estado_caja, estado_local=estado_local)

    except Exception as e:
        print("❌ ERROR gastos_cargadas:", e)
        return jsonify(success=False, msg=str(e)), 500


@app.route('/guardar_gastos_lote', methods=['POST'])
@login_required
def guardar_gastos_lote():
    payload = request.get_json() or {}
    caja   = payload.get('caja')
    fecha  = payload.get('fecha')
    turno  = payload.get('turno')
    items  = payload.get('transacciones', []) or []

    # Para auditores: usar local del payload o del request, sino usar get_local_param()
    local   = payload.get('local') or get_local_param()
    usuario = session.get('username') or 'sistema'

    if not (local and caja and fecha and turno):
        return jsonify(success=False, msg="Faltan local/caja/fecha/turno"), 400
    if not items:
        return jsonify(success=False, msg="No se recibieron gastos"), 400

    try:
        conn = get_db_connection()

        user_level = get_user_level()
        fecha_norm = _normalize_fecha(fecha)

        # Verificar si está auditado primero para dar mensaje más claro
        if is_local_auditado(conn, local, fecha_norm):
            conn.close()
            return jsonify(success=False, msg="❌ El local está AUDITADO para esta fecha. No se pueden realizar más modificaciones."), 403

        if not can_edit(conn, local, caja, turno, fecha_norm, user_level):
            conn.close()
            return jsonify(success=False, msg="No tenés permisos para guardar (caja/local cerrados para tu rol)."), 409

        cur = conn.cursor()
        for g in items:
            tipo = (g.get('tipo') or "").strip()
            obs  = (g.get('observaciones') or "").strip()

            # monto puede venir con separadores locales
            mstr = str(g.get('monto', '0')).replace('.', '').replace(',', '.')
            try:
                monto = float(mstr or 0)
            except Exception:
                monto = 0.0

            cur.execute("""
                INSERT INTO gastos_trns
                    (fecha, local, caja, turno, tipo, monto, observaciones, usuario, created_at, estado)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), %s)
            """, (_normalize_fecha(fecha), local, caja, turno, tipo, monto, obs, usuario, "revision"))

        conn.commit()
        cur.close(); conn.close()
        return jsonify(success=True)

    except Exception as e:
        print("❌ ERROR guardar_gastos_lote:", e)
        return jsonify(success=False, msg=str(e)), 500


@app.route('/gastos/<int:gasto_id>', methods=['PUT'])
@login_required
def actualizar_gasto(gasto_id):
    data  = request.get_json() or {}
    tipo  = (data.get('tipo') or "").strip()
    # Observaciones puede NO venir desde el lápiz, en ese caso conservamos la actual
    obs_in = data.get('observaciones')  # puede ser None
    monto = data.get('monto')

    # normalizo monto
    try:
        if isinstance(monto, str):
            monto = float(monto.replace('.', '').replace(',', '.'))
        else:
            monto = float(monto or 0)
    except Exception:
        return jsonify(success=False, msg="Monto inválido"), 400

    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)

        # 1. Obtener registro anterior para auditoría
        datos_anteriores = obtener_registro_anterior(conn, 'gastos_trns', gasto_id)

        cur.execute("""
            SELECT id, local, caja, fecha, turno, observaciones
            FROM gastos_trns
            WHERE id=%s
        """, (gasto_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify(success=False, msg="Registro no encontrado"), 404

        user_level = get_user_level()

        # Verificar si está auditado primero para dar mensaje más claro
        if is_local_auditado(conn, row['local'], row['fecha']):
            cur.close(); conn.close()
            return jsonify(success=False, msg="❌ El local está AUDITADO para esta fecha. No se pueden realizar más modificaciones."), 403

        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], user_level):
            cur.close(); conn.close()
            return jsonify(success=False, msg="No tenés permisos para actualizar (caja/local cerrados para tu rol)."), 409

        final_obs = (obs_in if obs_in is not None else (row.get('observaciones') or "")).strip()

        cur2 = conn.cursor()
        cur2.execute("""
            UPDATE gastos_trns
               SET tipo=%s, monto=%s, observaciones=%s
             WHERE id=%s
        """, (tipo, monto, final_obs, gasto_id))

        conn.commit()

        # 2. Registrar auditoría
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='gastos_trns',
            registro_id=gasto_id,
            datos_anteriores=datos_anteriores,
            datos_nuevos={'tipo': tipo, 'monto': monto, 'observaciones': final_obs},
            contexto_override={'local': row['local'], 'caja': row['caja'], 'fecha_operacion': row['fecha'], 'turno': row['turno']}
        )

        cur2.close(); cur.close(); conn.close()
        return jsonify(success=True)

    except Exception as e:
        print("❌ ERROR actualizar_gasto:", e)
        return jsonify(success=False, msg=str(e)), 500


@app.route('/gastos/<int:gasto_id>', methods=['DELETE'])
@login_required
def borrar_gasto(gasto_id):
    try:
        conn = get_db_connection()
        cur  = conn.cursor(dictionary=True)

        # 1. Obtener registro anterior para auditoría
        datos_anteriores = obtener_registro_anterior(conn, 'gastos_trns', gasto_id)

        cur.execute("""
            SELECT id, local, caja, fecha, turno
            FROM gastos_trns
            WHERE id=%s
        """, (gasto_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify(success=False, msg="Registro no encontrado"), 404

        user_level = get_user_level()

        # Verificar si está auditado primero para dar mensaje más claro
        if is_local_auditado(conn, row['local'], row['fecha']):
            cur.close(); conn.close()
            return jsonify(success=False, msg="❌ El local está AUDITADO para esta fecha. No se pueden realizar más modificaciones."), 403

        if not can_edit(conn, row['local'], row['caja'], row['turno'], row['fecha'], user_level):
            cur.close(); conn.close()
            return jsonify(success=False, msg="No tenés permisos para borrar (caja/local cerrados para tu rol)."), 409

        cur2 = conn.cursor()
        cur2.execute("DELETE FROM gastos_trns WHERE id=%s", (gasto_id,))
        conn.commit()
        deleted = cur2.rowcount

        # 2. Registrar auditoría
        registrar_auditoria(
            conn=conn,
            accion='DELETE',
            tabla='gastos_trns',
            registro_id=gasto_id,
            datos_anteriores=datos_anteriores,
            contexto_override={'local': row['local'], 'caja': row['caja'], 'fecha_operacion': row['fecha'], 'turno': row['turno']}
        )

        cur2.close(); cur.close(); conn.close()
        return jsonify(success=True, deleted=deleted)

    except Exception as e:
        print("❌ ERROR borrar_gasto:", e)
        return jsonify(success=False, msg=str(e)), 500























# ========= API LOCALES =========
@app.route('/api/locales', methods=['GET'])
@login_required
def api_locales():
    """
    Endpoint para obtener la lista de todos los locales.
    Usado por auditores para filtrar en Carga de Datos.
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT DISTINCT local FROM locales ORDER BY local")
        locales = cur.fetchall()
        resultado = [l['local'] for l in locales if l['local']]
        print(f"📍 API /api/locales devolviendo: {resultado}")
        return jsonify(locales=resultado)
    except Exception as e:
        print(f"❌ Error en /api/locales: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(error=str(e)), 500
    finally:
        cur.close()
        conn.close()


# ========= API CAJAS POR LOCAL =========
@app.route('/api/cajas', methods=['GET'])
@login_required
def api_cajas():
    """
    Endpoint para obtener las cajas de un local específico.
    Usado por auditores para filtrar en Carga de Datos.
    Query param: local
    """
    local = request.args.get('local')
    if not local:
        return jsonify([])

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT DISTINCT caja
            FROM (
                SELECT DISTINCT caja FROM ventas_trns WHERE local = %s
                UNION
                SELECT DISTINCT caja FROM remesas_trns WHERE local = %s
                UNION
                SELECT DISTINCT caja FROM tarjetas_trns WHERE local = %s
            ) AS todas_cajas
            ORDER BY caja
        """, (local, local, local))
        cajas = cur.fetchall()
        return jsonify([c['caja'] for c in cajas if c['caja']])
    except Exception as e:
        return jsonify(error=str(e)), 500
    finally:
        cur.close()
        conn.close()


# ========= API TURNOS POR LOCAL =========
@app.route('/api/turnos', methods=['GET'])
@login_required
def api_turnos():
    """
    Endpoint para obtener los turnos de un local específico.
    Usado por auditores para filtrar en Carga de Datos.
    Query param: local
    """
    local = request.args.get('local')
    if not local:
        return jsonify([])

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Usar la MISMA lógica que index() y encargado() (líneas 541-542)
        # Esta query devuelve TODAS las filas donde local coincide
        # Si hay 2 filas para "Alma Cerrito" (una con "Turno día" y otra con "Turno noche"), devuelve ambas
        cur.execute("SELECT turnos FROM locales WHERE local = %s", (local,))
        turnos = [r[0] for r in cur.fetchall() if r[0]] or ['UNI']

        print(f"[/api/turnos] Local='{local}', Turnos={turnos}")
        return jsonify(turnos)
    except Exception as e:
        print(f"❌ Error en /api/turnos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(['UNI'])
    finally:
        cur.close()
        conn.close()


# ___________ RESUMEN_LOCAL.HTML _____________________________-#
# ========= Vista =========
@app.route("/resumen-local")
@login_required
def resumen_local():
    # Variables mínimas para la plantilla
    session.setdefault("username", session.get("username", "USUARIO"))
    session.setdefault("local",    session.get("local",    "Mi Local"))
    session.setdefault("society",  session.get("society",  "Mi Sociedad SA"))
    return render_template("resumen_local.html")


@app.route("/auditoria")
@login_required
@role_min_required(3)  # Solo auditor
def auditoria():
    """
    Página de carga masiva para auditoría.
    Recibe local y fecha como query params desde resumen_local.
    Los filtros están deshabilitados - solo se usan los parámetros de URL.
    """
    import sys
    local = request.args.get('local') or session.get('local')
    fecha = request.args.get('fecha')

    print(f"🔍 /auditoria - Parámetros recibidos: local={local}, fecha={fecha}", file=sys.stderr, flush=True)
    print(f"🔍 /auditoria - Session: {session.get('local')}", file=sys.stderr, flush=True)
    print(f"🔍 /auditoria - Query args: {dict(request.args)}", file=sys.stderr, flush=True)

    # Si no hay parámetros, usar valores de sesión/defaults
    if not local:
        local = session.get('local', '')
    if not fecha:
        from datetime import date
        fecha = date.today().isoformat()

    print(f"🔍 /auditoria - Valores finales: local='{local}', fecha='{fecha}'", file=sys.stderr, flush=True)
    print(f"🔍 /auditoria - Renderizando template", file=sys.stderr, flush=True)

    return render_template("auditor.html", local=local, fecha=fecha)


# ========= Helpers =========
from flask import request, jsonify, session, make_response

def _qsum(cur, sql, params):
    """Suma segura: devuelve float siempre."""
    cur.execute(sql, params)
    row = cur.fetchone()
    return float(row[0]) if row and row[0] is not None else 0.0

def _leftpad(num, width):
    try:
        s = str(int(num))
    except Exception:
        s = str(num or "0")
    return s.zfill(width)

def _format_z(pv, nz):
    return f"{_leftpad(pv,5)}-{_leftpad(nz,8)}"

def _fetch_ventas_z(conn, cur, fecha, local):
    """
    Devuelve (items, total):
      items = [{'pv': '5', 'punto_venta': '5', 'numero_z': '1691', 'monto': 1000.0, 'z_display': '00005-00001691'}, ...]
      total = suma de los montos
    (Versión legacy sin nombre de tabla; usa facturas_trns con tipo='Z')
    """
    return _fetch_ventas_z_dyn(cur, fecha, local, "facturas_trns", tipo='Z')

def _sum_cuenta_corriente(conn, cur, fecha, local):
    """Total de Cuenta Corriente, tolerante a diferencias de columnas."""
    try:
        return _qsum(
            cur,
            "SELECT COALESCE(SUM(monto),0) FROM cuenta_corriente_trns WHERE DATE(fecha)=%s AND local=%s",
            (fecha, local),
        )
    except Exception:
        conn.rollback()
        try:
            return _qsum(
                cur,
                "SELECT COALESCE(SUM(importe),0) FROM cuenta_corriente_trns WHERE DATE(fecha)=%s AND local=%s",
                (fecha, local),
            )
        except Exception:
            conn.rollback()
            return 0.0

def _sum_tips_tarjetas_breakdown(cur, table_name, fecha, local):
    """
    Devuelve dict con tips por marca desde tarjetas_trns (usando columna monto_tip).
    Suma por tarjeta (marca normalizada).
    """
    marcas = [
        "VISA", "VISA DEBITO", "VISA PREPAGO",
        "MASTERCARD", "MASTERCARD DEBITO", "MASTERCARD PREPAGO",
        "CABAL", "CABAL DEBITO",
        "AMEX", "MAESTRO",
        "NARANJA", "DECIDIR", "DINERS",
        "PAGOS INMEDIATOS",
        "MAS DELIVERY"
    ]

    breakdown = {m: 0.0 for m in marcas}

    try:
        for marca in marcas:
            cur.execute(
                f"""
                SELECT COALESCE(SUM(monto_tip),0)
                FROM {table_name}
                WHERE DATE(fecha)=%s AND local=%s AND UPPER(tarjeta)=%s
                """,
                (fecha, local, marca.upper())
            )
            row = cur.fetchone()
            breakdown[marca] = float(row[0] or 0.0) if row else 0.0
    except Exception:
        # Si falla, dejamos todo en 0
        pass

    total_tarjetas = float(sum(breakdown.values()))
    return breakdown, total_tarjetas


def _get_diferencias_detalle(cur, fecha, local, usar_snap=False):
    """
    Calcula las diferencias por caja/turno para mostrar en el acordeón.
    Solo devuelve las cajas/turnos donde diferencia != 0.

    Retorna: lista de dicts [{'caja': 'Caja 1', 'turno': 'MAÑANA', 'diferencia': -100.50, 'descargo': 'Faltó...'}]
    """
    f = _normalize_fecha(fecha)
    resultado = []

    # Primero, obtener todas las cajas/turnos del local ese día
    if usar_snap:
        # Si está cerrado, usar snap_*
        cur.execute("""
            SELECT DISTINCT caja, turno
            FROM (
                SELECT caja, turno FROM snap_ventas WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM snap_remesas WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM snap_tarjetas WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM snap_mercadopago WHERE local=%s AND DATE(fecha)=%s
            ) AS todas
            ORDER BY caja, turno
        """, (local, f) * 4)
    else:
        # Si está abierto, usar tablas normales
        cur.execute("""
            SELECT DISTINCT caja, turno
            FROM (
                SELECT caja, turno FROM ventas_trns WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM remesas_trns WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM tarjetas_trns WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM mercadopago_trns WHERE local=%s AND DATE(fecha)=%s
            ) AS todas
            ORDER BY caja, turno
        """, (local, f, local, f, local, f, local, f))

    cajas_turnos = cur.fetchall()

    for ct in cajas_turnos:
        caja = ct[0]
        turno = ct[1]

        # Calcular venta_total para esta caja/turno
        if usar_snap:
            cur.execute("""
                SELECT COALESCE(SUM(venta_total_sistema),0)
                FROM snap_ventas
                WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s
            """, (f, local, caja, turno))
        else:
            cur.execute("""
                SELECT COALESCE(SUM(venta_total_sistema),0)
                FROM ventas_trns
                WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s
            """, (f, local, caja, turno))

        row = cur.fetchone()
        venta_total = float(row[0] or 0.0) if row else 0.0

        # Calcular total_cobrado (efectivo + tarjeta + mp + rappi + pedidosya + cuenta_cte + gastos)
        total_cobrado = 0.0

        # Efectivo (remesas)
        if usar_snap:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM snap_remesas WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s", (f, local, caja, turno))
        else:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM remesas_trns WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s", (f, local, caja, turno))
        row = cur.fetchone()
        total_cobrado += float(row[0] or 0.0) if row else 0.0

        # Tarjetas
        if usar_snap:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM snap_tarjetas WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s", (f, local, caja, turno))
        else:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM tarjetas_trns WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s", (f, local, caja, turno))
        row = cur.fetchone()
        total_cobrado += float(row[0] or 0.0) if row else 0.0

        # MercadoPago (NORMAL)
        if usar_snap:
            cur.execute("SELECT COALESCE(SUM(importe),0) FROM snap_mercadopago WHERE DATE(fecha)=%s AND local=%s AND tipo='NORMAL' AND caja=%s AND turno=%s", (f, local, caja, turno))
        else:
            cur.execute("SELECT COALESCE(SUM(importe),0) FROM mercadopago_trns WHERE DATE(fecha)=%s AND local=%s AND tipo='NORMAL' AND caja=%s AND turno=%s", (f, local, caja, turno))
        row = cur.fetchone()
        total_cobrado += float(row[0] or 0.0) if row else 0.0

        # Rappi
        if usar_snap:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM snap_rappi WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s", (f, local, caja, turno))
        else:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM rappi_trns WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s", (f, local, caja, turno))
        row = cur.fetchone()
        total_cobrado += float(row[0] or 0.0) if row else 0.0

        # PedidosYa
        if usar_snap:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM snap_pedidosya WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s", (f, local, caja, turno))
        else:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM pedidosya_trns WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s", (f, local, caja, turno))
        row = cur.fetchone()
        total_cobrado += float(row[0] or 0.0) if row else 0.0

        # Cuenta corriente (facturas CC)
        if usar_snap:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM snap_facturas WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s AND tipo='CC'", (f, local, caja, turno))
        else:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM facturas_trns WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s AND tipo='CC'", (f, local, caja, turno))
        row = cur.fetchone()
        total_cobrado += float(row[0] or 0.0) if row else 0.0

        # Gastos
        if usar_snap:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM snap_gastos WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s", (f, local, caja, turno))
        else:
            cur.execute("SELECT COALESCE(SUM(monto),0) FROM gastos_trns WHERE DATE(fecha)=%s AND local=%s AND caja=%s AND turno=%s", (f, local, caja, turno))
        row = cur.fetchone()
        total_cobrado += float(row[0] or 0.0) if row else 0.0

        # Anticipos Consumidos - SIEMPRE consultar de tablas normales
        # (no hay snap_anticipos todavía, se agregará en futuras versiones)
        # Nuevo sistema: usar anticipos_recibidos + anticipos_estados_caja
        cur.execute("""
            SELECT COALESCE(SUM(ar.importe),0)
            FROM anticipos_recibidos ar
            JOIN anticipos_estados_caja aec ON aec.anticipo_id = ar.id
            WHERE aec.local=%s AND aec.caja=%s AND DATE(aec.fecha)=%s AND aec.turno=%s
              AND aec.estado = 'consumido'
        """, (local, caja, f, turno))
        row = cur.fetchone()
        total_cobrado += float(row[0] or 0.0) if row else 0.0

        # Calcular diferencia
        diferencia = total_cobrado - venta_total

        # Solo incluir si diferencia != 0
        if abs(diferencia) > 0.01:  # Tolerancia de 1 centavo
            # Buscar descargo desde cajas_estado
            cur.execute("""
                SELECT observacion
                FROM cajas_estado
                WHERE local=%s AND caja=%s AND turno=%s AND DATE(fecha_operacion)=%s
                ORDER BY id DESC
                LIMIT 1
            """, (local, caja, turno, f))
            row_obs = cur.fetchone()
            descargo = (row_obs[0] or '').strip() if row_obs else ''

            resultado.append({
                'caja': caja,
                'turno': turno,
                'diferencia': round(diferencia, 2),
                'descargo': descargo
            })

    return resultado


# ========= API GLOBAL =========
@app.route('/api/resumen_local', methods=['GET'])
@login_required
def api_resumen_local():
    """
    Resumen global por local+fecha (suma todas las cajas/turnos).
    Si el local está CERRADO ese día, lee de SNAPSHOTS (tablas snap_*).

    Parámetro adicional 'source' (solo para nivel 3):
    - 'operacion': Forzar lectura de tablas snap_* (datos de operación/cierre nivel 2)
    - 'admin': Forzar lectura de tablas normales (datos de administración/auditoría)
    Si no se especifica 'source', usa la lógica por defecto (cerrado=snap, abierto=normal)
    """
    local = (request.args.get('local') or session.get('local') or '').strip()
    fecha_s = (request.args.get('fecha') or '').strip()
    source = (request.args.get('source') or '').strip().lower()  # 'operacion' o 'admin'

    if not local or not fecha_s:
        return jsonify(error="Parámetros insuficientes: fecha y local son requeridos"), 400

    f = _normalize_fecha(fecha_s)
    if not f:
        return jsonify(error="Fecha inválida (formato esperado YYYY-MM-DD)"), 400

    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        # Verificar si el local está cerrado (siempre necesitamos este dato para el payload)
        cur.execute("""
            SELECT COALESCE(MIN(estado), 1) AS min_estado
            FROM cierres_locales
            WHERE local=%s AND fecha=%s
        """, (local, f))
        row = cur.fetchone()
        local_cerrado = (row is not None and row[0] is not None and int(row[0]) == 0)

        # Determinar si usar tablas snap o normales
        usar_snap = False

        if source:
            # Si se especifica 'source', usar ese valor (para niveles 2 y 3)
            usar_snap = (source == 'operacion')
        else:
            # Lógica por defecto: si el local está cerrado, usar snapshots
            usar_snap = local_cerrado

        # Tablas según el flag usar_snap
        # IMPORTANTE: Facturas siempre se leen de facturas_trns porque los auditores
        # pueden agregar/editar facturas incluso después del cierre del local
        if usar_snap:
            T_REMESAS    = "snap_remesas"
            T_TARJETAS   = "snap_tarjetas"
            T_MP         = "snap_mercadopago"
            T_RAPPI      = "snap_rappi"
            T_PEDIDOSYA  = "snap_pedidosya"
            T_GASTOS     = "snap_gastos"
            T_VENTAS     = "snap_ventas"
            T_FACTURAS   = "facturas_trns"  # SIEMPRE leer de facturas_trns (auditores editan post-cierre)
        else:
            T_REMESAS    = "remesas_trns"
            T_TARJETAS   = "tarjetas_trns"
            T_MP         = "mercadopago_trns"
            T_RAPPI      = "rappi_trns"
            T_PEDIDOSYA  = "pedidosya_trns"
            T_GASTOS     = "gastos_trns"
            T_VENTAS     = "ventas_trns"
            T_FACTURAS   = "facturas_trns"  # Contiene Z, A, B, CC

        # ===== RESUMEN: venta total, facturas y discovery =====
        venta_total = _qsum(
            cur,
            f"SELECT COALESCE(SUM(venta_total_sistema),0) FROM {T_VENTAS} WHERE DATE(fecha)=%s AND local=%s",
            (f, local),
        )

        # Breakdown Z por PV + Número (tabla facturas con tipo='Z')
        z_items, vta_z_total = _fetch_ventas_z_dyn(cur, f, local, T_FACTURAS, tipo='Z')

        # Total de todas las facturas (Z + A + B + CC)
        total_facturas = _qsum(
            cur,
            f"""SELECT COALESCE(SUM(monto),0)
                FROM {T_FACTURAS}
                WHERE DATE(fecha)=%s AND local=%s AND tipo IN ('Z','A','B','CC')""",
            (f, local),
        )

        # Discovery = Venta Total - (Z + A + B + CC)
        discovery = max((venta_total or 0.0) - (total_facturas or 0.0), 0.0)

        # ===== EFECTIVO (Remesas) =====
        efectivo_remesas = _qsum(
            cur,
            f"SELECT COALESCE(SUM(monto),0) FROM {T_REMESAS} WHERE DATE(fecha)=%s AND local=%s",
            (f, local),
        )

        # COMPENSACIÓN: Restar anticipos en efectivo consumidos en este local/fecha
        compensacion_efectivo_local = _qsum(
            cur,
            """
            SELECT COALESCE(SUM(aec.importe_consumido), 0)
            FROM anticipos_estados_caja aec
            JOIN anticipos_recibidos ar ON ar.id = aec.anticipo_id
            JOIN medios_anticipos ma ON ma.id = ar.medio_pago_id
            WHERE aec.fecha = %s
              AND aec.local = %s
              AND aec.estado = 'consumido'
              AND ma.es_efectivo = 1
            """,
            (f, local)
        ) or 0.0

        efectivo_neto = efectivo_remesas - compensacion_efectivo_local

        # ===== TARJETAS (ventas por marca) =====
        marcas = [
            "VISA", "VISA DEBITO", "VISA PREPAGO",
            "MASTERCARD", "MASTERCARD DEBITO", "MASTERCARD PREPAGO",
            "CABAL", "CABAL DEBITO",
            "AMEX", "MAESTRO",
            "NARANJA", "DECIDIR", "DINERS",
            "PAGOS INMEDIATOS",
            "MAS DELIVERY"
        ]
        def _sum_tarjeta_from(table_name, marca):
            return _qsum(
                cur,
                f"""
                SELECT COALESCE(SUM(monto),0)
                FROM {table_name}
                WHERE DATE(fecha)=%s AND local=%s AND UPPER(tarjeta)=%s
                """,
                (f, local, marca.upper()),
            ) or 0.0

        tarjetas_det = {m: _sum_tarjeta_from(T_TARJETAS, m) for m in marcas}
        tarjeta_total = float(sum(tarjetas_det.values()))

        # ===== MP, Rappi, PedidosYa =====
        mp_total = _qsum(
            cur,
            f"SELECT COALESCE(SUM(importe),0) FROM {T_MP} WHERE DATE(fecha)=%s AND local=%s AND UPPER(tipo)='NORMAL'",
            (f, local),
        )
        tips_mp = _qsum(
            cur,
            f"SELECT COALESCE(SUM(importe),0) FROM {T_MP} WHERE DATE(fecha)=%s AND local=%s AND UPPER(tipo)='TIP'",
            (f, local),
        )
        rappi_total = _qsum(
            cur,
            f"SELECT COALESCE(SUM(monto),0) FROM {T_RAPPI} WHERE DATE(fecha)=%s AND local=%s",
            (f, local),
        )
        pedidosya_total = _qsum(
            cur,
            f"SELECT COALESCE(SUM(monto),0) FROM {T_PEDIDOSYA} WHERE DATE(fecha)=%s AND local=%s",
            (f, local),
        )

        # ===== GASTOS (total + detalle por tipo) =====
        gastos_total = _qsum(
            cur,
            f"SELECT COALESCE(SUM(monto),0) FROM {T_GASTOS} WHERE DATE(fecha)=%s AND local=%s",
            (f, local),
        )

        # Detalle de gastos por tipo (con JOIN a tipos_gastos)
        gastos_detalle = {}
        try:
            cur.execute(
                f"""
                SELECT g.tipo, COALESCE(tg.descripcion, g.tipo) as descripcion, SUM(g.monto) as total
                FROM {T_GASTOS} g
                LEFT JOIN tipos_gastos tg ON g.tipo = tg.codigo
                WHERE DATE(g.fecha)=%s AND g.local=%s
                GROUP BY g.tipo, tg.descripcion
                ORDER BY tg.orden, g.tipo
                """,
                (f, local)
            )
            rows = cur.fetchall() or []
            for r in rows:
                tipo_codigo = r[0]
                tipo_desc = r[1]
                monto_tipo = float(r[2] or 0)
                gastos_detalle[tipo_codigo] = {
                    "descripcion": tipo_desc,
                    "monto": monto_tipo
                }
        except Exception as e:
            print(f"⚠️ Error obteniendo detalle de gastos: {e}")
            gastos_detalle = {}

        # ===== ANTICIPOS CONSUMIDOS (total + detalle) =====
        # IMPORTANTE: Los anticipos SIEMPRE se consultan de tablas normales
        # (no hay snap_anticipos todavía, se agregará en futuras versiones)
        anticipos_total = 0.0
        anticipos_items = []
        # Removido el check de usar_snap - siempre consultar de tablas normales
        if True:  # Anticipos solo existen en tablas normales (no hay snap todavía)
            # Total - Nuevo sistema: usar anticipos_recibidos + anticipos_estados_caja
            anticipos_total = _qsum(
                cur,
                """SELECT COALESCE(SUM(ar.importe),0)
                   FROM anticipos_recibidos ar
                   JOIN anticipos_estados_caja aec ON aec.anticipo_id = ar.id
                   WHERE aec.local=%s AND DATE(aec.fecha)=%s
                     AND aec.estado = 'consumido'""",
                (local, f),
            ) or 0.0

            print(f"🔍 DEBUG Anticipos - Local: {local}, Fecha: {f}, Total: {anticipos_total}, usar_snap: {usar_snap}")

            # Detalle (items individuales) - Nuevo sistema
            try:
                cur.execute("""
                    SELECT ar.id, ar.fecha_pago, ar.medio_pago, ar.observaciones, ar.importe, aec.usuario, ar.cliente, aec.caja, aec.fecha
                    FROM anticipos_recibidos ar
                    JOIN anticipos_estados_caja aec ON aec.anticipo_id = ar.id
                    WHERE aec.local=%s AND DATE(aec.fecha)=%s
                      AND aec.estado = 'consumido'
                    ORDER BY ar.id ASC
                """, (local, f))
                rows = cur.fetchall() or []
                print(f"🔍 DEBUG Anticipos - Rows encontrados: {len(rows)}")
                for r in rows:
                    print(f"   - Cliente: {r[6]}, Fecha aec: {r[8]}, Importe: {r[4]}")
                    anticipos_items.append({
                        "id": r[0],
                        "fecha_anticipo_recibido": str(r[1]) if r[1] else "",
                        "medio_pago": r[2] or "",
                        "comentario": f"{r[6] or ''} - {r[3] or ''}" if r[6] or r[3] else "",  # Cliente + observaciones
                        "monto": float(r[4] or 0),
                        "created_by": r[5] or "",  # usuario que consumió
                        "caja": r[7] or ""  # caja donde se consumió
                    })
            except Exception as e:
                print(f"⚠️ Error obteniendo detalle de anticipos: {e}")
                import traceback
                traceback.print_exc()

        # ===== FACTURAS (A, B, CC) =====
        # Helper para obtener items de facturas
        def _fetch_facturas_items(tipo):
            try:
                cur.execute(
                    f"""
                    SELECT punto_venta, nro_factura, monto
                    FROM {T_FACTURAS}
                    WHERE DATE(fecha)=%s AND local=%s AND tipo=%s
                    ORDER BY punto_venta, nro_factura
                    """,
                    (f, local, tipo)
                )
                rows = cur.fetchall() or []
                items = []
                for r in rows:
                    pv = r[0]
                    nro = r[1]
                    monto = float(r[2] or 0)
                    item = {
                        "pv": str(pv) if pv is not None else "0",
                        "punto_venta": str(pv) if pv is not None else "0",
                        "nro_factura": str(nro),
                        "monto": monto,
                        "z_display": _format_z(pv, nro)
                    }
                    items.append(item)
                return items, sum(i['monto'] for i in items)
            except Exception:
                return [], 0.0

        # Facturas A y B con items
        facturas_a_items, facturas_a = _fetch_facturas_items('A')
        facturas_b_items, facturas_b = _fetch_facturas_items('B')
        facturas_total = float(facturas_a or 0.0) + float(facturas_b or 0.0)

        # Facturas CC (Cuenta Corriente)
        facturas_cc = _qsum(
            cur,
            f"SELECT COALESCE(SUM(monto),0) FROM {T_FACTURAS} WHERE DATE(fecha)=%s AND local=%s AND tipo='CC'",
            (f, local),
        )

        # ===== CTA CTE (legacy - si existe en cuenta_corriente_trns) =====
        cta_cte_legacy = _sum_cuenta_corriente(conn, cur, f, local)
        # Sumamos CC de facturas + cta cte legacy
        cta_cte_total = float(facturas_cc or 0.0) + float(cta_cte_legacy or 0.0)

        # ===== TIPS TARJETAS (detalle por marca + total) desde tarjetas_trns =====
        tips_tarj_breakdown, tips_tarj_total = _sum_tips_tarjetas_breakdown(cur, T_TARJETAS, f, local)
        tips_total = float(tips_tarj_total or 0.0) + float(tips_mp or 0.0)

        # ===== Totales del panel =====
        # Las facturas A, B, Z NO suman al total cobrado (solo sirven para calcular discovery)
        # Los TIPS tampoco suman al total cobrado
        # Suman: efectivo, tarjetas, MP, rappi, pedidosya, cuenta corriente (CC), gastos, anticipos
        total_cobrado = float(sum([
            efectivo_neto or 0.0,
            tarjeta_total or 0.0,
            mp_total or 0.0,
            rappi_total or 0.0,
            pedidosya_total or 0.0,
            cta_cte_total or 0.0,  # Cuenta corriente (facturas CC)
            gastos_total or 0.0,   # Gastos justifican la venta
            anticipos_total or 0.0,  # Anticipos consumidos
        ]))

        info_total = float(sum([
            efectivo_neto or 0.0,
            tarjeta_total or 0.0,
            mp_total or 0.0,
            rappi_total or 0.0,
            pedidosya_total or 0.0,
            facturas_total or 0.0,
            gastos_total or 0.0,
            cta_cte_total or 0.0,
            tips_total or 0.0,
            anticipos_total or 0.0,
        ]))

        # ===== Diferencias detalladas por caja/turno =====
        diferencias_detalle = _get_diferencias_detalle(cur, fecha_s, local, usar_snap)

        # Agregar timestamp del servidor para validación de datos
        from datetime import datetime
        import pytz
        tz_arg = pytz.timezone('America/Argentina/Buenos_Aires')
        server_timestamp = datetime.now(tz_arg).isoformat()

        payload = {
            "fecha": fecha_s,
            "local": local,
            "local_cerrado": bool(local_cerrado),
            "server_timestamp": server_timestamp,  # ← Timestamp para validación de integridad
            "data_source": "snap" if usar_snap else "normal",  # ← Indicar fuente de datos

            "resumen": {
                "venta_total": float(venta_total or 0.0),
                "total_cobrado": total_cobrado,
                "diferencia": total_cobrado - float(venta_total or 0.0),
                "diferencias_detalle": diferencias_detalle
            },

            "index": {
                "vta_z_total": float(vta_z_total or 0.0),
                "discovery": float(discovery or 0.0),
                "z_items": z_items  # cada item trae pv, numero_z, monto y z_display
            },

            "info": {
                "total": info_total,
                "efectivo": {
                    "total": float(efectivo_neto or 0.0),
                    "remesas": float(efectivo_remesas or 0.0),
                    "neto_efectivo": float(efectivo_neto or 0.0)
                },
                "tarjeta": {
                    "total": tarjeta_total,
                    "breakdown": {k: float(v or 0.0) for k, v in tarjetas_det.items()}
                },
                "mercadopago": {
                    "total": float(mp_total or 0.0),
                    "tips": float(tips_mp or 0.0)
                },
                "rappi": { "total": float(rappi_total or 0.0) },
                "pedidosya": { "total": float(pedidosya_total or 0.0) },
                "facturas": {
                    "total": float(facturas_total or 0.0),
                    "a": float(facturas_a or 0.0),
                    "b": float(facturas_b or 0.0),
                    "a_items": facturas_a_items,
                    "b_items": facturas_b_items
                },
                "gastos": {
                    "total": float(gastos_total or 0.0),
                    "detalle": gastos_detalle
                },
                "anticipos": {
                    "total": float(anticipos_total or 0.0),
                    "items": anticipos_items
                },
                "cuenta_cte": {
                    "total": float(cta_cte_total or 0.0),
                    "cc": float(facturas_cc or 0.0),
                    "legacy": float(cta_cte_legacy or 0.0)
                },
                "tips": {
                    "total": float(tips_total or 0.0),
                    "mp": float(tips_mp or 0.0),
                    "tarjetas": float(tips_tarj_total or 0.0),
                    "breakdown": tips_tarj_breakdown  # <<--- DETALLE POR TARJETA
                }
            }
        }

        resp = make_response(jsonify(payload))
        resp.headers['Cache-Control'] = 'no-store, max-age=0'
        return resp

    except Exception as e:
        try: conn.rollback()
        except: ...
        return jsonify(error=str(e)), 500
    finally:
        try: cur.close()
        except: ...
        try: conn.close()
        except: ...


def _fetch_ventas_z_dyn(cur, fecha, local, table_name="ventas_z_trns", tipo='Z'):
    """
    Devuelve (items, total) con detección flexible de columnas:
    - Punto de venta: punto_venta | pto_venta | pv | p_venta | punto_de_venta
    - Número Z:      numero_z | nro_z | z_numero | num_z
    - Monto:         monto | importe | total
    items: [{'pv': '5', 'punto_venta': '5', 'numero_z': '1691', 'monto': 1000.0, 'z_display': '00005-00001691'}, ...]
    """
    pv_cols    = ["punto_venta", "pto_venta", "pv", "p_venta", "punto_de_venta"]
    numero_cols = ["nro_factura", "nro_z", "z_numero", "num_z"]
    monto_cols  = ["monto", "importe", "total"]
    last_error = None

    # Intento con PV + Número
    for pvcol in pv_cols:
        for ncol in numero_cols:
            for mcol in monto_cols:
                try:
                    cur.execute(
                        f"""
                        SELECT {pvcol} AS pv, {ncol} AS znum, COALESCE(SUM({mcol}),0) AS total
                        FROM {table_name}
                        WHERE DATE(fecha)=%s AND local=%s AND tipo=%s
                        GROUP BY {pvcol}, {ncol}
                        ORDER BY {pvcol}, {ncol}
                        """,
                        (fecha, local, tipo),
                    )
                    rows = cur.fetchall() or []
                    items = []
                    for r in rows:
                        pv = r[0]
                        nz = r[1]
                        monto = float(r[2] or 0)
                        item = {
                            "pv": str(pv) if pv is not None else "0",
                            "punto_venta": str(pv) if pv is not None else "0",
                            "numero_z": str(nz),
                            "monto": monto,
                        }
                        item["z_display"] = _format_z(item["pv"], item["numero_z"])
                        items.append(item)
                    return items, sum(i['monto'] for i in items)
                except Exception as e:
                    last_error = e
                    continue

    # Fallback: si no existe PV, agrupo solo por número (PV = '0')
    for ncol in numero_cols:
        for mcol in monto_cols:
            try:
                cur.execute(
                    f"""
                    SELECT {ncol} AS znum, COALESCE(SUM({mcol}),0) AS total
                    FROM {table_name}
                    WHERE DATE(fecha)=%s AND local=%s AND tipo=%s
                    GROUP BY {ncol}
                    ORDER BY {ncol}
                    """,
                    (fecha, local, tipo),
                )
                rows = cur.fetchall() or []
                items = []
                for r in rows:
                    nz = r[0]
                    monto = float(r[1] or 0)
                    item = {
                        "pv": "0",
                        "punto_venta": "0",
                        "numero_z": str(nz),
                        "monto": monto,
                    }
                    item["z_display"] = _format_z(item["pv"], item["numero_z"])
                    items.append(item)
                return items, sum(i['monto'] for i in items)
            except Exception as e:
                last_error = e
                continue

    # Si nada funcionó, devolvemos vacío
    return [], 0.0




@app.route('/api/locales_options', methods=['GET'])
@login_required
def api_locales_options():
    lvl = get_user_level()
    conn = get_db_connection()
    try:
        if lvl >= 3:
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT local FROM locales ORDER BY local;")
            locales = [r[0] for r in cur.fetchall()]
            cur.close()
        else:
            locales = [session.get('local')]
        return jsonify(locales=locales)
    finally:
        try: conn.close()
        except: ...




## _______________________________ REPORTERIA REMESAS _______________________


# --- Reportería: Remesas ---
from datetime import date, datetime, timedelta
from io import BytesIO
from flask import send_file, jsonify, render_template, request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Página UI
@app.route("/tesoreria/home")
@login_required
@role_min_required(7)  # Tesoreros (role_level 7+)
def tesoreria_home_new():
    """
    Página de inicio para tesorería con dos opciones principales.
    """
    return render_template("tesoreria_home.html")

@app.route("/reporteria/remesas")
@login_required
@role_min_required(7)  # Tesoreros (role_level 7+) - Carga individual de remesas
def ui_reporteria_remesas():
    """
    Vista de carga individual de remesas para tesoreros.
    Permite cargar el monto real de cada bolsa/remesa.
    NOTA: Esta es la vista HISTÓRICA con filtro de fecha.
    """
    return render_template("reporte_remesas.html")

@app.route("/reporteria/remesas-trabajo")
@login_required
@role_min_required(7)  # Tesoreros (role_level 7+)
def ui_reporteria_remesas_trabajo():
    """
    Vista de TRABAJO para tesoreros.
    Muestra SOLO remesas en estado TRAN (pendientes de contabilizar).
    Sin filtro de fecha - muestra todas las TRAN.
    Read-only - solo visualización.
    """
    return render_template("reporte_remesas_trabajo.html")

@app.route("/reporteria/remesas-tesoreria")
@login_required
@role_min_required(7)  # Tesoreros (role_level 7+) - Resumen por local (solo lectura)
def reporte_remesas_tesoreria_page():
    """
    Vista de resumen agrupado por local para tesorería.
    SOLO LECTURA - Muestra lo cargado previamente en /reporteria/remesas.
    Sirve para reclamar diferencias a los locales.
    """
    return render_template("reporte_remesas_tesoreria.html")


@app.route("/reporteria/resumen-tesoreria")
@login_required
@role_min_required(7)  # Tesoreros y admin_tesoreria
def reporte_resumen_tesoreria_page():
    """
    Nuevo reporte visual de tesorería por local con formato similar a Excel de gerencia.
    - Detecta automáticamente si es lunes (muestra fin de semana: Vie, Sáb, Dom)
    - Días normales: solo fecha actual (TRAN vs Real)
    - Incluye sección 'No Enviados' con cálculo de relevancia
    """
    return render_template("resumen_tesoreria.html")


# Helpers DB
def _get_locales(cur):
    """
    Obtiene lista de locales activos.
    Compatible con cursores dictionary y normales.
    """
    # 1) tabla "locales" si existe
    try:
        cur.execute("SELECT DISTINCT local FROM locales ORDER BY local")
        rows = cur.fetchall()
        if rows:
            # Detectar si es dictionary cursor
            if isinstance(rows[0], dict):
                return [r['local'] for r in rows if r['local']]
            else:
                return [r[0] for r in rows if r[0]]
    except Exception:
        pass

    # 2) fallback: distintos locales que tengan actividad
    cur.execute("""
        SELECT DISTINCT local
        FROM (
            SELECT local FROM ventas_trns
            UNION ALL
            SELECT local FROM remesas_trns
        ) t
        ORDER BY 1
    """)
    rows = cur.fetchall()

    # Detectar si es dictionary cursor
    if rows and isinstance(rows[0], dict):
        return [r['local'] for r in rows]
    else:
        return [r[0] for r in rows]

def _sum_remesas_retiradas_misma_fecha(cur, local, fecha):
    """
    Suma de remesas que:
    - Pertenecen a la fecha de caja = fecha
    - Fueron marcadas como retiradas (independientemente de cuándo)
    """
    cur.execute("""
        SELECT COALESCE(SUM(monto),0)
        FROM remesas_trns
        WHERE local=%s
          AND DATE(fecha) = %s
          AND retirada IN (1, 'Si', 'Sí', 'sí', 'si', 'SI', 'SÍ')
    """, (local, fecha))
    r = cur.fetchone()
    return float(r[0] or 0)

def _sum_remesas_otras_fechas_retiradas_hoy(cur, local, fecha_retiro):
    """
    Suma de remesas que:
    - Pertenecen a fechas de caja ANTERIORES a fecha_retiro
    - Fueron marcadas como retiradas en fecha_retiro
    """
    cur.execute("""
        SELECT COALESCE(SUM(monto),0)
        FROM remesas_trns
        WHERE local=%s
          AND DATE(fecha) < %s
          AND DATE(fecha_retirada) = %s
          AND retirada IN (1, 'Si', 'Sí', 'sí', 'si', 'SI', 'SÍ')
    """, (local, fecha_retiro, fecha_retiro))
    r = cur.fetchone()
    return float(r[0] or 0)

def _sum_remesas_no_retiradas(cur, local, hasta_fecha):
    """
    Suma de remesas pendientes de retiro hasta una fecha.
    NOTA: Esta función se mantiene para compatibilidad pero no se usa en el nuevo reporte.
    """
    cur.execute("""
        SELECT COALESCE(SUM(monto),0)
        FROM remesas_trns
        WHERE local=%s
          AND retirada NOT IN (1, 'Si', 'Sí', 'sí', 'si', 'SI', 'SÍ')
          AND (retirada = 0 OR retirada = 'No' OR retirada IS NULL OR retirada = '')
          AND DATE(fecha) <= %s
    """, (local, hasta_fecha))
    r = cur.fetchone()
    return float(r[0] or 0)

def _dia_es_lunes(d):
    return d.weekday() == 0  # 0=Lunes

def _nombre_dia_es(d):
    # Lunes..Domingo capitalizado
    dias = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
    return dias[d.weekday()]

# Construye el reporte (datos) - NUEVA LÓGICA PARA TESORERÍA
def _build_remesas_report(fecha_sel: date):
    """
    Reporte rediseñado para Tesorería.

    Calcula el TEÓRICO A RECIBIR para una fecha de retiro:
    - Remesas de esa fecha de caja que fueron retiradas
    - Remesas de fechas anteriores que se retiraron en esa fecha

    El objetivo es que tesorería sepa cuánto efectivo debería recibir ese día.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        dia_label = _nombre_dia_es(fecha_sel)
        locales = _get_locales(cur)
        rows = []

        # Totales
        tot = {
            "remesas_del_dia": 0.0,
            "remesas_anteriores": 0.0,
            "total_teorico": 0.0,
        }

        for loc in locales:
            # 1. Remesas de esta fecha de caja que fueron retiradas
            remesas_del_dia = _sum_remesas_retiradas_misma_fecha(cur, loc, fecha_sel)

            # 2. Remesas de fechas anteriores que se retiraron HOY
            remesas_anteriores = _sum_remesas_otras_fechas_retiradas_hoy(cur, loc, fecha_sel)

            # 3. Total teórico = lo que tesorería debería recibir
            total_teorico = remesas_del_dia + remesas_anteriores

            row = {
                "fecha": fecha_sel.isoformat(),
                "local": loc,
                "remesas_del_dia": remesas_del_dia,
                "remesas_anteriores": remesas_anteriores,
                "total_teorico": total_teorico,
            }

            # Acumular totales
            tot["remesas_del_dia"] += remesas_del_dia
            tot["remesas_anteriores"] += remesas_anteriores
            tot["total_teorico"] += total_teorico

            rows.append(row)

        return {
            "fecha": fecha_sel.isoformat(),
            "dia_label": dia_label,
            "rows": rows,
            "totals": tot,
        }
    finally:
        cur.close()
        conn.close()

# ==================== NUEVO: REPORTE MATRICIAL ====================
def _build_remesas_matrix_report(fecha_desde: date, fecha_hasta: date):
    """
    Reporte MATRICIAL para Tesorería.

    Retorna datos estructurados por LOCAL (filas) x FECHAS (columnas)
    Permite ver de un vistazo múltiples días y hacer click para expandir detalle.

    Estructura de retorno:
    {
        "fechas": ["2025-01-15", "2025-01-16", ...],
        "locales": ["Local 1", "Local 2", ...],
        "matriz": {
            "Local 1": {
                "2025-01-15": {"teorico": 1800, "real": 1700, "dif": -100, "remesas": [...]},
                "2025-01-16": {...}
            },
            ...
        },
        "totales_por_fecha": {
            "2025-01-15": {"teorico": 3100, "real": 3000, "dif": -100},
            ...
        }
    }
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        locales = _get_locales(cur)

        # Generar lista de fechas en el rango
        fechas = []
        current_date = fecha_desde
        while current_date <= fecha_hasta:
            fechas.append(current_date.isoformat())
            current_date += timedelta(days=1)

        # Estructura para la matriz
        matriz = {}
        totales_por_fecha = {f: {"teorico": 0, "real": 0, "dif": 0} for f in fechas}

        for local in locales:
            matriz[local] = {}

            for fecha_str in fechas:
                fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d").date()

                # Obtener remesas retiradas en esta fecha para este local
                cur.execute("""
                    SELECT
                        id, nro_remesa, precinto, monto, retirada_por,
                        fecha as fecha_caja, caja, turno, fecha_retirada
                    FROM remesas_trns
                    WHERE local = %s
                      AND DATE(fecha_retirada) = %s
                      AND retirada IN (1, 'Si', 'Sí', 'sí', 'si', 'SI', 'SÍ')
                    ORDER BY id
                """, (local, fecha_obj))

                remesas = cur.fetchall() or []

                # Calcular teórico
                teorico = sum(float(r['monto'] or 0) for r in remesas)

                # Obtener montos reales de cada remesa individual
                reales_por_remesa = {}
                real = 0
                try:
                    # Usar un nuevo cursor para evitar "Unread result found"
                    cur2 = conn.cursor(dictionary=True)
                    cur2.execute("""
                        SELECT remesa_id, monto_real
                        FROM tesoreria_recibido
                        WHERE local = %s AND fecha_retiro = %s AND remesa_id IS NOT NULL
                    """, (local, fecha_obj))

                    for row in cur2.fetchall() or []:
                        reales_por_remesa[row['remesa_id']] = float(row['monto_real']) if row['monto_real'] else 0
                        real += reales_por_remesa[row['remesa_id']]

                    cur2.close()
                except Exception as e:
                    # Tabla no existe o error, usar 0
                    print(f"⚠️ Error consultando tesoreria_recibido: {e}")
                    pass
                dif = teorico - real
                # Determinar estado basado en la diferencia
                if real == 0 and teorico > 0:
                    estado = 'en_transito'
                elif abs(dif) < 0.01:  # Tolerancia de 1 centavo
                    estado = 'recibido'
                elif dif != 0:
                    estado = 'con_diferencia'
                else:
                    estado = None
                observaciones = ''

                # Preparar remesas para el detalle (convertir a formato serializable)
                remesas_detail = []
                for r in remesas:
                    # Convertir fechas a string de forma segura
                    fecha_caja_str = None
                    if r.get('fecha_caja'):
                        if hasattr(r['fecha_caja'], 'isoformat'):
                            fecha_caja_str = r['fecha_caja'].isoformat()
                        else:
                            fecha_caja_str = str(r['fecha_caja'])

                    fecha_retirada_str = None
                    if r.get('fecha_retirada'):
                        if hasattr(r['fecha_retirada'], 'isoformat'):
                            fecha_retirada_str = r['fecha_retirada'].isoformat()
                        else:
                            fecha_retirada_str = str(r['fecha_retirada'])

                    remesas_detail.append({
                        'id': r.get('id'),
                        'nro_remesa': r.get('nro_remesa'),
                        'precinto': r.get('precinto'),
                        'monto': float(r.get('monto') or 0),
                        'real': reales_por_remesa.get(r.get('id'), 0),  # Agregar monto real individual
                        'retirada_por': r.get('retirada_por'),
                        'fecha_caja': fecha_caja_str,
                        'caja': r.get('caja'),
                        'turno': r.get('turno'),
                        'fecha_retirada': fecha_retirada_str
                    })

                matriz[local][fecha_str] = {
                    "teorico": teorico,
                    "real": real,
                    "dif": dif,
                    "estado": estado,
                    "observaciones": observaciones,
                    "remesas": remesas_detail,
                    "tiene_remesas": len(remesas) > 0
                }

                # Acumular totales por fecha
                totales_por_fecha[fecha_str]["teorico"] += teorico
                totales_por_fecha[fecha_str]["real"] += real
                totales_por_fecha[fecha_str]["dif"] += dif

        return {
            "fecha_desde": fecha_desde.isoformat(),
            "fecha_hasta": fecha_hasta.isoformat(),
            "fechas": fechas,
            "locales": locales,
            "matriz": matriz,
            "totales_por_fecha": totales_por_fecha
        }
    finally:
        cur.close()
        conn.close()

# API JSON - Mantener el actual para compatibilidad
@app.route("/api/reportes/remesas")
@login_required
def api_reportes_remesas():
    f = request.args.get("fecha")
    try:
        fecha_sel = datetime.strptime(f, "%Y-%m-%d").date() if f else date.today()
    except Exception:
        return jsonify(error="fecha inválida (YYYY-MM-DD)"), 400

    data = _build_remesas_report(fecha_sel)
    return jsonify(data)

# NUEVO: API para reporte matricial
@app.route("/api/reportes/remesas-matriz")
@login_required
@role_min_required(3)  # Solo auditores y tesorería
def api_reportes_remesas_matriz():
    fecha_desde_str = request.args.get("fecha_desde")
    fecha_hasta_str = request.args.get("fecha_hasta")

    try:
        if fecha_desde_str:
            fecha_desde = datetime.strptime(fecha_desde_str, "%Y-%m-%d").date()
        else:
            # Por defecto: últimos 7 días
            fecha_desde = date.today() - timedelta(days=6)

        if fecha_hasta_str:
            fecha_hasta = datetime.strptime(fecha_hasta_str, "%Y-%m-%d").date()
        else:
            fecha_hasta = date.today()

        # Validar que el rango no sea muy grande (máximo 31 días)
        if (fecha_hasta - fecha_desde).days > 31:
            return jsonify(error="El rango de fechas no puede ser mayor a 31 días"), 400

        data = _build_remesas_matrix_report(fecha_desde, fecha_hasta)
        return jsonify(data)
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"ERROR en api_reportes_remesas_matriz: {error_detail}")
        return jsonify(error=f"Error al generar reporte: {str(e)}", details=error_detail), 400


@app.route('/api/tesoreria/obtener-real', methods=['GET'])
@login_required
@role_min_required(7)
def api_tesoreria_obtener_real():
    """
    Obtiene el monto real registrado para un local y fecha de retiro.
    Consulta la tabla tesoreria_recibido.
    """
    local = request.args.get('local', '').strip()
    fecha_retiro = request.args.get('fecha_retiro', '').strip()

    if not (local and fecha_retiro):
        return jsonify(success=False, msg='Faltan parámetros'), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Sumar TODAS las remesas individuales de este local y fecha
        cur.execute("""
            SELECT
                SUM(monto_real) as monto_real_total,
                SUM(monto_teorico) as monto_teorico_total,
                COUNT(*) as cantidad_remesas
            FROM tesoreria_recibido
            WHERE local = %s AND fecha_retiro = %s
        """, (local, fecha_retiro))

        resultado = cur.fetchone()
        cur.close()
        conn.close()

        if resultado and resultado['cantidad_remesas'] > 0:
            monto_real = float(resultado['monto_real_total']) if resultado['monto_real_total'] else 0
            monto_teorico = float(resultado['monto_teorico_total']) if resultado['monto_teorico_total'] else 0

            # Determinar estado basado en diferencia
            diferencia = abs(monto_teorico - monto_real)
            if monto_real == 0:
                estado = 'en_transito'
            elif diferencia < 0.01:
                estado = 'recibido'
            else:
                estado = 'con_diferencia'

            return jsonify(
                success=True,
                monto_real=monto_real,
                monto_teorico=monto_teorico,
                estado=estado,
                observaciones='',
                cantidad_remesas=resultado['cantidad_remesas']
            )
        else:
            # No hay registros, devolver 0
            return jsonify(success=True, monto_real=0, monto_teorico=0, estado='en_transito', observaciones='', cantidad_remesas=0)

    except Exception as e:
        print(f"❌ ERROR obtener_real: {e}")
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/tesoreria/remesas-tran', methods=['GET'])
@login_required
@role_min_required(7)
@rate_limited(max_requests=60, window_seconds=60)
def api_tesoreria_remesas_tran():
    """
    Obtiene TODAS las remesas en estado TRAN (pendientes de contabilizar).
    Sin filtro de fecha - retorna todas las que están en tránsito.

    SEGURIDAD:
    - Rate limiting: 60 requests/min ✅
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Obtener todas las remesas en estado TRAN
        cur.execute("""
            SELECT
                r.id,
                r.fecha,
                r.precinto,
                r.nro_remesa,
                r.local,
                r.caja,
                r.turno,
                r.monto as monto_teorico,
                r.fecha_retirada,
                r.retirada_por,
                r.estado_contable,
                r.fecha_estado_tran,
                COALESCE(t.monto_real, 0) as monto_real
            FROM remesas_trns r
            LEFT JOIN tesoreria_recibido t ON t.remesa_id = r.id
            WHERE r.estado_contable = 'TRAN'
            ORDER BY r.fecha_retirada DESC, r.local, r.caja
        """)

        remesas = cur.fetchall() or []

        # Convertir decimales y fechas a formatos serializables
        for remesa in remesas:
            if remesa.get('monto_teorico'):
                remesa['monto_teorico'] = float(remesa['monto_teorico'])
            if remesa.get('monto_real'):
                remesa['monto_real'] = float(remesa['monto_real'])
            if remesa.get('fecha'):
                remesa['fecha'] = remesa['fecha'].isoformat() if hasattr(remesa['fecha'], 'isoformat') else str(remesa['fecha'])
            if remesa.get('fecha_retirada'):
                remesa['fecha_retirada'] = remesa['fecha_retirada'].isoformat() if hasattr(remesa['fecha_retirada'], 'isoformat') else str(remesa['fecha_retirada'])
            if remesa.get('fecha_estado_tran'):
                remesa['fecha_estado_tran'] = remesa['fecha_estado_tran'].isoformat() if hasattr(remesa['fecha_estado_tran'], 'isoformat') else str(remesa['fecha_estado_tran'])

            # Debug log para remesa 956
            if remesa.get('id') == 956 or remesa.get('nro_remesa') == '438410':
                print(f"🔍 DEBUG Remesa 956/438410: fecha={remesa.get('fecha')}, fecha_retirada={remesa.get('fecha_retirada')}, local={remesa.get('local')}")

        cur.close()
        conn.close()

        return jsonify(success=True, remesas=remesas, count=len(remesas))

    except Exception as e:
        print(f"❌ ERROR remesas-tran: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/tesoreria/remesas-detalle', methods=['GET'])
@login_required
@role_min_required(7)
@rate_limited(max_requests=60, window_seconds=60)
def api_tesoreria_remesas_detalle():
    """
    Obtiene TODAS las remesas retiradas en una fecha específica.
    Retorna un array con cada remesa individual (no agrupadas).
    Incluye el monto real y fecha_sello si ya fue registrado en tesoreria_recibido.

    Parámetros:
    - fecha_retiro: fecha en que se marcó como retirada (obligatorio)
    - fecha_sello: fecha de apertura/contabilización (opcional, filtra adicionalmente)

    SEGURIDAD:
    - Rate limiting: 60 requests/min ✅
    """
    fecha_retiro = request.args.get('fecha_retiro', '').strip()
    fecha_sello = request.args.get('fecha_sello', '').strip()

    if not fecha_retiro:
        return jsonify(success=False, msg='Falta fecha_retiro'), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Obtener todas las remesas retiradas en esta fecha
        # IMPORTANTE: fecha_retirada es la fecha en que se marcó como retirada
        cur.execute("""
            SELECT
                id,
                fecha AS fecha_caja,
                precinto,
                nro_remesa,
                local,
                caja,
                turno,
                monto,
                fecha_retirada,
                retirada_por
            FROM remesas_trns
            WHERE retirada IN (1, 'Si', 'Sí', 'sí', 'si', 'SI', 'SÍ')
              AND DATE(fecha_retirada) = %s
            ORDER BY COALESCE(NULLIF(precinto, ''), '9999'), nro_remesa
        """, (fecha_retiro,))

        remesas = cur.fetchall() or []

        # DEBUG: Verificar si precinto y nro_remesa vienen de remesas_trns
        print(f"🔍 DEBUG remesas-detalle: fecha={fecha_retiro}, cantidad={len(remesas)}")
        if remesas:
            print(f"🔍 Primera remesa: precinto={remesas[0].get('precinto')}, nro_remesa={remesas[0].get('nro_remesa')}, local={remesas[0].get('local')}")

        # Para cada remesa, consultar si tiene monto real registrado
        # Usar remesa_id como clave principal (más confiable que precinto/nro_remesa)
        if fecha_sello:
            # Si se proporciona fecha_sello, filtrar por ella
            cur.execute("""
                SELECT remesa_id, monto_real, fecha_sello
                FROM tesoreria_recibido
                WHERE fecha_retiro = %s
                  AND fecha_sello = %s
                  AND remesa_id IS NOT NULL
            """, (fecha_retiro, fecha_sello))
        else:
            # Sin filtro de fecha_sello, traer todas
            cur.execute("""
                SELECT remesa_id, monto_real, fecha_sello
                FROM tesoreria_recibido
                WHERE fecha_retiro = %s AND remesa_id IS NOT NULL
            """, (fecha_retiro,))

        reales = {}
        fechas_sello = {}
        for row in cur.fetchall() or []:
            reales[row['remesa_id']] = float(row['monto_real']) if row['monto_real'] else 0
            fechas_sello[row['remesa_id']] = row['fecha_sello'].isoformat() if row.get('fecha_sello') and hasattr(row['fecha_sello'], 'isoformat') else str(row.get('fecha_sello')) if row.get('fecha_sello') else None

        # Adjuntar el real y fecha_sello a cada remesa individual (si existe)
        for remesa in remesas:
            remesa['real'] = reales.get(remesa['id'], 0)
            remesa['fecha_sello'] = fechas_sello.get(remesa['id'], None)

            # Convertir decimales a float
            if remesa.get('monto'):
                remesa['monto'] = float(remesa['monto'])

        cur.close()
        conn.close()

        print(f"🔍 DEBUG remesas-detalle: Retornando {len(remesas)} remesas")

        return jsonify(success=True, remesas=remesas)

    except Exception as e:
        print(f"❌ ERROR remesas-detalle: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/tesoreria/guardar-remesa', methods=['POST'])
@login_required
@role_min_required(7)
@tesoreria_secured(max_requests=30, window_seconds=60)
def api_tesoreria_guardar_remesa():
    """
    Guarda el monto real de UNA remesa específica.
    Actualiza o crea registro en tesoreria_recibido.
    Solo permite guardar si la fecha NO está aprobada.

    SEGURIDAD:
    - CSRF protection ✅
    - Rate limiting: 30 requests/min ✅
    - Audit logging ✅
    """
    data = request.get_json() or {}
    remesa_id = data.get('remesa_id')
    local = data.get('local', '').strip()
    fecha_retiro = data.get('fecha_retiro', '').strip()
    precinto = data.get('precinto', '').strip()
    nro_remesa = data.get('nro_remesa', '').strip()
    monto_real = data.get('monto_real', 0)
    monto_teorico = data.get('monto_teorico', 0)

    if not (local and fecha_retiro and remesa_id):
        return jsonify(success=False, msg='Faltan datos requeridos (local, fecha_retiro, remesa_id)'), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        username = session.get('username', 'SYSTEM')
        user_id = session.get('user_id')

        # Verificar si la fecha está aprobada
        cur.execute("""
            SELECT estado FROM tesoreria_aprobaciones
            WHERE fecha_retiro = %s
        """, (fecha_retiro,))
        aprobacion = cur.fetchone()

        if aprobacion and aprobacion['estado'] == 'aprobado':
            cur.close()
            conn.close()
            return jsonify(success=False, msg='No se puede editar. La conciliación de esta fecha ya fue aprobada.'), 403

        # Obtener monto_real anterior para audit log
        cur.execute("""
            SELECT monto_real FROM tesoreria_recibido WHERE remesa_id = %s
        """, (remesa_id,))
        old_record = cur.fetchone()
        old_monto = float(old_record['monto_real']) if old_record else 0

        # Determinar estado según diferencia
        dif = float(monto_teorico) - float(monto_real)
        if monto_real == 0:
            estado = 'en_transito'
        elif abs(dif) < 0.01:  # Tolerancia de 1 centavo
            estado = 'recibido'
        else:
            estado = 'con_diferencia'

        # Obtener fecha de hoy (fecha_sello = fecha de apertura del bolsín)
        from datetime import date
        fecha_sello = date.today().isoformat()

        # Insertar o actualizar remesa individual usando remesa_id
        cur.execute("""
            INSERT INTO tesoreria_recibido
                (remesa_id, local, fecha_retiro, precinto, nro_remesa, monto_teorico, monto_real, estado, fecha_sello, registrado_por, registrado_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE
                monto_real = VALUES(monto_real),
                monto_teorico = VALUES(monto_teorico),
                estado = VALUES(estado),
                fecha_sello = VALUES(fecha_sello),
                registrado_por = VALUES(registrado_por),
                registrado_at = NOW()
        """, (remesa_id, local, fecha_retiro, precinto, nro_remesa, monto_teorico, monto_real, estado, fecha_sello, username))

        # NUEVO: Actualizar estado_contable en remesas_trns si se contabilizó
        if float(monto_real) > 0:
            cur.execute("""
                UPDATE remesas_trns
                SET estado_contable = 'Contabilizada',
                    fecha_estado_contabilizada = NOW()
                WHERE id = %s
                  AND estado_contable != 'Contabilizada'
            """, (remesa_id,))

        conn.commit()

        # NUEVO: Registrar en audit log si cambió el monto
        if float(monto_real) != old_monto:
            AuditLogger.log_remesa_change(
                conn, remesa_id, 'monto_real', old_monto, float(monto_real), user_id
            )

        cur.close()
        conn.close()

        return jsonify(success=True, msg='Guardado correctamente', estado=estado)

    except Exception as e:
        print(f"❌ ERROR guardar-remesa: {e}")
        import traceback
        traceback.print_exc()
        try:
            conn.rollback()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/tesoreria/aprobar-conciliacion', methods=['POST'])
@login_required
@role_min_required(8)
@tesoreria_secured(max_requests=10, window_seconds=60)
def api_aprobar_conciliacion():
    """
    Aprueba la conciliación de una fecha específica.
    Solo admin de tesorería puede aprobar.

    SEGURIDAD:
    - CSRF protection ✅
    - Rate limiting: 10 requests/min ✅
    """
    data = request.get_json() or {}
    fecha_retiro = data.get('fecha_retiro', '').strip()
    observaciones = data.get('observaciones', '').strip()

    if not fecha_retiro:
        return jsonify(success=False, msg='Falta fecha_retiro'), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        username = session.get('username', 'SYSTEM')

        # Insertar o actualizar aprobación
        cur.execute("""
            INSERT INTO tesoreria_aprobaciones
                (fecha_retiro, estado, aprobado_por, aprobado_at, observaciones)
            VALUES (%s, 'aprobado', %s, NOW(), %s)
            ON DUPLICATE KEY UPDATE
                estado = 'aprobado',
                aprobado_por = VALUES(aprobado_por),
                aprobado_at = NOW(),
                observaciones = VALUES(observaciones)
        """, (fecha_retiro, username, observaciones))

        # Registrar en auditoría
        cur.execute("""
            INSERT INTO tesoreria_aprobaciones_audit
                (fecha_retiro, accion, usuario, observaciones)
            VALUES (%s, 'aprobar', %s, %s)
        """, (fecha_retiro, username, observaciones))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify(success=True, msg='Conciliación aprobada correctamente')

    except Exception as e:
        print(f"❌ ERROR aprobar-conciliacion: {e}")
        import traceback
        traceback.print_exc()
        try:
            conn.rollback()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/tesoreria/desaprobar-conciliacion', methods=['POST'])
@login_required
@role_min_required(8)
@tesoreria_secured(max_requests=10, window_seconds=60)
def api_desaprobar_conciliacion():
    """
    Desaprueba la conciliación de una fecha específica.

    SEGURIDAD:
    - CSRF protection ✅
    - Rate limiting: 10 requests/min ✅
    """
    data = request.get_json() or {}
    fecha_retiro = data.get('fecha_retiro', '').strip()
    observaciones = data.get('observaciones', '').strip()

    if not fecha_retiro:
        return jsonify(success=False, msg='Falta fecha_retiro'), 400

    if not observaciones:
        return jsonify(success=False, msg='Debes proporcionar un motivo para desaprobar'), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        username = session.get('username', 'SYSTEM')

        # Actualizar aprobación a desaprobado
        cur.execute("""
            INSERT INTO tesoreria_aprobaciones
                (fecha_retiro, estado, desaprobado_por, desaprobado_at, observaciones)
            VALUES (%s, 'desaprobado', %s, NOW(), %s)
            ON DUPLICATE KEY UPDATE
                estado = 'desaprobado',
                desaprobado_por = VALUES(desaprobado_por),
                desaprobado_at = NOW(),
                observaciones = VALUES(observaciones)
        """, (fecha_retiro, username, observaciones))

        # Registrar en auditoría
        cur.execute("""
            INSERT INTO tesoreria_aprobaciones_audit
                (fecha_retiro, accion, usuario, observaciones)
            VALUES (%s, 'desaprobar', %s, %s)
        """, (fecha_retiro, username, observaciones))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify(success=True, msg='Conciliación desaprobada. Los tesoreros pueden volver a editar.')

    except Exception as e:
        print(f"❌ ERROR desaprobar-conciliacion: {e}")
        import traceback
        traceback.print_exc()
        try:
            conn.rollback()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/tesoreria/devolver-a-local/<int:remesa_id>', methods=['POST'])
@login_required
@role_min_required(7)
@tesoreria_secured(max_requests=20, window_seconds=60)
def api_devolver_remesa_a_local(remesa_id):
    """
    Devuelve una remesa de estado TRAN a estado Local.
    Se usa cuando la remesa no llegó físicamente a tesorería.
    Registra en auditoría.

    SEGURIDAD:
    - CSRF protection ✅
    - Rate limiting: 20 requests/min ✅
    - Audit logging ✅
    """
    data = request.get_json() or {}
    motivo = data.get('motivo', '').strip()

    if not motivo:
        motivo = 'Remesa no recibida físicamente'

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        username = session.get('username', 'SYSTEM')
        user_id = session.get('user_id')

        # Obtener datos actuales de la remesa
        cur.execute("""
            SELECT id, local, fecha, caja, turno, estado_contable, fecha_retirada, retirada_por
            FROM remesas_trns
            WHERE id = %s
        """, (remesa_id,))
        remesa = cur.fetchone()

        if not remesa:
            cur.close()
            conn.close()
            return jsonify(success=False, msg='Remesa no encontrada'), 404

        if remesa['estado_contable'] != 'TRAN':
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f'La remesa está en estado {remesa["estado_contable"]}, solo se pueden devolver remesas en TRAN'), 400

        # Guardar datos anteriores para auditoría
        datos_anteriores = {
            'estado_contable': remesa['estado_contable'],
            'retirada': 1,
            'retirada_por': remesa['retirada_por'],
            'fecha_retirada': remesa['fecha_retirada'].strftime('%Y-%m-%d') if remesa['fecha_retirada'] and hasattr(remesa['fecha_retirada'], 'strftime') else str(remesa['fecha_retirada']) if remesa['fecha_retirada'] else None
        }

        # Actualizar remesa: volver a estado Local
        cur.execute("""
            UPDATE remesas_trns
            SET estado_contable = 'Local',
                retirada = 0,
                retirada_por = NULL,
                fecha_retirada = NULL,
                fecha_estado_tran = NULL,
                ult_mod = NOW()
            WHERE id = %s
        """, (remesa_id,))

        # Eliminar registro de tesoreria_recibido si existe
        cur.execute("""
            DELETE FROM tesoreria_recibido
            WHERE remesa_id = %s
        """, (remesa_id,))

        conn.commit()

        # Datos nuevos para auditoría
        datos_nuevos = {
            'estado_contable': 'Local',
            'retirada': 0,
            'retirada_por': None,
            'fecha_retirada': None,
            'motivo_devolucion': motivo
        }

        # Registrar en auditoría
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='remesas_trns',
            registro_id=remesa_id,
            datos_anteriores=datos_anteriores,
            datos_nuevos=datos_nuevos,
            observaciones=f'Remesa devuelta a Local. Motivo: {motivo}'
        )

        cur.close()
        conn.close()

        return jsonify(success=True, msg='Remesa devuelta a estado Local correctamente')

    except Exception as e:
        print(f"❌ ERROR devolver-a-local: {e}")
        import traceback
        traceback.print_exc()
        try:
            conn.rollback()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/tesoreria/estado-aprobacion', methods=['GET'])
@login_required
@role_min_required(7)  # Tesoreros (level 7+) pueden consultar, solo admin (level 8) puede aprobar
def api_estado_aprobacion():
    """
    Obtiene el estado de aprobación de una fecha.
    Disponible para tesoreros (7+) y admin_tesoreria (8+).
    """
    fecha_retiro = request.args.get('fecha_retiro', '').strip()

    if not fecha_retiro:
        return jsonify(success=False, msg='Falta fecha_retiro'), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT estado, aprobado_por, aprobado_at, desaprobado_por, desaprobado_at, observaciones
            FROM tesoreria_aprobaciones
            WHERE fecha_retiro = %s
        """, (fecha_retiro,))

        aprobacion = cur.fetchone()
        cur.close()
        conn.close()

        if aprobacion and aprobacion['estado'] == 'aprobado':
            return jsonify(
                success=True,
                aprobado=True,
                aprobado_por=aprobacion['aprobado_por'],
                fecha_aprobacion=aprobacion['aprobado_at'].isoformat() if aprobacion['aprobado_at'] else None,
                aprobacion=aprobacion  # Mantener compatibilidad con código existente
            )
        else:
            return jsonify(
                success=True,
                aprobado=False,
                aprobacion={'estado': 'pendiente'} if not aprobacion else aprobacion
            )

    except Exception as e:
        print(f"❌ ERROR estado-aprobacion: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/tesoreria/audit-log', methods=['GET'])
@login_required
@role_min_required(8)  # Solo admin_tesoreria
@rate_limited(max_requests=30, window_seconds=60)
def api_tesoreria_audit_log():
    """
    Obtiene el log de auditoría de cambios en remesas.
    Solo disponible para admin_tesoreria (level 8+).

    Query params:
    - fecha_desde: YYYY-MM-DD (opcional)
    - fecha_hasta: YYYY-MM-DD (opcional)
    - remesa_id: int (opcional)
    - usuario: string (opcional)
    - limit: int (default: 100, max: 1000)

    SEGURIDAD:
    - Solo admin_tesoreria ✅
    - Rate limiting: 30 requests/min ✅
    """
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    remesa_id = request.args.get('remesa_id')
    usuario = request.args.get('usuario')
    limit = min(int(request.args.get('limit', 100)), 1000)

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Construir query dinámicamente según filtros
        where_clauses = []
        params = []

        if fecha_desde:
            where_clauses.append("DATE(t.changed_at) >= %s")
            params.append(fecha_desde)

        if fecha_hasta:
            where_clauses.append("DATE(t.changed_at) <= %s")
            params.append(fecha_hasta)

        if remesa_id:
            where_clauses.append("t.remesa_id = %s")
            params.append(int(remesa_id))

        if usuario:
            where_clauses.append("t.changed_by_username LIKE %s")
            params.append(f"%{usuario}%")

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        query = f"""
            SELECT
                t.id,
                t.remesa_id,
                r.local,
                r.precinto,
                r.nro_remesa,
                r.fecha as fecha_remesa,
                t.field_changed,
                t.old_value,
                t.new_value,
                t.changed_by_user_id,
                t.changed_by_username,
                t.ip_address,
                t.changed_at
            FROM tesoreria_audit_log t
            LEFT JOIN remesas_trns r ON r.id = t.remesa_id
            WHERE {where_sql}
            ORDER BY t.changed_at DESC
            LIMIT %s
        """

        params.append(limit)
        cur.execute(query, tuple(params))

        logs = cur.fetchall() or []

        # Convertir datetime a string
        for log in logs:
            if log.get('changed_at'):
                log['changed_at'] = log['changed_at'].isoformat()
            if log.get('fecha_remesa'):
                log['fecha_remesa'] = log['fecha_remesa'].isoformat()

        cur.close()
        conn.close()

        return jsonify(success=True, logs=logs, count=len(logs))

    except Exception as e:
        print(f"❌ ERROR audit-log: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/tesoreria/resumen-por-local', methods=['GET'])
@login_required
@role_min_required(7)
def api_tesoreria_resumen_por_local():
    """
    Nuevo reporte de tesorería agrupado por local con formato de gerencia.

    Detecta automáticamente:
    - Si es LUNES → Muestra fin de semana (Viernes, Sábado, Domingo)
    - Otros días → Solo fecha actual

    Segmenta locales en 3 grupos:
    1. Principal (todos menos Nómade, Polo House, Narda Sucre)
    2. Nómade + Polo House
    3. Narda Sucre

    También incluye sección "No Enviados" con cálculo de relevancia.
    """
    from datetime import datetime, timedelta

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Obtener fecha del parámetro o usar fecha actual
        fecha_param = request.args.get('fecha')
        if fecha_param:
            try:
                hoy = datetime.strptime(fecha_param, '%Y-%m-%d').date()
            except:
                hoy = datetime.now().date()
        else:
            hoy = datetime.now().date()

        es_lunes = hoy.weekday() == 0  # 0 = Lunes

        # Determinar fechas a mostrar
        if es_lunes:
            # Fin de semana: Viernes, Sábado, Domingo
            viernes = hoy - timedelta(days=3)
            sabado = hoy - timedelta(days=2)
            domingo = hoy - timedelta(days=1)
            fechas = [viernes, sabado, domingo]
            fecha_labels = [
                viernes.strftime('%d/%m'),
                sabado.strftime('%d/%m'),
                domingo.strftime('%d/%m')
            ]
        else:
            # Solo hoy
            fechas = [hoy]
            fecha_labels = [hoy.strftime('%d/%m')]

        # Obtener todos los locales activos
        cur.execute("SELECT DISTINCT local FROM locales ORDER BY local")
        todos_locales = [r['local'] for r in cur.fetchall() if r['local']]

        # Definir segmentos
        segmento2 = ['Nómade', 'Polo House']
        segmento3 = ['Narda Sucre']
        segmento1 = [l for l in todos_locales if l not in segmento2 and l not in segmento3]

        segmentos = [
            {'nombre': 'Principal', 'locales': segmento1},
            {'nombre': 'Segmento 2', 'locales': segmento2},
            {'nombre': 'Narda Sucre', 'locales': segmento3}
        ]

        # Para cada local, obtener datos por fecha
        datos_por_local = {}

        for local in todos_locales:
            datos_por_local[local] = {
                'saldo_a_retirar': 0  # Remesas de HOY que no fueron retiradas
            }

            for idx, fecha in enumerate(fechas):
                fecha_label = fecha_labels[idx]

                # Obtener monto teórico de todas las remesas retiradas en esta fecha
                # SIN filtro de estado - necesitamos el teórico aunque ya esté contabilizada
                # Incluye remesas que ya fueron contabilizadas (estado_contable = 'Contabilizada')
                cur.execute("""
                    SELECT
                        SUM(r.monto) as tran_total,
                        COUNT(*) as cantidad_tran
                    FROM remesas_trns r
                    WHERE r.local = %s
                      AND DATE(r.fecha_retirada) = %s
                      AND (r.retirada = 1 OR r.estado_contable IN ('TRAN', 'Contabilizada'))
                """, (local, fecha))
                tran_data = cur.fetchone()
                tran_total = float(tran_data['tran_total']) if tran_data and tran_data['tran_total'] else 0

                # Obtener monto real contabilizado (suma de tesoreria_recibido)
                cur.execute("""
                    SELECT SUM(monto_real) as real_total
                    FROM tesoreria_recibido
                    WHERE local = %s
                      AND DATE(fecha_retiro) = %s
                """, (local, fecha))
                real_data = cur.fetchone()
                real_total = float(real_data['real_total']) if real_data and real_data['real_total'] else 0

                diferencia = real_total - tran_total

                # Guardar indexado por fecha_label para acceso fácil en frontend
                datos_por_local[local][fecha_label] = {
                    'tran': tran_total,
                    'real': real_total,
                    'diferencia': diferencia
                }

            # Calcular "Saldo a retirar" (remesas del DÍA ANTERIOR en estado Local)
            # Si hoy es 13/01, muestra remesas Local de 12/01
            fecha_ayer = hoy - timedelta(days=1)
            cur.execute("""
                SELECT SUM(monto) as saldo_pendiente
                FROM remesas_trns
                WHERE local = %s
                  AND DATE(fecha) = %s
                  AND estado_contable = 'Local'
                  AND retirada = 0
            """, (local, fecha_ayer))
            saldo_data = cur.fetchone()
            datos_por_local[local]['saldo_a_retirar'] = float(saldo_data['saldo_pendiente']) if saldo_data and saldo_data['saldo_pendiente'] else 0

        # Calcular "No Enviados" (remesas en estado Local, ordenadas por relevancia)
        cur.execute("""
            SELECT
                local,
                fecha as fecha_caja,
                monto,
                DATEDIFF(CURDATE(), fecha) as dias_transcurridos,
                (monto * (DATEDIFF(CURDATE(), fecha) + 1)) as relevancia
            FROM remesas_trns
            WHERE estado_contable = 'Local'
              AND retirada = 0
            ORDER BY relevancia DESC
            LIMIT 50
        """)
        no_enviados = cur.fetchall()

        # Convertir fechas a string
        for item in no_enviados:
            if item.get('fecha_caja'):
                item['fecha_caja'] = item['fecha_caja'].isoformat()

        cur.close()
        conn.close()

        return jsonify(
            success=True,
            es_lunes=es_lunes,
            fechas=fecha_labels,
            fechas_full=[f.strftime('%Y-%m-%d') for f in fechas],
            segmentos=segmentos,
            datos=datos_por_local,
            no_enviados=no_enviados
        )

    except Exception as e:
        print(f"❌ ERROR resumen-por-local: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


# NUEVO: Endpoint para registrar monto real recibido por tesorería
@app.route("/api/tesoreria/registrar-recibido", methods=['POST'])
@login_required
@role_min_required(3)  # Auditores (3+) y Tesorería (7+)
def registrar_monto_recibido():
    """
    Permite a tesorería registrar el monto real recibido para un local y fecha.

    Body JSON:
    {
        "local": "Fabric Sushi",
        "fecha_retiro": "2025-01-15",
        "monto_teorico": 1800,
        "monto_real": 1700,
        "observaciones": "Faltaban $100, se está investigando"
    }
    """
    data = request.get_json() or {}

    local = data.get('local', '').strip()
    fecha_retiro_str = data.get('fecha_retiro', '').strip()
    monto_teorico = data.get('monto_teorico', 0)
    monto_real = data.get('monto_real', 0)
    observaciones = data.get('observaciones', '').strip()

    if not local:
        return jsonify(success=False, msg="El local es requerido"), 400

    if not fecha_retiro_str:
        return jsonify(success=False, msg="La fecha de retiro es requerida"), 400

    try:
        fecha_retiro = datetime.strptime(fecha_retiro_str, "%Y-%m-%d").date()
    except Exception:
        return jsonify(success=False, msg="Formato de fecha inválido (usar YYYY-MM-DD)"), 400

    try:
        monto_real = float(monto_real)
        monto_teorico = float(monto_teorico)
    except Exception:
        return jsonify(success=False, msg="Los montos deben ser numéricos"), 400

    if monto_real < 0:
        return jsonify(success=False, msg="El monto real no puede ser negativo"), 400

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        username = session.get('username', 'SYSTEM')

        # Calcular diferencia y estado
        diferencia = monto_teorico - monto_real

        # Determinar estado automáticamente
        if monto_real > 0:
            if abs(diferencia) > 100:
                estado = 'con_diferencia'
            else:
                estado = 'recibido'
        else:
            estado = 'en_transito'

        # Insertar o actualizar registro de tesorería
        # NOTA: 'diferencia' es una columna generada, no la insertamos manualmente
        cur.execute("""
            INSERT INTO tesoreria_recibido
                (local, fecha_retiro, monto_teorico, monto_real, estado, observaciones, registrado_por, registrado_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE
                monto_real = VALUES(monto_real),
                monto_teorico = VALUES(monto_teorico),
                estado = VALUES(estado),
                observaciones = VALUES(observaciones),
                registrado_por = VALUES(registrado_por),
                registrado_at = NOW()
        """, (local, fecha_retiro, monto_teorico, monto_real, estado, observaciones, username))

        conn.commit()

        # Obtener el registro actualizado para devolver
        cur.execute("""
            SELECT local, fecha_retiro, monto_teorico, monto_real, diferencia, estado, observaciones
            FROM tesoreria_recibido
            WHERE local = %s AND fecha_retiro = %s
        """, (local, fecha_retiro))

        row = cur.fetchone()

        # Registrar en auditoría
        registrar_auditoria(
            conn=conn,
            accion='INSERT_UPDATE',
            tabla='tesoreria_recibido',
            registro_id=f"{local}_{fecha_retiro}",
            datos_anteriores={},
            datos_nuevos={
                'local': local,
                'fecha_retiro': fecha_retiro_str,
                'monto_teorico': monto_teorico,
                'monto_real': monto_real,
                'diferencia': monto_teorico - monto_real,
                'observaciones': observaciones
            },
            descripcion=f"Tesorería registró monto real recibido - Local: {local}, Fecha: {fecha_retiro_str}, Teórico: ${monto_teorico}, Real: ${monto_real}"
        )

        return jsonify(
            success=True,
            msg="Monto recibido registrado correctamente",
            data={
                'local': row[0],
                'fecha_retiro': row[1].isoformat() if row[1] else None,
                'monto_teorico': float(row[2]),
                'monto_real': float(row[3]),
                'diferencia': float(row[4]),
                'estado': row[5],
                'observaciones': row[6]
            }
        )
    except Exception as e:
        conn.rollback()
        return jsonify(success=False, msg=f"Error al registrar: {str(e)}"), 500
    finally:
        cur.close()
        conn.close()

# Excel (XLSX) con estilo
@app.route("/api/reportes/remesas.xlsx")
@login_required
def api_reportes_remesas_xlsx():
    f = request.args.get("fecha")
    try:
        fecha_sel = datetime.strptime(f, "%Y-%m-%d").date() if f else date.today()
    except Exception:
        return "fecha inválida", 400

    data = _build_remesas_report(fecha_sel)

    wb = Workbook()
    ws = wb.active
    ws.title = "Remesas"

    # Estilos
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    ap_fill     = PatternFill("solid", fgColor="D9D9D9")
    bold        = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    right       = Alignment(horizontal="right", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    # Cabeceras - Nueva estructura simple
    headers = [
        "Fecha de Retiro",
        "Local",
        "Remesas del Día",
        "Remesas Fechas Anteriores",
        "Total Teórico a Recibir"
    ]

    ws.append(headers)

    # Header style
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.font = bold
        c.alignment = center
        c.fill = header_fill
        c.border = border

    # Filas
    for r in data["rows"]:
        ws.append([
            datetime.strptime(r["fecha"], "%Y-%m-%d").strftime("%d/%m/%Y"),
            r["local"],
            r["remesas_del_dia"],
            r["remesas_anteriores"],
            r["total_teorico"]
        ])

    # Formato numérico y bordes
    num_fmt = '#,##0.00'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for idx, cell in enumerate(row, start=1):
            cell.border = border
            if idx >= 3:  # columnas numéricas
                cell.number_format = num_fmt
                cell.alignment = right

    # Sombrear columna de Total Teórico (última columna)
    total_col = len(headers)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, total_col).fill = ap_fill

    # Totales
    ws.append([])
    totals_row = ws.max_row + 1
    tot = data["totals"]
    ws.append([
        "TOTAL GENERAL", "",
        tot["remesas_del_dia"],
        tot["remesas_anteriores"],
        tot["total_teorico"]
    ])

    for col in range(1, len(headers)+1):
        c = ws.cell(row=totals_row, column=col)
        c.font = bold
        c.border = border
        if col >= 3:  # num
            c.number_format = num_fmt
            c.alignment = right
        if headers[col-1] in ("Total FDS","Ap Proyectada"):
            c.fill = ap_fill

    # Anchos
    col_widths = {
        1: 12,  # Fecha Sello
        2: 26,  # Locales
    }
    for i in range(3, len(headers)+1):
        col_widths[i] = 16
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    # Salida
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"reporte_remesas_{fecha_sel.isoformat()}.xlsx"
    return send_file(bio, as_attachment=True,
                     download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")





## REGISTRO Y LOGIN NUEVO ________________________________

# utils_auth.py (o dentro de app.py si preferís archivo único)
# --- imports necesarios ---
import uuid, bcrypt
from datetime import datetime
from flask import render_template, request, jsonify, session
from functools import wraps
from mysql.connector import Error

# ---------- Decoradores mínimos ----------

# ---------- (Re)usar la función create_user ----------
def create_user(username, password_plain, role_name, local, society, pages_slugs, status='active'):
    """
    Crea usuario, asegura pages y relaciones. Devuelve dict con {user_id}
    """
    user_id = str(uuid.uuid4())
    pw_hash = bcrypt.hashpw(password_plain.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    conn = get_db_connection()
    try:
        cur = conn.cursor()

        # Verificar username único
        cur.execute("SELECT 1 FROM users WHERE username=%s", (username,))
        if cur.fetchone():
            raise ValueError("El usuario ya existe")

        # Role
        cur.execute("SELECT id FROM roles WHERE name=%s", (role_name,))
        row = cur.fetchone()
        if not row:
            raise ValueError(f"Rol '{role_name}' no existe (cajero|encargado|auditor|jefe_auditor|admin_anticipos)")
        role_id = row[0]

        # Insert user
        cur.execute("""
            INSERT INTO users (id, username, password, role_id, local, society, status, created_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (user_id, username, pw_hash, role_id, local, society, status, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))

        # Asegurar páginas y vincular
        for slug in pages_slugs:
            cur.execute("INSERT IGNORE INTO pages (slug) VALUES (%s)", (slug,))
            conn.commit()
            cur.execute("SELECT id FROM pages WHERE slug=%s", (slug,))
            pid = cur.fetchone()[0]
            cur.execute("INSERT IGNORE INTO user_pages (user_id, page_id) VALUES (%s,%s)", (user_id, pid))

        conn.commit()
        return {"user_id": user_id}
    except ValueError as ve:
        conn.rollback()
        raise ve
    except Error as e:
        conn.rollback()
        raise e
    finally:
        conn.close()

# ---------- UI: página de creación ----------
@app.route('/admin/create-user', methods=['GET'])
@login_required
@role_min_required(2)  # Encargado (2) o Auditor (3)
def ui_create_user():
    return render_template('create_user.html')

# ---------- API: crear usuario ----------
@app.route('/api/users', methods=['POST'])
@login_required
@role_min_required(2)
def api_create_user():
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '')
    role     = (data.get('role') or '').strip()
    status   = (data.get('status') or 'active').strip() or 'active'
    local    = (data.get('local') or '').strip()
    society  = (data.get('society') or '').strip()
    pages    = data.get('pages') or []

    # Validaciones base
    if not (username and password and role and local and society):
        return jsonify(success=False, msg='Faltan campos obligatorios'), 400
    if len(password) < 4:
        return jsonify(success=False, msg='La contraseña debe tener al menos 4 caracteres'), 400
    if role not in ('cajero','encargado','auditor','jefe_auditor','admin_anticipos'):
        return jsonify(success=False, msg='Rol inválido'), 400
    if status not in ('active','inactive'):
        return jsonify(success=False, msg='Estado inválido'), 400

    # Sanitizar pages
    pages = [str(p).strip() for p in pages if str(p).strip()]

    try:
        out = create_user(username, password, role, local, society, pages, status=status)
        return jsonify(success=True, user_id=out['user_id'])
    except ValueError as ve:
        return jsonify(success=False, msg=str(ve)), 409
    except Error as e:
        return jsonify(success=False, msg=f'Error de base de datos: {e}'), 500
    except Exception as e:
        return jsonify(success=False, msg=f'Error: {e}'), 500


# ---------- Login / Logout ----------

def _fetch_user_by_username(username):
    conn = get_db_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT u.id, u.username, u.password, u.local, u.society, u.status, u.first_login,
                   r.id AS role_id, r.name AS role_name, r.level AS role_level
            FROM users u
            JOIN roles r ON u.role_id = r.id
            WHERE u.username=%s
            LIMIT 1
        """, (username,))
        user = cur.fetchone()
        if not user:
            return None

        # Páginas por ROL (canon)
        cur.execute("""
            SELECT p.slug
            FROM role_pages rp
            JOIN pages p ON p.id = rp.page_id
            WHERE rp.role_id=%s
        """, (user['role_id'],))
        role_pages = {row['slug'] for row in cur.fetchall()}

        # (Opcional) extras por usuario — si no querés extras, dejá user_pages = set()
        cur.execute("""
            SELECT p.slug
            FROM user_pages up
            JOIN pages p ON up.page_id = p.id
            WHERE up.user_id=%s
        """, (user['id'],))
        user_pages = {row['slug'] for row in cur.fetchall()}

        user['pages'] = sorted(role_pages | user_pages)  # rol ∪ extras
        return user
    finally:
        conn.close()


def _update_last_access(user_id):
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("UPDATE users SET last_access=NOW() WHERE id=%s", (user_id,))
        conn.commit()
    finally:
        conn.close()


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        nxt = request.args.get('next', '')
        return render_template('login.html', error=None, next=nxt)

    # POST (form tradicional o JSON)
    data = request.form if request.form else (request.get_json(silent=True) or {})
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()

    user = _fetch_user_by_username(username)

    def _fail(msg='Usuario o contraseña inválidos'):
        if request.form:
            return render_template('login.html', error=msg), 401
        return jsonify(ok=False, msg=msg), 401

    # Verificar que el usuario esté activo
    if not user or user.get('status') != 'active':
        return _fail()

    # LÓGICA ESPECIAL PARA PRIMER LOGIN:
    # Si first_login=FALSE, cualquier contraseña es válida y se guarda como la definitiva
    is_first_login = (user.get('first_login') == 0 or user.get('first_login') == False)

    if is_first_login:
        # Primer login: aceptar cualquier contraseña ingresada y guardarla
        if not password or len(password) < 4:
            return _fail('La contraseña debe tener al menos 4 caracteres')

        # Hashear y guardar la contraseña que el usuario ingresó
        try:
            pw_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("""
                UPDATE users
                SET password=%s, first_login=TRUE
                WHERE id=%s
            """, (pw_hash, user['id']))
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            return _fail(f'Error al guardar contraseña: {str(e)}')
    else:
        # Login normal: verificar contraseña
        try:
            hashed = (user.get('password') or '').encode('utf-8')
            if not bcrypt.checkpw(password.encode('utf-8'), hashed):
                return _fail()
        except Exception:
            return _fail()

    # Sesión (mismas claves que usa tu app)
    session.clear()
    session.permanent     = True  # Usar PERMANENT_SESSION_LIFETIME configurado
    session['user_id']    = user['id']
    session['username']   = user['username']
    session['local']      = user.get('local')
    session['society']    = user.get('society')
    session['role']       = user.get('role_name')                 # 'cajero'|'encargado'|'auditor'|'jefe_auditor'
    session['role_level'] = int(user.get('role_level') or 1)      # 1|2|3|4
    session['pages']      = user.get('pages') or []
    _update_last_access(user['id'])

    # Si vino por HTML -> redirigir respetando next y rol
    if request.form:
        return redirect_after_login()

    # Si vino por API -> sugerir destino
    return jsonify(ok=True, redirect=route_for_current_role())


@app.route('/primer-login')
@login_required
def primer_login():
    """Página para cambiar contraseña en el primer login"""
    return render_template('primer_login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


## __________________________ FIN LOGIN NUEVO ________________________








## ____________ endpoints nivel 2 ______________________

# === ruta nueva (Nivel 2) ===
@app.route('/encargado', endpoint='carga_datos_encargado')
@role_min_required(2)  # encargado+ crea usuarios
@login_required
def carga_datos_encargado():
    # seguridad de acceso por rol: mínimo 2
    if get_user_level() < 2:
        return jsonify(success=False, msg='Acceso sólo para nivel 2+'), 403

    local = session.get('local')
    conn = get_db_connection()
    cur  = conn.cursor()

    # cajas disponibles para el local
    cur.execute("SELECT cantidad_cajas FROM locales WHERE local = %s LIMIT 1", (local,))
    row = cur.fetchone()
    cantidad_cajas = row[0] if row else 1

    # turnos habilitados
    cur.execute("SELECT turnos FROM locales WHERE local = %s", (local,))
    turnos = [r[0] for r in cur.fetchall()] or ['UNI']

    cur.close(); conn.close()

    # Renderiza copia del index (tu HTML clon)
    return render_template('index_encargado.html',
                           cantidad_cajas=cantidad_cajas,
                           turnos=turnos,
                           role_level=get_user_level())













def get_turnos_del_dia(conn, local: str, fecha):
    """
    Devuelve la lista de turnos existentes para (local, fecha).
    Fuente principal: cajas_estado. Si no hay, intenta ver en tablas operativas.
    """
    f = _normalize_fecha(fecha)
    cur = conn.cursor()
    turnos = set()

    # Fuente principal: cajas_estado (más confiable para cierres)
    cur.execute("""
        SELECT DISTINCT turno
          FROM cajas_estado
         WHERE local=%s AND DATE(fecha_operacion)=%s
           AND turno IS NOT NULL AND turno <> ''
    """, (local, f))
    for (t,) in cur.fetchall() or []:
        turnos.add(str(t))

    # Fallbacks por si no hay filas en cajas_estado (raro pero posible)
    if not turnos:
        tablas = [
            ("remesas_trns",     "fecha"),
            ("tarjetas_trns",    "fecha"),
            ("mercadopago_trns", "fecha"),
            ("gastos_trns",      "fecha"),
            ("rappi_trns",       "fecha"),
            ("pedidosya_trns",   "fecha"),
            ("ventas_base",      "fecha"),
            ("ventas_z",         "fecha"),
            ("tips_tarjetas",    "fecha"),
        ]
        for tabla, col_fecha in tablas:
            try:
                cur.execute(f"""
                    SELECT DISTINCT turno
                      FROM {tabla}
                     WHERE local=%s AND DATE({col_fecha})=%s
                       AND turno IS NOT NULL AND turno <> ''
                """, (local, f))
                for (t,) in cur.fetchall() or []:
                    turnos.add(str(t))
            except Exception:
                # Si alguna tabla no existe / no tiene turno en tu entorno, la salteamos
                pass

    cur.close()
    return sorted(turnos)
def create_snapshot_for_local_by_day(conn, local: str, fecha, made_by: str):
    """
    Genera snapshots por CADA turno existente en el día.
    Usa la función ya existente: create_snapshot_for_local(conn, local, fecha, turno, made_by)
    """
    f = _normalize_fecha(fecha)
    turnos = get_turnos_del_dia(conn, local, f)

    # Si no detectamos turnos, igualmente intentamos con uno "UNI" (por si tu operación es de turno único)
    if not turnos:
        turnos = ["UNI"]

    for turno in turnos:
        # Tu función ya existente; no la reescribimos
        create_snapshot_for_local(conn, local, f, turno, made_by=made_by)



def _validar_imagenes_antes_cierre_local(conn, local, fecha):
    """
    Valida que para cada caja/turno que tenga datos cargados,
    exista al menos una imagen adjuntada en la pestaña correspondiente.

    NOTA: Rappi y PedidosYa NO son obligatorios (a veces no hay).

    Retorna (ok: bool, msg: str, faltantes: list)
    """
    cur = conn.cursor(dictionary=True)
    f = _normalize_fecha(fecha)
    faltantes = []

    try:
        # Obtener todas las cajas/turnos del local ese día
        cur.execute("""
            SELECT DISTINCT caja, turno
            FROM (
                SELECT caja, turno FROM remesas_trns WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM mercadopago_trns WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM tarjetas_trns WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM gastos_trns WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM facturas_trns WHERE local=%s AND DATE(fecha)=%s
            ) AS all_data
            ORDER BY caja, turno
        """, (local, f) * 5)

        cajas_turnos = cur.fetchall() or []

        # Para cada caja/turno, validar cada pestaña (EXCEPTO rappi y pedidosya)
        for ct in cajas_turnos:
            caja = ct['caja']
            turno = ct['turno']

            # Mapeo: tabla -> config de pestaña
            # IMPORTANTE: Rappi y PedidosYa NO están aquí (no son obligatorios)
            pestanas = {
                'remesas_trns': {'tabs': ['remesas'], 'nombre': 'Remesas'},
                'mercadopago_trns': {'tabs': ['mercadopago'], 'nombre': 'Mercado Pago'},
                'tarjetas_trns': {'tabs': ['tarjetas'], 'nombre': 'Tarjetas'},
                'gastos_trns': {'tabs': ['gastos'], 'nombre': 'Gastos'},
                'facturas_trns': {'tabs': ['ventas_z'], 'nombre': 'Facturas'}
            }

            for tabla, config in pestanas.items():
                # Verificar si hay datos en esta tabla
                cur.execute(f"""
                    SELECT COUNT(*) as cnt
                    FROM {tabla}
                    WHERE local=%s AND caja=%s AND turno=%s AND DATE(fecha)=%s
                """, (local, caja, turno, f))

                row = cur.fetchone()
                tiene_datos = row and row['cnt'] > 0

                if tiene_datos:
                    # Verificar si hay imágenes adjuntadas (buscar con cualquiera de las variaciones del tab)
                    tab_variants = config['tabs']
                    placeholders = ','.join(['%s'] * len(tab_variants))

                    cur.execute(f"""
                        SELECT COUNT(*) as cnt
                        FROM imagenes_adjuntos
                        WHERE tab IN ({placeholders})
                          AND local=%s AND caja=%s AND turno=%s AND DATE(fecha)=%s
                          AND estado='active'
                    """, (*tab_variants, local, caja, turno, f))

                    row_img = cur.fetchone()
                    tiene_imagenes = row_img and row_img['cnt'] > 0

                    if not tiene_imagenes:
                        faltantes.append({
                            'caja': caja,
                            'turno': turno,
                            'pestana': config['nombre'],
                            'tab': tab_variants[0]  # usar el primero como referencia
                        })

        if faltantes:
            # Construir mensaje amigable
            msgs = []
            for f in faltantes:
                msgs.append(f"{f['pestana']} (Caja: {f['caja']}, Turno: {f['turno']})")

            msg = "Falta subir imágenes en:\n• " + "\n• ".join(msgs)
            return False, msg, faltantes

        return True, "", []

    finally:
        cur.close()


## ______________________ ACTUALIZAR OBSERVACIÓN DE DIFERENCIA _______________________
@app.route('/api/actualizar_observacion_diferencia', methods=['POST'])
@login_required
@role_min_required(2)  # Solo nivel 2 (encargado)
def api_actualizar_observacion_diferencia():
    """
    Permite a nivel 2 editar la observación de una diferencia
    SOLO cuando el local está abierto.
    """
    data = request.get_json() or {}
    local = session.get('local')
    fecha = data.get('fecha')
    caja = data.get('caja')
    turno = data.get('turno')
    observacion = data.get('observacion', '').strip()

    if not (local and fecha and caja and turno):
        return jsonify(success=False, msg='Faltan parámetros (local, fecha, caja, turno)'), 400

    f = _normalize_fecha(fecha)
    if not f:
        return jsonify(success=False, msg='Fecha inválida'), 400

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    try:
        # Verificar que el local NO esté cerrado
        cur.execute("""
            SELECT COUNT(*) as cnt
            FROM cierres_locales
            WHERE local=%s AND DATE(fecha)=%s AND estado=0
        """, (local, f))
        row = cur.fetchone()
        if row and row['cnt'] > 0:
            cur.close()
            conn.close()
            return jsonify(success=False, msg='❌ El local está cerrado. No se puede editar la observación.'), 403

        # Actualizar la observación en cajas_estado
        cur.execute("""
            UPDATE cajas_estado
            SET observacion = %s
            WHERE local=%s AND caja=%s AND turno=%s AND DATE(fecha_operacion)=%s
        """, (observacion, local, caja, turno, f))

        conn.commit()
        affected = cur.rowcount

        cur.close()
        conn.close()

        if affected > 0:
            return jsonify(success=True, msg='Observación actualizada')
        else:
            return jsonify(success=False, msg='No se encontró el registro de caja para actualizar'), 404

    except Exception as e:
        try:
            conn.rollback()
        except:
            pass
        try:
            cur.close()
        except:
            pass
        try:
            conn.close()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


## ______________________ CIERRE LOCAL _______________________
@app.route('/api/cierre_local', methods=['POST'])
@login_required
@role_min_required(2)  # mínimo L2
def api_cierre_local():
    data  = request.get_json() or {}
    # Para auditores: usar local del body, sino usar get_local_param()
    local = data.get('local') or get_local_param()
    fecha = data.get('fecha')

    if not (local and fecha):
        return jsonify(success=False, msg='falta local/fecha'), 400

    f = _normalize_fecha(fecha)
    if not f:
        return jsonify(success=False, msg='fecha inválida'), 400

    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)
    try:
        # (1) PRIMERO: Validar que existan imágenes para cada pestaña con datos
        ok_imgs, msg_imgs, faltantes_imgs = _validar_imagenes_antes_cierre_local(conn, local, fecha)
        if not ok_imgs:
            return jsonify(success=False, msg=msg_imgs, faltantes=faltantes_imgs), 400

        # (1.5) Validar que NO haya anticipos pendientes sin consumir
        cur.execute("""
            SELECT
                ar.id, ar.cliente, ar.importe, ar.numero_transaccion,
                ar.observaciones
            FROM anticipos_recibidos ar
            WHERE ar.local = %s
              AND ar.fecha_evento = %s
              AND ar.estado = 'pendiente'
        """, (local, f))
        anticipos_pendientes = cur.fetchall() or []

        if anticipos_pendientes:
            lista_anticipos = []
            for a in anticipos_pendientes:
                cliente = a['cliente']
                importe = f"${a['importe']:,.2f}".replace(',', '.')
                lista_anticipos.append(f"{cliente} ({importe})")

            msg_anticipos = (f"No se puede cerrar el local: hay {len(anticipos_pendientes)} anticipo(s) sin consumir:\n" +
                           "\n".join(f"• {ant}" for ant in lista_anticipos))
            return jsonify(success=False, msg=msg_anticipos, anticipos_pendientes=anticipos_pendientes), 400

        # (2) SEGUNDO: Verificar que TODAS las cajas estén cerradas (estado=0)
        # Obtener todas las cajas/turnos que tienen datos ese día
        cur.execute("""
            SELECT DISTINCT caja, turno
            FROM (
                SELECT caja, turno FROM ventas_trns WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM remesas_trns WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM tarjetas_trns WHERE local=%s AND DATE(fecha)=%s
                UNION
                SELECT caja, turno FROM mercadopago_trns WHERE local=%s AND DATE(fecha)=%s
            ) AS todas_cajas
            ORDER BY caja, turno
        """, (local, f, local, f, local, f, local, f))
        cajas_con_datos = cur.fetchall() or []

        # Verificar cuáles están cerradas
        cajas_sin_cerrar = []
        for row in cajas_con_datos:
            caja = row['caja']
            turno = row['turno']

            # Verificar si existe un cierre para esta caja/turno
            cur.execute("""
                SELECT estado
                FROM cajas_estado
                WHERE local=%s AND caja=%s AND turno=%s AND DATE(fecha_operacion)=%s
                ORDER BY id DESC
                LIMIT 1
            """, (local, caja, turno, f))
            cierre = cur.fetchone()

            # Si no existe cierre o está abierta (estado=1), agregarla a la lista
            if not cierre or cierre['estado'] == 1:
                cajas_sin_cerrar.append(f"{caja} - {turno}")

        if cajas_sin_cerrar:
            msg = f"Todas las cajas deben estar cerradas antes de cerrar el local. Faltan cerrar: {', '.join(cajas_sin_cerrar)}"
            return jsonify(success=False, msg=msg, cajas_sin_cerrar=cajas_sin_cerrar), 409

        # (2) Marcar/crear cierre de local (sin columna turno)
        cur.execute("""
          INSERT INTO cierres_locales (local, fecha, estado, closed_by, closed_at)
          VALUES (%s,%s,0,%s,NOW())
          ON DUPLICATE KEY UPDATE
            estado=VALUES(estado),
            closed_by=VALUES(closed_by),
            closed_at=VALUES(closed_at)
        """, (local, f, session.get('username')))

        # (3) Pasar estado='ok' en todas las tablas operativas del día (sin turno)
        updates = [
          ("remesas_trns",     "UPDATE remesas_trns     SET estado='ok' WHERE local=%s AND DATE(fecha)=%s"),
          ("mercadopago_trns", "UPDATE mercadopago_trns SET estado='ok' WHERE local=%s AND DATE(fecha)=%s"),
          ("tarjetas_trns",    "UPDATE tarjetas_trns    SET estado='ok' WHERE local=%s AND DATE(fecha)=%s"),
          ("gastos_trns",      "UPDATE gastos_trns      SET estado='ok' WHERE local=%s AND DATE(fecha)=%s"),
          ("rappi_trns",       "UPDATE rappi_trns       SET estado='ok' WHERE local=%s AND DATE(fecha)=%s"),
          ("pedidosya_trns",   "UPDATE pedidosya_trns   SET estado='ok' WHERE local=%s AND DATE(fecha)=%s"),
          ("ventas_base",      "UPDATE ventas_base      SET estado='ok' WHERE local=%s AND DATE(fecha)=%s"),
          ("ventas_z",         "UPDATE ventas_z         SET estado='ok' WHERE local=%s AND DATE(fecha)=%s"),
          ("tips_tarjetas",    "UPDATE tips_tarjetas    SET estado='ok' WHERE local=%s AND DATE(fecha)=%s"),
        ]
        for _, sql in updates:
            try:
                cur.execute(sql, (local, f))
            except Exception:
                # Por si tenés alguna tabla aún no creada en todos los entornos
                pass

        # (4) Congelar snapshot por CADA turno del día (usa tu función existente con 'turno')
        create_snapshot_for_local_by_day(conn, local, f, made_by=session.get('username'))

        conn.commit()
        return jsonify(success=True, msg="Cierre del local confirmado y snapshots por turno creados")
    except Exception as e:
        conn.rollback()
        print("❌ cierre_local:", e)
        return jsonify(success=False, msg=str(e)), 500
    finally:
        try: cur.close()
        except: ...
        try: conn.close()
        except: ...


## ______________________ MARCAR COMO AUDITADO _______________________
@app.route('/api/marcar_auditado', methods=['POST'])
@login_required
@role_min_required(3)  # solo auditor (L3+)
def api_marcar_auditado():
    """
    Marca un local como auditado para una fecha específica.
    El local debe estar cerrado para poder ser marcado como auditado.
    Una vez auditado, el local se vuelve inmutable (nadie puede editar).
    """
    data = request.get_json() or {}
    # Para auditores: usar local del body, sino usar get_local_param()
    local = data.get('local') or get_local_param()
    fecha = data.get('fecha')
    observaciones = data.get('observaciones', '')

    if not (local and fecha):
        return jsonify(success=False, msg='Faltan parámetros: local y fecha son requeridos'), 400

    f = _normalize_fecha(fecha)
    if not f:
        return jsonify(success=False, msg='Fecha inválida (formato esperado YYYY-MM-DD)'), 400

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # Verificar que el local esté cerrado
        cur.execute("""
            SELECT estado FROM cierres_locales
            WHERE local=%s AND DATE(fecha)=%s
        """, (local, f))
        row = cur.fetchone()

        if not row:
            cur.close()
            conn.close()
            return jsonify(success=False, msg='El local no está cerrado para esta fecha'), 400

        if row['estado'] != 0:
            cur.close()
            conn.close()
            return jsonify(success=False, msg='El local debe estar cerrado (estado=0) para ser auditado'), 400

        # Verificar si ya está auditado
        cur.execute("""
            SELECT id FROM locales_auditados
            WHERE local=%s AND DATE(fecha)=%s
        """, (local, f))

        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify(success=False, msg='Este local ya está marcado como auditado'), 409

        # NUEVO: Verificar que no haya anticipos pendientes
        cur.execute("""
            SELECT
                ar.id, ar.cliente, ar.importe, ar.numero_transaccion,
                ar.observaciones
            FROM anticipos_recibidos ar
            WHERE ar.local = %s
              AND ar.fecha_evento = %s
              AND ar.estado = 'pendiente'
        """, (local, f))
        anticipos_pendientes = cur.fetchall() or []

        if anticipos_pendientes:
            cur.close()
            conn.close()
            # Construir mensaje detallado
            msg_anticipos = []
            for ant in anticipos_pendientes:
                msg_anticipos.append(f"• {ant['cliente']} - ${ant['importe']} (ID: {ant['id']})")

            msg = (
                f"❌ No se puede marcar como auditado. Hay {len(anticipos_pendientes)} anticipo(s) sin consumir:\n\n" +
                "\n".join(msg_anticipos) +
                "\n\nDebés consumir o eliminar todos los anticipos antes de auditar."
            )
            return jsonify(
                success=False,
                msg=msg,
                anticipos_pendientes=anticipos_pendientes
            ), 400

        # Marcar como auditado
        cur.execute("""
            INSERT INTO locales_auditados
            (local, fecha, auditado_por, fecha_auditoria, observaciones)
            VALUES (%s, %s, %s, NOW(), %s)
        """, (local, f, session.get('username'), observaciones))

        conn.commit()

        # ========== SINCRONIZACIÓN AUTOMÁTICA CON OPPEN ==========
        # Enviar facturas a Oppen después de marcar como auditado
        try:
            from modules.oppen_integration import sync_facturas_to_oppen, sync_recibo_to_oppen

            print(f"🔄 Iniciando sincronización de facturas con Oppen para {local} - {f}...")
            resultado_oppen = sync_facturas_to_oppen(conn, local, f)

            # Construir mensaje de resultado
            msg_base = f"Local {local} marcado como auditado para {f}"

            if resultado_oppen['success']:
                if resultado_oppen['total'] > 0:
                    msg_oppen = f"\n✅ {resultado_oppen['exitosas']} factura(s) enviada(s) a Oppen exitosamente"
                else:
                    msg_oppen = "\nℹ️ No había facturas para sincronizar con Oppen"
            else:
                msg_oppen = f"\n⚠️ Algunas facturas no pudieron enviarse a Oppen: {resultado_oppen['fallidas']}/{resultado_oppen['total']}"
                if resultado_oppen['errores']:
                    # Mostrar primer error como ejemplo
                    primer_error = resultado_oppen['errores'][0]
                    msg_oppen += f"\nPrimer error: {primer_error.get('error', 'Error desconocido')}"

            # ========== CREAR RECIBO EN OPPEN ==========
            # Después de crear facturas, crear el recibo vinculando las facturas Z
            msg_recibo = ""
            resultado_recibo = None

            if resultado_oppen['success'] and resultado_oppen['exitosas'] > 0:
                try:
                    print(f"🔄 Creando recibo en Oppen para {local} - {f}...")
                    resultado_recibo = sync_recibo_to_oppen(conn, local, f)

                    if resultado_recibo['recibo_creado']:
                        msg_recibo = f"\n✅ Recibo creado en Oppen: {resultado_recibo['message']}"
                        if resultado_recibo.get('sernr'):
                            msg_recibo += f" (SerNr: {resultado_recibo['sernr']})"
                    else:
                        msg_recibo = f"\nℹ️ {resultado_recibo['message']}"

                except Exception as e:
                    logger.error(f"Error creando recibo: {e}")
                    msg_recibo = f"\n⚠️ Error creando recibo: {str(e)}"

            cur.close()
            conn.close()

            return jsonify(
                success=True,
                msg=msg_base + msg_oppen + msg_recibo,
                oppen_sync=resultado_oppen,
                recibo_sync=resultado_recibo
            )

        except ImportError as e:
            # Si falla la importación del módulo, continuar sin Oppen
            print(f"⚠️ Módulo de Oppen no disponible: {e}")
            cur.close()
            conn.close()
            return jsonify(
                success=True,
                msg=f"Local {local} marcado como auditado para {f} (sincronización con Oppen no disponible)"
            )
        except Exception as e:
            # Si falla Oppen, el local YA ESTÁ AUDITADO (commit ya se hizo)
            # Solo notificamos el error de Oppen
            print(f"❌ Error sincronizando con Oppen: {e}")
            cur.close()
            conn.close()
            return jsonify(
                success=True,
                msg=f"Local {local} marcado como auditado para {f}\n⚠️ Error sincronizando con Oppen: {str(e)}",
                oppen_error=str(e)
            )

    except Exception as e:
        conn.rollback()
        print("❌ marcar_auditado:", e)
        return jsonify(success=False, msg=str(e)), 500
    finally:
        try: cur.close()
        except: ...
        try: conn.close()
        except: ...


@app.route('/api/estado_auditoria', methods=['GET'])
@login_required
def api_estado_auditoria():
    """
    Verifica si un local está auditado para una fecha específica.
    Retorna: { auditado: bool, info: {...} }
    """
    local = get_local_param()
    fecha = request.args.get('fecha')

    if not (local and fecha):
        return jsonify(success=False, msg='Faltan parámetros: local y fecha'), 400

    f = _normalize_fecha(fecha)
    if not f:
        return jsonify(success=False, msg='Fecha inválida'), 400

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT id, local, fecha, auditado_por, fecha_auditoria, observaciones
            FROM locales_auditados
            WHERE local=%s AND DATE(fecha)=%s
        """, (local, f))

        row = cur.fetchone()

        if row:
            return jsonify(success=True, auditado=True, info=row)
        else:
            return jsonify(success=True, auditado=False, info=None)

    except Exception as e:
        print("❌ estado_auditoria:", e)
        return jsonify(success=False, msg=str(e)), 500
    finally:
        try: cur.close()
        except: ...
        try: conn.close()
        except: ...


@app.get("/healthz")
def healthz():
    return {"ok": True}, 200

def create_snapshot_for_local(conn, local:str, fecha, turno:str, made_by:str):
    """
    Congela el estado 'aceptado por L2' (lo que hay en tablas operativas
    tras pasar a estado='ok') en tablas snap_* para (local, fecha, turno).
    Idempotente: si existe snapshot, lo reutiliza y repuebla su contenido.
    """
    f = _normalize_fecha(fecha)
    cur = conn.cursor()

    # Cabecera: upsert que preserve el id para reinsertar detalles
    cur.execute("""
        INSERT INTO cierre_snapshots (local, fecha, turno, made_at, made_by)
        VALUES (%s,%s,%s,NOW(),%s)
        ON DUPLICATE KEY UPDATE made_at=VALUES(made_at), made_by=VALUES(made_by)
    """, (local, f, turno, made_by))

    # Obtener id (si ya existía, lo recuperamos)
    cur.execute("SELECT id FROM cierre_snapshots WHERE local=%s AND fecha=%s AND turno=%s",
                (local, f, turno))
    snapshot_id = cur.fetchone()[0]

    # Limpiamos detalles previos de ese snapshot para repoblar
    for tbl in (
        "snap_remesas",
        "snap_tarjetas",
        "snap_mercadopago",
        "snap_ventas",
        "snap_facturas",
        "snap_gastos",
        "snap_rappi",
        "snap_pedidosya",
    ):
        cur.execute(f"DELETE FROM {tbl} WHERE snapshot_id=%s", (snapshot_id,))

    # --- Remesas ---
    cur.execute("""
        INSERT INTO snap_remesas
        (snapshot_id, id_src, usuario, local, caja, turno, fecha, nro_remesa, precinto, monto,
         retirada, retirada_por, fecha_retirada, ult_mod, estado)
        SELECT %s, r.id, r.usuario, r.local, r.caja, r.turno, DATE(r.fecha), r.nro_remesa,
               r.precinto, r.monto, r.retirada, r.retirada_por, r.fecha_retirada, r.ult_mod, r.estado
          FROM remesas_trns r
         WHERE r.local=%s AND DATE(r.fecha)=%s AND r.turno=%s
    """, (snapshot_id, local, f, turno))

    # --- Tarjetas ---
    cur.execute("""
        INSERT INTO snap_tarjetas
        (snapshot_id, id_src, usuario, local, caja, turno, fecha, tarjeta, terminal, lote, monto, monto_tip, estado)
        SELECT %s, t.id, t.usuario, t.local, t.caja, t.turno, DATE(t.fecha), t.tarjeta, t.terminal, t.lote, t.monto, t.monto_tip, t.estado
          FROM tarjetas_trns t
         WHERE t.local=%s AND DATE(t.fecha)=%s AND t.turno=%s
    """, (snapshot_id, local, f, turno))

    # --- Mercado Pago ---
    cur.execute("""
        INSERT INTO snap_mercadopago
        (snapshot_id, id_src, usuario, local, caja, turno, fecha, tipo, terminal, comprobante, importe, estado)
        SELECT %s, m.id, m.usuario, m.local, m.caja, m.turno, DATE(m.fecha),
               UPPER(m.tipo), m.terminal, m.comprobante, m.importe, m.estado
          FROM mercadopago_trns m
         WHERE m.local=%s AND DATE(m.fecha)=%s AND m.turno=%s
    """, (snapshot_id, local, f, turno))

    # --- Ventas (base) ---
    cur.execute("""
        INSERT INTO snap_ventas
        (snapshot_id, id_src, usuario, local, caja, turno, fecha, venta_total_sistema, estado)
        SELECT %s, v.id, v.usuario, v.local, v.caja, v.turno, DATE(v.fecha), v.venta_total_sistema, v.estado
          FROM ventas_trns v
         WHERE v.local=%s AND DATE(v.fecha)=%s AND v.turno=%s
    """, (snapshot_id, local, f, turno))

    # --- Facturas (Z, A, B, CC) ---
    cur.execute("""
        INSERT INTO snap_facturas
        (snapshot_id, id_src, local, caja, turno, fecha, tipo, punto_venta, nro_factura, monto, estado)
        SELECT %s, f.id, f.local, f.caja, f.turno, DATE(f.fecha), f.tipo, f.punto_venta, f.nro_factura, f.monto, f.estado
          FROM facturas_trns f
         WHERE f.local=%s AND DATE(f.fecha)=%s AND f.turno=%s
    """, (snapshot_id, local, f, turno))

    # --- Gastos ---
    cur.execute("""
        INSERT INTO snap_gastos
        (snapshot_id, id_src, local, caja, turno, fecha, tipo, monto, observaciones, estado)
        SELECT %s, g.id, g.local, g.caja, g.turno, DATE(g.fecha), g.tipo, g.monto, g.observaciones, g.estado
          FROM gastos_trns g
         WHERE g.local=%s AND DATE(g.fecha)=%s AND g.turno=%s
    """, (snapshot_id, local, f, turno))

    # --- Rappi ---
    cur.execute("""
        INSERT INTO snap_rappi
        (snapshot_id, id_src, local, caja, turno, fecha, transaccion, monto, estado)
        SELECT %s, r.id, r.local, r.caja, r.turno, DATE(r.fecha), r.transaccion, r.monto, r.estado
          FROM rappi_trns r
         WHERE r.local=%s AND DATE(r.fecha)=%s AND r.turno=%s
    """, (snapshot_id, local, f, turno))

    # --- PedidosYa ---
    cur.execute("""
        INSERT INTO snap_pedidosya
        (snapshot_id, id_src, local, caja, turno, fecha, transaccion, monto, estado)
        SELECT %s, r.id, r.local, r.caja, r.turno, DATE(r.fecha), r.transaccion, r.monto, r.estado
          FROM pedidosya_trns r
         WHERE r.local=%s AND DATE(r.fecha)=%s AND r.turno=%s
    """, (snapshot_id, local, f, turno))

    cur.close()
    # commit lo hace quien llama



@app.get("/estado_local")
@login_required
def estado_local():
    # Permitir que el auditor pase 'local' por query string
    local = request.args.get("local") or get_local_param()
    fecha = request.args.get("fecha") or session.get('fecha')

    if not (local and fecha):
        return jsonify(ok=False, msg="Faltan parámetros"), 400

    fecha_normalizada = _normalize_fecha(fecha)
    app.logger.info(f"[estado_local] Consultando: local={local}, fecha={fecha_normalizada}")

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Lógica simple: si existe un registro en cierres_locales => CERRADO (estado=0)
        # Si NO existe registro => ABIERTO (estado=1)
        cur.execute("""
            SELECT COUNT(*)
            FROM cierres_locales
            WHERE local=%s AND fecha=%s
        """, (local, fecha_normalizada))
        row = cur.fetchone()
        count = row[0] if row else 0

        # Si existe al menos 1 registro => local CERRADO (estado=0)
        # Si NO existe registro => local ABIERTO (estado=1)
        estado = 0 if count > 0 else 1

        app.logger.info(f"[estado_local] Resultado: count={count}, estado={estado} ({'CERRADO' if estado == 0 else 'ABIERTO'})")

        return jsonify(ok=True, estado=estado)
    finally:
        cur.close(); conn.close()


def crear_snapshot_local(conn, local, fecha, usuario):
    """
    Congela (copia) la foto del día por local a tablas *snap*.
    Idempotente: podés borrar e insertar, o usar UPSERT por (local, fecha, ...).
    """
    # --- REMESAS (ejemplo simple) ---
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO remesas_snap (local, fecha, caja, turno, nro_remesa, precinto, monto, retirada, retirada_por, fuente_usuario, ts_snap)
        SELECT local, DATE(fecha) AS fecha, caja, turno, nro_remesa, precinto, monto, retirada, retirada_por, %s, NOW()
        FROM remesas_trns
        WHERE local=%s AND DATE(fecha)=%s
        ON DUPLICATE KEY UPDATE
          monto=VALUES(monto), retirada=VALUES(retirada), retirada_por=VALUES(retirada_por), ts_snap=VALUES(ts_snap)
    """, (usuario, local, fecha))
    cur.close()

    # --- TARJETAS (ejemplo simple) ---
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO tarjetas_snap (local, fecha, caja, turno, tarjeta, terminal, lote, monto, fuente_usuario, ts_snap)
        SELECT local, DATE(fecha), caja, turno, tarjeta, terminal, lote, monto, %s, NOW()
        FROM tarjetas_trns
        WHERE local=%s AND DATE(fecha)=%s
        ON DUPLICATE KEY UPDATE
          monto=VALUES(monto), ts_snap=VALUES(ts_snap)
    """, (usuario, local, fecha))
    cur.close()

    # --- TIPS TARJETAS (opcional) ---
    cur = conn.cursor()
    cur.execute(f"""
        INSERT INTO tips_tarjetas_snap
          (local, fecha, caja, turno, terminal, lote, {", ".join(_ALL_TIP_COLS)}, fuente_usuario, ts_snap)
        SELECT local, DATE(fecha), caja, turno, terminal, lote, {", ".join(_ALL_TIP_COLS)}, %s, NOW()
        FROM tips_tarjetas
        WHERE local=%s AND DATE(fecha)=%s
        ON DUPLICATE KEY UPDATE
          ts_snap=VALUES(ts_snap) {"" if not _ALL_TIP_COLS else ","}
          {", ".join([f"{c}=VALUES({c})" for c in _ALL_TIP_COLS])}
    """, (usuario, local, fecha))
    cur.close()

    # >>> Agregá aquí otras fuentes (MP, Rappi, Gastos, etc.) con el mismo patrón.


# ========================================================================
# ENDPOINT TEMPORAL: Aumentar tamaño de columna 'tab' en imagenes_adjuntos
# 

## ______________________ GESTIÓN DE USUARIOS _______________________

@app.route('/gestion-usuarios')
@login_required
@role_min_required(2)  # Mínimo nivel 2 (encargado)
def gestion_usuarios():
    """
    Página de gestión de usuarios.
    Si el usuario es admin_anticipos (nivel 6), redirigir a gestión de usuarios de anticipos.
    """
    user_level = get_user_level()
    if user_level == 6:
        # Admin de anticipos ve la página específica de gestión de usuarios de anticipos
        return render_template('gestion_usuarios_anticipos.html')
    # Otros niveles ven la página normal
    return render_template('gestion_usuarios.html')


@app.route('/gestion-anticipos')
@login_required
@role_min_required(3)  # Accesible para auditores (nivel 3) y superiores
def gestion_anticipos_page():
    """
    Página de gestión de anticipos recibidos.
    Accesible para:
    - auditores (nivel 3): pueden consumir/desconsumir si local no está auditado
    - anticipos (nivel 4): solo ve y crea anticipos en sus locales asignados
    - admin_anticipos (nivel 6): gestiona todos los anticipos
    """
    return render_template('gestion_anticipos.html')


@app.route('/api/mi_nivel', methods=['GET'])
@login_required
def api_mi_nivel():
    """Devuelve el nivel y local del usuario actual"""
    return jsonify(
        level=get_user_level(),
        local=session.get('local', '')
    )


@app.route('/api/mi_perfil_anticipos', methods=['GET'])
@login_required
def api_mi_perfil_anticipos():
    """
    Devuelve el perfil del usuario actual para anticipos.
    Incluye:
    - level: nivel del usuario
    - allowed_locales: locales a los que tiene acceso ([] significa todos)
    - can_edit: si puede editar anticipos
    - can_delete: si puede eliminar anticipos
    - can_consume: si puede consumir/desconsumir anticipos
    """
    user_level = get_user_level()
    allowed_locales = get_user_allowed_locales()

    # Auditores (nivel 3+) y admin_anticipos (nivel 6+) ven todos los locales
    if user_level >= 3:
        allowed_locales = []  # [] = acceso a todos los locales

    return jsonify(
        success=True,
        level=user_level,
        allowed_locales=allowed_locales,
        can_edit=(user_level >= 6),  # Solo admin_anticipos puede editar
        can_delete=(user_level >= 6),  # Solo admin_anticipos puede eliminar
        can_consume=(user_level >= 3),  # Auditores pueden consumir/desconsumir si local no está auditado
        has_full_access=(user_level >= 6)  # Admin tiene acceso total
    )


@app.route('/api/usuarios/listar', methods=['GET'])
@login_required
@role_min_required(2)
def api_usuarios_listar():
    """
    Lista usuarios que el usuario actual puede ver:
    - Nivel 2 (encargado): ve cajeros de su local
    - Nivel 3 (auditor): ve encargados
    - Nivel 4 (jefe_auditor): ve auditores
    """
    user_level = get_user_level()
    user_local = session.get('local')

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    try:
        # Determinar qué roles puede ver según el nivel
        if user_level == 2:
            # Encargado ve cajeros de su local
            cur.execute("""
                SELECT u.id, u.username, r.name as role, u.local, u.society, u.first_login,
                       CASE r.name
                           WHEN 'cajero' THEN 1
                           WHEN 'encargado' THEN 2
                           WHEN 'auditor' THEN 3
                           WHEN 'jefe_auditor' THEN 4
                           ELSE 0
                       END as role_level
                FROM users u
                JOIN roles r ON u.role_id = r.id
                WHERE r.name = 'cajero' AND u.local = %s
                ORDER BY u.username
            """, (user_local,))
        elif user_level == 3:
            # Auditor ve encargados
            cur.execute("""
                SELECT u.id, u.username, r.name as role, u.local, u.society, u.first_login,
                       CASE r.name
                           WHEN 'cajero' THEN 1
                           WHEN 'encargado' THEN 2
                           WHEN 'auditor' THEN 3
                           WHEN 'jefe_auditor' THEN 4
                           ELSE 0
                       END as role_level
                FROM users u
                JOIN roles r ON u.role_id = r.id
                WHERE r.name = 'encargado'
                ORDER BY u.username
            """)
        elif user_level >= 4:
            # Jefe Auditor ve auditores
            cur.execute("""
                SELECT u.id, u.username, r.name as role, u.local, u.society, u.first_login,
                       CASE r.name
                           WHEN 'cajero' THEN 1
                           WHEN 'encargado' THEN 2
                           WHEN 'auditor' THEN 3
                           WHEN 'jefe_auditor' THEN 4
                           ELSE 0
                       END as role_level
                FROM users u
                JOIN roles r ON u.role_id = r.id
                WHERE r.name = 'auditor'
                ORDER BY u.username
            """)
        else:
            cur.close()
            conn.close()
            return jsonify(success=False, msg='Sin permisos'), 403

        users = cur.fetchall() or []
        cur.close()
        conn.close()

        return jsonify(success=True, users=users)

    except Exception as e:
        try:
            cur.close()
        except:
            pass
        try:
            conn.close()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/usuarios/crear', methods=['POST'])
@login_required
@role_min_required(2)
def api_usuarios_crear():
    """
    Crea un nuevo usuario según los permisos:
    - Nivel 2 crea cajeros
    - Nivel 3 crea encargados
    - Nivel 4 crea auditores
    - Nivel 5 (jefe_auditor) crea admin_anticipos
    """
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    rol = data.get('rol', '').strip()
    local = data.get('local', '').strip()

    # Validación básica
    if not (username and rol):
        return jsonify(success=False, msg='Faltan datos requeridos'), 400

    user_level = get_user_level()

    # Validar que el usuario puede crear ese rol
    if user_level == 2 and rol != 'cajero':
        return jsonify(success=False, msg='Solo podés crear cajeros'), 403
    elif user_level == 3 and rol != 'encargado':
        return jsonify(success=False, msg='Solo podés crear encargados'), 403
    elif user_level == 4 and rol != 'auditor':
        return jsonify(success=False, msg='Solo podés crear auditores'), 403
    elif user_level >= 5 and rol not in ('auditor', 'admin_anticipos'):
        return jsonify(success=False, msg='Solo podés crear auditores o administradores de anticipos'), 403

    # Si se está creando un auditor o admin_anticipos y no hay local, asignar "OFICINA CENTRAL"
    if rol in ('auditor', 'admin_anticipos') and not local:
        local = 'OFICINA CENTRAL'

    # Para otros roles, el local es obligatorio
    if not local:
        return jsonify(success=False, msg='El local es requerido'), 400

    # Obtener society del local desde la tabla locales
    conn_temp = get_db_connection()
    cur_temp = conn_temp.cursor(dictionary=True)
    try:
        cur_temp.execute("SELECT society FROM locales WHERE local = %s LIMIT 1", (local,))
        local_info = cur_temp.fetchone()
        society = local_info['society'] if local_info and local_info.get('society') else ''
        cur_temp.close()
        conn_temp.close()
    except:
        society = ''
        try:
            cur_temp.close()
        except:
            pass
        try:
            conn_temp.close()
        except:
            pass

    try:
        # Crear usuario con contraseña dummy (será reemplazada en el primer login)
        # El usuario establecerá su propia contraseña la primera vez que se logee
        dummy_password = '__PRIMER_LOGIN_PENDIENTE__'  # Contraseña temporal que nunca se usará
        pages_slugs = ['index', 'resumen_local']  # Páginas por defecto
        result = create_user(username, dummy_password, rol, local, society, pages_slugs, status='active')

        # Marcar como primer login pendiente
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("UPDATE users SET first_login=FALSE WHERE id=%s", (result['user_id'],))
        conn.commit()
        cur.close()
        conn.close()

        return jsonify(success=True, user_id=result['user_id'], username=username)

    except ValueError as ve:
        return jsonify(success=False, msg=str(ve)), 400
    except Exception as e:
        return jsonify(success=False, msg=f'Error al crear usuario: {str(e)}'), 500


@app.route('/api/usuarios/eliminar/<user_id>', methods=['DELETE'])
@login_required
@role_min_required(2)
def api_usuarios_eliminar(user_id):
    """
    Elimina un usuario.
    Solo nivel 2+ puede eliminar usuarios de nivel inferior.
    """
    user_level = get_user_level()

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    try:
        # Obtener información del usuario a eliminar
        cur.execute("""
            SELECT u.id, u.username, r.level AS role_level
            FROM users u
            JOIN roles r ON u.role_id = r.id
            WHERE u.id = %s
        """, (user_id,))
        user_to_delete = cur.fetchone()

        if not user_to_delete:
            return jsonify(success=False, msg='Usuario no encontrado'), 404

        target_level = int(user_to_delete['role_level'])

        # Verificar permisos: solo puedo eliminar usuarios de nivel inferior
        if target_level >= user_level:
            return jsonify(success=False, msg='No tenés permisos para eliminar este usuario'), 403

        # No permitir eliminar al usuario actual
        if user_id == session.get('user_id'):
            return jsonify(success=False, msg='No podés eliminarte a vos mismo'), 400

        # Eliminar relaciones en user_pages
        cur.execute("DELETE FROM user_pages WHERE user_id = %s", (user_id,))

        # Eliminar usuario
        cur.execute("DELETE FROM users WHERE id = %s", (user_id,))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify(success=True, msg=f'Usuario "{user_to_delete["username"]}" eliminado correctamente')

    except Exception as e:
        conn.rollback()
        cur.close()
        conn.close()
        return jsonify(success=False, msg=f'Error al eliminar usuario: {str(e)}'), 500


## ______________________ PRIMER LOGIN Y CAMBIO DE CONTRASEÑA _______________________

@app.route('/api/cambiar_password_primer_login', methods=['POST'])
@login_required
def api_cambiar_password_primer_login():
    """
    Permite al usuario cambiar su contraseña en el primer login.
    Solo funciona si el usuario tiene status='pending_first_login'.
    """
    data = request.get_json() or {}
    nueva_password = data.get('nueva_password', '').strip()

    if not nueva_password:
        return jsonify(success=False, msg='La contraseña no puede estar vacía'), 400

    if len(nueva_password) < 6:
        return jsonify(success=False, msg='La contraseña debe tener al menos 6 caracteres'), 400

    user_id = session.get('user_id')
    if not user_id:
        return jsonify(success=False, msg='Sesión inválida'), 401

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    try:
        # Verificar que el usuario esté en primer login (first_login=FALSE)
        cur.execute("SELECT id, first_login FROM users WHERE id=%s", (user_id,))
        user = cur.fetchone()

        if not user:
            cur.close()
            conn.close()
            return jsonify(success=False, msg='Usuario no encontrado'), 404

        if user.get('first_login') == 1 or user.get('first_login') == True:
            cur.close()
            conn.close()
            return jsonify(success=False, msg='Este usuario ya realizó su primer login'), 400

        # Hashear nueva contraseña
        pw_hash = bcrypt.hashpw(nueva_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

        # Actualizar contraseña y marcar first_login como TRUE
        cur.execute("""
            UPDATE users
            SET password=%s, first_login=TRUE
            WHERE id=%s
        """, (pw_hash, user_id))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify(success=True, msg='Contraseña actualizada correctamente')

    except Exception as e:
        try:
            conn.rollback()
        except:
            pass
        try:
            cur.close()
        except:
            pass
        try:
            conn.close()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


# ==================================================================================
# ENDPOINTS: Gestión de usuarios de anticipos (solo para admin_anticipos nivel 6)
# ==================================================================================

@app.route('/api/usuarios_anticipos/listar', methods=['GET'])
@login_required
@role_min_required(6)  # Solo admin_anticipos
def api_usuarios_anticipos_listar():
    """
    Lista todos los usuarios con rol 'anticipos' (nivel 4) con sus locales asignados.
    Solo accesible para admin_anticipos (nivel 6).
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Obtener usuarios con rol anticipos (nivel 4)
        cur.execute("""
            SELECT
                u.id,
                u.username,
                u.status,
                u.created_at,
                r.name as role_name,
                r.level as role_level
            FROM users u
            INNER JOIN roles r ON r.id = u.role_id
            WHERE r.level = 4
            ORDER BY u.username ASC
        """)
        usuarios = cur.fetchall()

        # Para cada usuario, obtener sus locales asignados
        for usuario in usuarios:
            cur.execute("""
                SELECT local
                FROM user_local_permissions
                WHERE username = %s
                ORDER BY local ASC
            """, (usuario['username'],))
            locales_rows = cur.fetchall()
            usuario['locales_asignados'] = [row['local'] for row in locales_rows]

        cur.close()
        conn.close()

        return jsonify(success=True, usuarios=usuarios)

    except Exception as e:
        print(f"❌ ERROR api_usuarios_anticipos_listar: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/usuarios_anticipos/crear', methods=['POST'])
@login_required
@role_min_required(6)  # Solo admin_anticipos
def api_usuarios_anticipos_crear():
    """
    Crea un nuevo usuario con rol 'anticipos' (nivel 4).
    Body: {
        "username": "juan.anticipos",
        "locales": ["Ribs Infanta", "La Mala"]  // Opcional
    }
    """
    try:
        data = request.get_json() or {}
        username = (data.get('username') or '').strip()
        locales = data.get('locales') or []

        if not username:
            return jsonify(success=False, msg='Username requerido'), 400

        if len(username) < 3:
            return jsonify(success=False, msg='Username debe tener al menos 3 caracteres'), 400

        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que el usuario no exista
        cur.execute("SELECT id FROM users WHERE username = %s", (username,))
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f'El usuario {username} ya existe'), 400

        # Obtener el role_id del rol 'anticipos' (nivel 4)
        cur.execute("SELECT id FROM roles WHERE level = 4 LIMIT 1")
        role_row = cur.fetchone()
        if not role_row:
            cur.close()
            conn.close()
            return jsonify(success=False, msg='Rol anticipos no encontrado. Ejecutá la migración de roles.'), 500

        role_id = role_row['id']

        # Generar UUID para el usuario
        user_id = str(uuid.uuid4())

        # Crear el usuario
        # Password vacío + first_login=0 = cualquier contraseña en el primer login será válida
        # Los campos local y society se dejan vacíos porque los anticipos se manejan por user_local_permissions
        cur.execute("""
            INSERT INTO users (id, username, password, role_id, local, society, status, first_login, created_at)
            VALUES (%s, %s, '', %s, '', '', 'active', 0, NOW())
        """, (user_id, username, role_id))

        # Nota: No usar lastrowid porque estamos usando UUID, no AUTO_INCREMENT

        # Asignar locales si se especificaron
        admin_username = session.get('username', 'sistema')
        for local in locales:
            local = local.strip()
            if local:
                try:
                    cur.execute("""
                        INSERT INTO user_local_permissions (username, local, created_by, created_at)
                        VALUES (%s, %s, %s, NOW())
                    """, (username, local, admin_username))
                except Exception as e:
                    print(f"⚠️  Warning: No se pudo asignar local {local}: {e}")

        conn.commit()

        # Registrar en auditoría
        from modules.tabla_auditoria import registrar_auditoria
        registrar_auditoria(
            conn=conn,
            accion='INSERT',
            tabla='users',
            registro_id=username,  # Usar username en lugar de UUID para compatibilidad
            datos_nuevos={
                'id': user_id,
                'username': username,
                'role': 'anticipos',
                'locales_asignados': locales
            },
            descripcion=f"Usuario anticipos creado: {username}"
        )

        cur.close()
        conn.close()

        return jsonify(success=True, msg=f'Usuario {username} creado correctamente', user_id=user_id)

    except Exception as e:
        print(f"❌ ERROR api_usuarios_anticipos_crear: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/usuarios_anticipos/<username>/locales/asignar', methods=['POST'])
@login_required
@role_min_required(6)  # Solo admin_anticipos
def api_usuarios_anticipos_asignar_local(username):
    """
    Asigna un local a un usuario de anticipos.
    Body: { "local": "Ribs Infanta" }
    """
    try:
        data = request.get_json() or {}
        local = (data.get('local') or '').strip()

        if not local:
            return jsonify(success=False, msg='Local requerido'), 400

        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que el usuario existe y es de anticipos
        cur.execute("""
            SELECT u.id, r.level
            FROM users u
            INNER JOIN roles r ON r.id = u.role_id
            WHERE u.username = %s
        """, (username,))
        user = cur.fetchone()

        if not user:
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f'Usuario {username} no encontrado'), 404

        if user['level'] != 4:
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f'El usuario {username} no es de tipo anticipos'), 400

        # Verificar si ya tiene el local asignado
        cur.execute("""
            SELECT id FROM user_local_permissions
            WHERE username = %s AND local = %s
        """, (username, local))

        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f'El local {local} ya está asignado a {username}'), 400

        # Asignar el local
        admin_username = session.get('username', 'sistema')
        cur.execute("""
            INSERT INTO user_local_permissions (username, local, created_by, created_at)
            VALUES (%s, %s, %s, NOW())
        """, (username, local, admin_username))

        conn.commit()

        # Registrar en auditoría
        from modules.tabla_auditoria import registrar_auditoria
        registrar_auditoria(
            conn=conn,
            accion='INSERT',
            tabla='user_local_permissions',
            registro_id=cur.lastrowid,
            datos_nuevos={
                'username': username,
                'local': local,
                'created_by': admin_username
            },
            descripcion=f"Local {local} asignado a {username}"
        )

        cur.close()
        conn.close()

        return jsonify(success=True, msg=f'Local {local} asignado correctamente a {username}')

    except Exception as e:
        print(f"❌ ERROR api_usuarios_anticipos_asignar_local: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/usuarios_anticipos/<username>/locales/<local>/quitar', methods=['DELETE'])
@login_required
@role_min_required(6)  # Solo admin_anticipos
def api_usuarios_anticipos_quitar_local(username, local):
    """
    Quita un local asignado a un usuario de anticipos.
    """
    try:
        from urllib.parse import unquote
        local = unquote(local)  # Decodificar URL encoding

        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que el usuario existe y es de anticipos
        cur.execute("""
            SELECT u.id, r.level
            FROM users u
            INNER JOIN roles r ON r.id = u.role_id
            WHERE u.username = %s
        """, (username,))
        user = cur.fetchone()

        if not user:
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f'Usuario {username} no encontrado'), 404

        if user['level'] != 4:
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f'El usuario {username} no es de tipo anticipos'), 400

        # Eliminar el permiso
        cur.execute("""
            DELETE FROM user_local_permissions
            WHERE username = %s AND local = %s
        """, (username, local))

        rows_affected = cur.rowcount
        conn.commit()

        # Registrar en auditoría
        from modules.tabla_auditoria import registrar_auditoria
        registrar_auditoria(
            conn=conn,
            accion='DELETE',
            tabla='user_local_permissions',
            registro_id=0,
            datos_anteriores={
                'username': username,
                'local': local
            },
            descripcion=f"Local {local} removido de {username}"
        )

        cur.close()
        conn.close()

        if rows_affected == 0:
            return jsonify(success=False, msg=f'El local {local} no estaba asignado a {username}'), 404

        return jsonify(success=True, msg=f'Local {local} removido correctamente de {username}')

    except Exception as e:
        print(f"❌ ERROR api_usuarios_anticipos_quitar_local: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/usuarios_anticipos/resetear_password', methods=['POST'])
@login_required
@role_min_required(6)  # Solo admin_anticipos
def api_usuarios_anticipos_resetear_password():
    """
    Resetea la contraseña de un usuario de anticipos.
    El usuario podrá usar cualquier contraseña en el próximo login.
    Body: { "username": "juan.anticipos" }
    """
    try:
        data = request.get_json() or {}
        username = (data.get('username') or '').strip()

        if not username:
            return jsonify(success=False, msg='Username requerido'), 400

        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que el usuario existe y es de anticipos
        cur.execute("""
            SELECT u.id, r.level
            FROM users u
            INNER JOIN roles r ON r.id = u.role_id
            WHERE u.username = %s
        """, (username,))
        user = cur.fetchone()

        if not user:
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f'Usuario {username} no encontrado'), 404

        if user['level'] != 4:
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f'El usuario {username} no es de tipo anticipos'), 400

        # Resetear password (vacío) y first_login=0
        cur.execute("""
            UPDATE users
            SET password = '', first_login = 0
            WHERE username = %s
        """, (username,))

        conn.commit()

        # Registrar en auditoría
        from modules.tabla_auditoria import registrar_auditoria
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='users',
            registro_id=user['id'],
            datos_nuevos={
                'password_reset': True,
                'first_login': 0
            },
            descripcion=f"Contraseña reseteada para {username}"
        )

        cur.close()
        conn.close()

        return jsonify(success=True, msg=f'Contraseña reseteada para {username}')

    except Exception as e:
        print(f"❌ ERROR api_usuarios_anticipos_resetear_password: {e}")
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


## ====== ANTICIPOS PARA AUDITORES (SOLO VISUALIZACIÓN) ======

@app.route('/anticipos-auditor')
@login_required
@role_min_required(3)  # Solo auditores (nivel 3+)
def anticipos_auditor_page():
    """
    Página de visualización de anticipos para auditores (solo lectura).
    Los auditores pueden ver todos los anticipos con filtros y detalles.
    """
    return render_template('anticipos_auditor.html')


@app.route('/api/anticipos_auditor/listar', methods=['GET'])
@login_required
@role_min_required(3)  # Solo auditores (nivel 3+)
def listar_anticipos_auditor():
    """
    Listar TODOS los anticipos recibidos para auditores (solo lectura).
    Los auditores pueden ver todos los anticipos sin restricción de locales.

    Params:
    - estado: (opcional) 'pendiente', 'consumido', 'eliminado_global'
    - local: (opcional) filtrar por local
    - fecha_desde, fecha_hasta: (opcional) filtrar por fecha_evento
    """
    estado = request.args.get('estado', '').strip()
    local = request.args.get('local', '').strip()
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Query similar al de listar_anticipos_recibidos pero sin restricción de permisos
        sql = """
            SELECT
                ar.id, ar.fecha_pago, ar.fecha_evento, ar.importe, ar.divisa,
                ar.tipo_cambio_fecha, ar.cliente,
                ar.numero_transaccion, ar.medio_pago, ar.observaciones, ar.local,
                ar.estado as estado_global,
                ar.created_by, ar.created_at, ar.updated_by, ar.updated_at,
                ar.deleted_by, ar.deleted_at,
                -- Verificar si fue consumido realmente
                (SELECT COUNT(*) FROM anticipos_estados_caja aec
                 WHERE aec.anticipo_id = ar.id AND aec.estado = 'consumido') as fue_consumido,
                -- Verificar si tiene adjunto
                CASE WHEN EXISTS (
                    SELECT 1 FROM imagenes_adjuntos ia
                    WHERE ia.entity_type = 'anticipo_recibido'
                      AND ia.entity_id = ar.id
                      AND ia.estado = 'active'
                ) THEN 1 ELSE 0 END as tiene_adjunto
            FROM anticipos_recibidos ar
            WHERE 1=1
        """
        params = []

        # Filtro por estado
        if estado:
            if estado == 'pendiente':
                sql += """ AND ar.estado = 'pendiente'
                          AND (SELECT COUNT(*) FROM anticipos_estados_caja aec
                               WHERE aec.anticipo_id = ar.id AND aec.estado = 'consumido') = 0 """
            elif estado == 'consumido':
                sql += """ AND (SELECT COUNT(*) FROM anticipos_estados_caja aec
                               WHERE aec.anticipo_id = ar.id AND aec.estado = 'consumido') > 0 """
            elif estado == 'eliminado_global':
                sql += " AND ar.estado = 'eliminado_global' "

        # Filtro por local
        if local:
            sql += " AND ar.local = %s "
            params.append(local)

        # Filtros por fecha_evento
        if fecha_desde:
            sql += " AND ar.fecha_evento >= %s "
            params.append(fecha_desde)

        if fecha_hasta:
            sql += " AND ar.fecha_evento <= %s "
            params.append(fecha_hasta)

        sql += " ORDER BY ar.fecha_evento DESC, ar.created_at DESC "

        cur.execute(sql, params)
        rows = cur.fetchall()

        # Procesar cada anticipo para convertir tipos
        anticipos = []
        for anticipo in rows:
            # Convertir Decimal a float
            if anticipo.get('importe') is not None:
                anticipo['importe'] = float(anticipo['importe'])
            if anticipo.get('tipo_cambio_fecha') is not None and isinstance(anticipo['tipo_cambio_fecha'], (int, float)):
                anticipo['tipo_cambio_fecha'] = float(anticipo['tipo_cambio_fecha'])

            # Convertir fechas a string
            if anticipo.get('fecha_pago'):
                if hasattr(anticipo['fecha_pago'], 'strftime'):
                    anticipo['fecha_pago'] = anticipo['fecha_pago'].strftime('%Y-%m-%d')
                else:
                    anticipo['fecha_pago'] = str(anticipo['fecha_pago'])

            if anticipo.get('fecha_evento'):
                if hasattr(anticipo['fecha_evento'], 'strftime'):
                    anticipo['fecha_evento'] = anticipo['fecha_evento'].strftime('%Y-%m-%d')
                else:
                    anticipo['fecha_evento'] = str(anticipo['fecha_evento'])

            # Convertir timestamps a ISO string
            for campo in ['created_at', 'updated_at', 'deleted_at']:
                if anticipo.get(campo):
                    if hasattr(anticipo[campo], 'isoformat'):
                        anticipo[campo] = anticipo[campo].isoformat()
                    else:
                        anticipo[campo] = str(anticipo[campo])

            # Si fue consumido en alguna caja, el estado debe ser 'consumido'
            if anticipo['fue_consumido'] > 0:
                anticipo['estado'] = 'consumido'
            else:
                anticipo['estado'] = anticipo['estado_global']

            # Eliminar campos auxiliares
            del anticipo['estado_global']
            del anticipo['fue_consumido']

            anticipos.append(anticipo)

        cur.close()
        conn.close()

        return jsonify(success=True, anticipos=anticipos)

    except Exception as e:
        print("❌ ERROR listar_anticipos_auditor:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


## ====== MEDIOS DE PAGO PARA ANTICIPOS ======

@app.route('/api/medios_anticipos/listar', methods=['GET'])
@login_required
@role_min_required(6)  # Solo admin_anticipos
def api_medios_anticipos_listar():
    """
    Listar todos los medios de pago para anticipos.
    Solo para admin_anticipos (level 6+)
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT
                id,
                nombre,
                activo,
                es_efectivo,
                created_at,
                updated_at
            FROM medios_anticipos
            ORDER BY
                CASE nombre
                    WHEN 'Efectivo' THEN 1
                    ELSE 2
                END,
                nombre ASC
        """)

        medios = cur.fetchall()

        # Convertir timestamps a string
        for medio in medios:
            if medio.get('created_at'):
                medio['created_at'] = medio['created_at'].isoformat() if hasattr(medio['created_at'], 'isoformat') else str(medio['created_at'])
            if medio.get('updated_at'):
                medio['updated_at'] = medio['updated_at'].isoformat() if hasattr(medio['updated_at'], 'isoformat') else str(medio['updated_at'])

        cur.close()
        conn.close()

        return jsonify(success=True, medios=medios)

    except Exception as e:
        print("❌ ERROR api_medios_anticipos_listar:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/medios_anticipos/crear', methods=['POST'])
@login_required
@role_min_required(6)  # Solo admin_anticipos
def api_medios_anticipos_crear():
    """
    Crear un nuevo medio de pago para anticipos.
    Solo para admin_anticipos (level 6+)

    Body JSON:
    {
        "nombre": "Nombre del medio",
        "es_efectivo": 0/1
    }
    """
    try:
        data = request.get_json()
        nombre = data.get('nombre', '').strip()
        es_efectivo = int(data.get('es_efectivo', 0))

        if not nombre:
            return jsonify(success=False, msg='El nombre es obligatorio'), 400

        conn = get_db_connection()
        cur = conn.cursor()

        # Verificar si ya existe
        cur.execute("SELECT id FROM medios_anticipos WHERE nombre = %s", (nombre,))
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f'Ya existe un medio de pago con el nombre "{nombre}"'), 400

        # Insertar
        cur.execute("""
            INSERT INTO medios_anticipos (nombre, es_efectivo, activo)
            VALUES (%s, %s, 1)
        """, (nombre, es_efectivo))

        conn.commit()
        nuevo_id = cur.lastrowid

        cur.close()
        conn.close()

        print(f"✅ Medio de pago creado: ID={nuevo_id}, nombre={nombre}, es_efectivo={es_efectivo}")

        return jsonify(success=True, msg='Medio de pago creado correctamente', id=nuevo_id)

    except Exception as e:
        print("❌ ERROR api_medios_anticipos_crear:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/medios_anticipos/<int:medio_id>', methods=['DELETE'])
@login_required
@role_min_required(6)  # Solo admin_anticipos
def api_medios_anticipos_eliminar(medio_id):
    """
    Eliminar (desactivar) un medio de pago.
    Solo para admin_anticipos (level 6+)

    IMPORTANTE: No se elimina físicamente, solo se desactiva (activo=0)
    para no romper referencias en anticipos existentes.
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que existe
        cur.execute("SELECT nombre, es_efectivo FROM medios_anticipos WHERE id = %s", (medio_id,))
        medio = cur.fetchone()

        if not medio:
            cur.close()
            conn.close()
            return jsonify(success=False, msg='Medio de pago no encontrado'), 404

        # No permitir eliminar "Efectivo"
        if medio['nombre'] == 'Efectivo':
            cur.close()
            conn.close()
            return jsonify(success=False, msg='No se puede eliminar el medio "Efectivo" (es el medio por defecto)'), 400

        # Verificar si hay anticipos usando este medio
        cur.execute("SELECT COUNT(*) as count FROM anticipos_recibidos WHERE medio_pago_id = %s", (medio_id,))
        result = cur.fetchone()
        anticipos_count = result['count'] if result else 0

        # Desactivar (no eliminar físicamente)
        cur.execute("UPDATE medios_anticipos SET activo = 0 WHERE id = %s", (medio_id,))
        conn.commit()

        cur.close()
        conn.close()

        msg = f'Medio de pago "{medio["nombre"]}" desactivado correctamente'
        if anticipos_count > 0:
            msg += f' (había {anticipos_count} anticipo(s) usando este medio)'

        print(f"✅ {msg}")

        return jsonify(success=True, msg=msg)

    except Exception as e:
        print("❌ ERROR api_medios_anticipos_eliminar:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/medios_anticipos/activos', methods=['GET'])
@login_required
def api_medios_anticipos_activos():
    """
    Listar solo medios de pago activos (para uso en formularios).
    Accesible para cualquier usuario autenticado.
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        cur.execute("""
            SELECT
                id,
                nombre,
                es_efectivo
            FROM medios_anticipos
            WHERE activo = 1
            ORDER BY
                CASE nombre
                    WHEN 'Efectivo' THEN 1
                    ELSE 2
                END,
                nombre ASC
        """)

        medios = cur.fetchall()

        cur.close()
        conn.close()

        return jsonify(success=True, medios=medios)

    except Exception as e:
        print("❌ ERROR api_medios_anticipos_activos:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


## ====== REMESAS NO RETIRADAS - GESTIÓN CENTRALIZADA ======

@app.route('/remesas-no-retiradas')
@login_required
@role_min_required(2)  # Encargados (nivel 2+) y auditores (nivel 3+)
def remesas_no_retiradas_page():
    """
    Página de gestión de remesas no retiradas.
    - Encargados: ven solo su local, pueden marcar como retiradas
    - Auditores: ven todos los locales, pueden editar
    """
    return render_template('remesas_no_retiradas.html')


@app.route('/api/remesas-no-retiradas/listar', methods=['GET'])
@login_required
@role_min_required(2)
def listar_remesas_no_retiradas():
    """
    Lista remesas no retiradas.
    - Encargados: solo ven su local
    - Auditores: ven todos los locales (con filtro opcional)

    Query params:
    - local (opcional): filtrar por local
    - fecha_desde (opcional): filtrar desde fecha
    - fecha_hasta (opcional): filtrar hasta fecha
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Obtener nivel del usuario
        user_id = session.get('user_id')
        cur.execute("SELECT role_id FROM users WHERE id = %s", (user_id,))
        user_row = cur.fetchone()
        if not user_row:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Usuario no encontrado"), 404

        cur.execute("SELECT level FROM roles WHERE id = %s", (user_row['role_id'],))
        role_row = cur.fetchone()
        user_level = role_row['level'] if role_row else 0

        # Construir query base
        # Filtrar SOLO las NO retiradas: excluir explícitamente valores "retirados"
        query = """
            SELECT
                id,
                local,
                caja,
                turno,
                fecha,
                nro_remesa,
                precinto,
                monto,
                retirada,
                retirada_por,
                fecha_retirada,
                usuario,
                ult_mod
            FROM remesas_trns
            WHERE retirada NOT IN (1, 'Si', 'Sí', 'sí', 'si', 'SI', 'SÍ')
              AND (retirada = 0 OR retirada = 'No' OR retirada IS NULL OR retirada = '')
        """
        params = []

        # Filtros según nivel
        if user_level < 3:
            # Encargado: solo su local
            user_local = session.get('local')
            if user_local:
                query += " AND local = %s"
                params.append(user_local)
        else:
            # Auditor: puede filtrar por local o ver todos
            local_filtro = request.args.get('local', '').strip()
            if local_filtro:
                query += " AND local = %s"
                params.append(local_filtro)

        # Filtros de fecha
        fecha_desde = request.args.get('fecha_desde', '').strip()
        if fecha_desde:
            query += " AND DATE(fecha) >= %s"
            params.append(fecha_desde)

        fecha_hasta = request.args.get('fecha_hasta', '').strip()
        if fecha_hasta:
            query += " AND DATE(fecha) <= %s"
            params.append(fecha_hasta)

        query += " ORDER BY fecha DESC, id DESC"

        cur.execute(query, params)
        rows = cur.fetchall()

        remesas = []
        for row in rows:
            remesa = dict(row)

            # Convertir tipos
            if remesa.get('monto'):
                remesa['monto'] = float(remesa['monto'])

            if remesa.get('fecha'):
                if hasattr(remesa['fecha'], 'strftime'):
                    remesa['fecha'] = remesa['fecha'].strftime('%Y-%m-%d')
                else:
                    remesa['fecha'] = str(remesa['fecha'])

            if remesa.get('fecha_retirada'):
                if hasattr(remesa['fecha_retirada'], 'strftime'):
                    remesa['fecha_retirada'] = remesa['fecha_retirada'].strftime('%Y-%m-%d')
                elif remesa['fecha_retirada']:
                    remesa['fecha_retirada'] = str(remesa['fecha_retirada'])

            if remesa.get('ult_mod'):
                if hasattr(remesa['ult_mod'], 'isoformat'):
                    remesa['ult_mod'] = remesa['ult_mod'].isoformat()
                else:
                    remesa['ult_mod'] = str(remesa['ult_mod'])

            remesas.append(remesa)

        cur.close()
        conn.close()

        return jsonify(success=True, remesas=remesas, user_level=user_level)

    except Exception as e:
        print("❌ ERROR listar_remesas_no_retiradas:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/remesas-no-retiradas/<int:remesa_id>/marcar-retirada', methods=['POST'])
@login_required
@role_min_required(2)
def marcar_remesa_retirada(remesa_id):
    """
    Marca una remesa como retirada.
    - Requiere: fecha_retirada, retirada_por
    - Registra en auditoría
    """
    try:
        data = request.get_json() or {}
        fecha_retirada = data.get('fecha_retirada', '').strip()
        retirada_por = data.get('retirada_por', '').strip()

        if not fecha_retirada:
            return jsonify(success=False, msg="La fecha de retiro es requerida"), 400

        if not retirada_por:
            return jsonify(success=False, msg="El nombre de quien retira es requerido"), 400

        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Obtener remesa actual para auditoría
        cur.execute("""
            SELECT id, local, caja, fecha, turno, monto, retirada, retirada_por, fecha_retirada, estado_contable
            FROM remesas_trns
            WHERE id = %s
        """, (remesa_id,))
        remesa = cur.fetchone()

        if not remesa:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Remesa no encontrada"), 404

        # VALIDACIÓN DE SEGURIDAD 1: Verificar que no esté ya retirada
        retirada_val = str(remesa['retirada']).strip() if remesa['retirada'] is not None else ''
        if retirada_val.lower() in ('1', 'si', 'sí', 'true') or remesa['retirada'] in (1, True):
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Esta remesa ya está marcada como retirada"), 400

        # VALIDACIÓN DE SEGURIDAD 2: Solo permitir si está en estado Local
        estado_actual = str(remesa.get('estado_contable', '')).strip().upper()
        if estado_actual not in ('', 'LOCAL', 'NONE'):
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f"No se puede marcar como retirada una remesa en estado {estado_actual}"), 400

        # VALIDACIÓN DE SEGURIDAD 3: Verificar permisos (encargados solo su local)
        user_level = session.get('role_level', 0)
        if user_level < 3:  # Encargado
            user_local = session.get('local')
            if remesa['local'] != user_local:
                cur.close()
                conn.close()
                return jsonify(success=False, msg="No tenés permisos para modificar remesas de otro local"), 403

        # Guardar datos anteriores para auditoría
        datos_anteriores = {
            'retirada': str(remesa['retirada']),
            'retirada_por': remesa['retirada_por'],
            'fecha_retirada': remesa['fecha_retirada'].strftime('%Y-%m-%d') if remesa['fecha_retirada'] and hasattr(remesa['fecha_retirada'], 'strftime') else str(remesa['fecha_retirada']) if remesa['fecha_retirada'] else None,
            'estado_contable': 'Local'
        }

        # Actualizar remesa y cambiar estado a TRAN
        cur.execute("""
            UPDATE remesas_trns
            SET retirada = 1,
                retirada_por = %s,
                fecha_retirada = %s,
                estado_contable = 'TRAN',
                fecha_estado_tran = %s,
                ult_mod = NOW()
            WHERE id = %s
        """, (retirada_por, fecha_retirada, fecha_retirada, remesa_id))

        conn.commit()

        # Datos nuevos para auditoría
        datos_nuevos = {
            'retirada': 1,
            'retirada_por': retirada_por,
            'fecha_retirada': fecha_retirada,
            'estado_contable': 'TRAN',
            'fecha_estado_tran': fecha_retirada
        }

        # Registrar en auditoría
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='remesas_trns',
            registro_id=remesa_id,
            datos_anteriores=datos_anteriores,
            datos_nuevos=datos_nuevos,
            descripcion=f"Remesa marcada como retirada - Local: {remesa['local']}, Fecha caja: {remesa['fecha']}, Monto: ${remesa['monto']}"
        )

        cur.close()
        conn.close()

        return jsonify(success=True, msg=f"Remesa marcada como retirada correctamente")

    except Exception as e:
        print("❌ ERROR marcar_remesa_retirada:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/remesas-no-retiradas/<int:remesa_id>/editar', methods=['PUT'])
@login_required
@role_min_required(3)  # Solo auditores
def editar_remesa_retirada(remesa_id):
    """
    Edita fecha_retirada o retirada_por de una remesa.
    Solo para auditores (correcciones).
    - Registra en auditoría
    """
    try:
        data = request.get_json() or {}
        fecha_retirada = data.get('fecha_retirada', '').strip()
        retirada_por = data.get('retirada_por', '').strip()

        if not fecha_retirada and not retirada_por:
            return jsonify(success=False, msg="Debés proporcionar al menos un campo a editar"), 400

        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Obtener remesa actual
        cur.execute("""
            SELECT id, local, caja, fecha, turno, monto, retirada, retirada_por, fecha_retirada, estado_contable
            FROM remesas_trns
            WHERE id = %s
        """, (remesa_id,))
        remesa = cur.fetchone()

        if not remesa:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Remesa no encontrada"), 404

        # VALIDACIÓN DE SEGURIDAD: Solo auditores pueden editar, pero no remesas contabilizadas
        estado_actual = str(remesa.get('estado_contable', '')).strip().upper()
        if estado_actual == 'CONTABILIZADA':
            cur.close()
            conn.close()
            return jsonify(success=False, msg="No se puede editar una remesa ya contabilizada. Contactá a un administrador."), 403

        # Guardar datos anteriores
        datos_anteriores = {
            'retirada': str(remesa['retirada']),
            'retirada_por': remesa['retirada_por'],
            'fecha_retirada': remesa['fecha_retirada'].strftime('%Y-%m-%d') if remesa['fecha_retirada'] and hasattr(remesa['fecha_retirada'], 'strftime') else str(remesa['fecha_retirada']) if remesa['fecha_retirada'] else None
        }

        # Construir UPDATE dinámico
        updates = []
        params = []

        if fecha_retirada:
            updates.append("fecha_retirada = %s")
            params.append(fecha_retirada)

        if retirada_por:
            updates.append("retirada_por = %s")
            params.append(retirada_por)

        updates.append("ult_mod = NOW()")
        params.append(remesa_id)

        query = f"UPDATE remesas_trns SET {', '.join(updates)} WHERE id = %s"
        cur.execute(query, params)
        conn.commit()

        # Datos nuevos
        datos_nuevos = {
            'retirada': str(remesa['retirada']),
            'retirada_por': retirada_por if retirada_por else remesa['retirada_por'],
            'fecha_retirada': fecha_retirada if fecha_retirada else datos_anteriores['fecha_retirada']
        }

        # Registrar en auditoría
        registrar_auditoria(
            conn=conn,
            accion='UPDATE',
            tabla='remesas_trns',
            registro_id=remesa_id,
            datos_anteriores=datos_anteriores,
            datos_nuevos=datos_nuevos,
            descripcion=f"Auditor editó datos de retiro - Local: {remesa['local']}, Fecha caja: {remesa['fecha']}"
        )

        cur.close()
        conn.close()

        return jsonify(success=True, msg="Datos de retiro actualizados correctamente")

    except Exception as e:
        print("❌ ERROR editar_remesa_retirada:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/remesas-no-retiradas/contador', methods=['GET'])
@login_required
@role_min_required(2)
def contador_remesas_no_retiradas():
    """
    Devuelve el contador de remesas no retiradas.
    - Encargados: solo su local
    - Auditores: todos los locales (con filtro opcional)
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Obtener nivel del usuario
        user_id = session.get('user_id')
        cur.execute("SELECT role_id FROM users WHERE id = %s", (user_id,))
        user_row = cur.fetchone()
        if not user_row:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Usuario no encontrado"), 404

        cur.execute("SELECT level FROM roles WHERE id = %s", (user_row['role_id'],))
        role_row = cur.fetchone()
        user_level = role_row['level'] if role_row else 0

        # Query según nivel
        if user_level < 3:
            # Encargado: solo su local
            user_local = session.get('local')
            cur.execute("""
                SELECT COUNT(*) as total
                FROM remesas_trns
                WHERE retirada NOT IN (1, 'Si', 'Sí', 'sí', 'si', 'SI', 'SÍ')
                  AND (retirada = 0 OR retirada = 'No' OR retirada IS NULL OR retirada = '')
                  AND local = %s
            """, (user_local,))
        else:
            # Auditor: todos los locales
            cur.execute("""
                SELECT COUNT(*) as total
                FROM remesas_trns
                WHERE retirada NOT IN (1, 'Si', 'Sí', 'sí', 'si', 'SI', 'SÍ')
                  AND (retirada = 0 OR retirada = 'No' OR retirada IS NULL OR retirada = '')
            """)

        row = cur.fetchone()
        total = row['total'] if row else 0

        cur.close()
        conn.close()

        return jsonify(success=True, total=total)

    except Exception as e:
        print("❌ ERROR contador_remesas_no_retiradas:", e)
        return jsonify(success=False, msg=str(e)), 500


# ===========================
# AUDITORÍA DE REMESAS - TRAZABILIDAD COMPLETA
# Solo accesible para admin_tesoreria (nivel 8)
# ===========================

@app.route('/auditoria-remesas')
@login_required
@role_min_required(8)  # Solo admin_tesoreria
def auditoria_remesas_page():
    """
    Página de auditoría completa de remesas.
    Muestra todos los cambios registrados en remesas_trns.
    """
    return render_template('auditoria_remesas.html')


@app.route('/api/auditoria-remesas/listar', methods=['GET'])
@login_required
@role_min_required(8)
def api_auditoria_remesas_listar():
    """
    Lista todos los registros de auditoría de remesas_trns con filtros.

    Query params:
    - nro_remesa (opcional): filtrar por número de remesa
    - precinto (opcional): filtrar por precinto
    - local (opcional): filtrar por local
    - fecha_desde (opcional): filtro de fecha desde (changed_at)
    - fecha_hasta (opcional): filtro de fecha hasta (changed_at)
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Obtener filtros
        nro_remesa = request.args.get('nro_remesa', '').strip()
        precinto = request.args.get('precinto', '').strip()
        local = request.args.get('local', '').strip()
        fecha_desde = request.args.get('fecha_desde', '').strip()
        fecha_hasta = request.args.get('fecha_hasta', '').strip()

        # Construir query base
        # La tabla auditoria tiene: fecha_hora, usuario, accion, tabla, registro_id,
        # datos_anteriores, datos_nuevos, descripcion
        query_parts = ["""
            SELECT
                a.id,
                DATE_FORMAT(a.fecha_hora, '%Y-%m-%d %H:%i:%s') as fecha_hora,
                a.usuario,
                a.usuario_email,
                a.usuario_nivel,
                a.usuario_ip,
                a.local as local_auditoria,
                a.accion,
                a.tabla,
                a.registro_id,
                a.datos_anteriores,
                a.datos_nuevos,
                a.datos_cambios,
                a.descripcion,
                a.endpoint
            FROM auditoria a
            WHERE a.tabla = 'remesas_trns'
        """]

        params = []

        # Filtros específicos (búsqueda en descripcion o datos JSON)
        if nro_remesa:
            query_parts.append("""
                AND (
                    a.descripcion LIKE %s
                    OR JSON_EXTRACT(a.datos_anteriores, '$.nro_remesa') = %s
                    OR JSON_EXTRACT(a.datos_nuevos, '$.nro_remesa') = %s
                )
            """)
            like_pattern = f"%{nro_remesa}%"
            params.extend([like_pattern, nro_remesa, nro_remesa])

        if precinto:
            query_parts.append("""
                AND (
                    a.descripcion LIKE %s
                    OR JSON_EXTRACT(a.datos_anteriores, '$.precinto') = %s
                    OR JSON_EXTRACT(a.datos_nuevos, '$.precinto') = %s
                )
            """)
            like_pattern = f"%{precinto}%"
            params.extend([like_pattern, precinto, precinto])

        if local:
            query_parts.append("""
                AND (
                    a.local = %s
                    OR a.descripcion LIKE %s
                    OR JSON_EXTRACT(a.datos_anteriores, '$.local') = %s
                    OR JSON_EXTRACT(a.datos_nuevos, '$.local') = %s
                )
            """)
            like_pattern = f"%{local}%"
            params.extend([local, like_pattern, local, local])

        # Filtros de fecha
        if fecha_desde:
            query_parts.append("AND DATE(a.fecha_hora) >= %s")
            params.append(fecha_desde)

        if fecha_hasta:
            query_parts.append("AND DATE(a.fecha_hora) <= %s")
            params.append(fecha_hasta)

        # Ordenar descendentemente (más recientes primero)
        query_parts.append("ORDER BY a.fecha_hora DESC")

        # Limitar resultados para performance
        query_parts.append("LIMIT 1000")

        # Ejecutar query
        query = " ".join(query_parts)
        print(f"🔍 Query auditoría: {query}")
        print(f"📊 Params: {params}")
        print(f"📊 Cantidad de params: {len(params)}")

        # Ejecutar con params solo si hay alguno
        if params:
            cur.execute(query, tuple(params))
        else:
            cur.execute(query)

        registros = cur.fetchall()

        cur.close()
        conn.close()

        print(f"✅ {len(registros)} registros de auditoría encontrados")

        return jsonify(success=True, registros=registros)

    except Exception as e:
        print("❌ ERROR api_auditoria_remesas_listar:", e)
        import traceback
        traceback.print_exc()
        return jsonify(success=False, msg=str(e)), 500


# ===========================
# GESTIÓN DE USUARIOS DE TESORERÍA
# Solo accesible para admin_anticipos (nivel 6) y admin_tesoreria (nivel 8)
# ===========================

@app.route('/gestion-usuarios-tesoreria')
@login_required
@role_min_required(6)  # admin_anticipos y superiores
def gestion_usuarios_tesoreria():
    """
    Página de gestión de usuarios de tesorería.
    Accesible para:
    - admin_anticipos (nivel 6): puede crear y eliminar usuarios de tesorería
    - admin_tesoreria (nivel 8): puede gestionar usuarios de tesorería
    """
    return render_template('gestion_usuarios_tesoreria.html')


@app.route('/api/usuarios-tesoreria/listar', methods=['GET'])
@login_required
@role_min_required(6)
def api_usuarios_tesoreria_listar():
    """
    Lista todos los usuarios con rol 'tesoreria' (nivel 7) y 'admin_tesoreria' (nivel 8).
    Solo accesible para admin_anticipos (nivel 6) y admin_tesoreria (nivel 8).
    """
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT
                u.id,
                u.username,
                r.name as role,
                r.level as role_level,
                u.local,
                u.society,
                u.status,
                u.first_login,
                u.created_at
            FROM users u
            JOIN roles r ON u.role_id = r.id
            WHERE r.name IN ('tesoreria', 'admin_tesoreria')
            ORDER BY r.level DESC, u.created_at DESC
        """)

        users = cur.fetchall() or []
        cur.close()
        conn.close()

        return jsonify(success=True, users=users)

    except Exception as e:
        try:
            cur.close()
        except:
            pass
        try:
            conn.close()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/usuarios-tesoreria/crear', methods=['POST'])
@login_required
@role_min_required(6)
def api_usuarios_tesoreria_crear():
    """
    Crea un nuevo usuario de tesorería.
    Solo admin_anticipos (nivel 6) puede crear usuarios con rol 'tesoreria' (nivel 7).
    Solo admin_tesoreria (nivel 8) puede crear otros admin_tesoreria.
    """
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    rol = data.get('rol', '').strip()  # 'tesoreria' o 'admin_tesoreria'
    local = data.get('local', 'TESORERIA').strip()

    # Validación básica
    if not (username and password and rol):
        return jsonify(success=False, msg='Faltan datos requeridos'), 400

    user_level = get_user_level()

    # Validar permisos según nivel
    if user_level == 6 and rol != 'tesoreria':
        return jsonify(success=False, msg='Solo podés crear usuarios de tesorería (nivel 7)'), 403
    elif user_level < 8 and rol == 'admin_tesoreria':
        return jsonify(success=False, msg='Solo admin_tesoreria puede crear otros admin_tesoreria'), 403

    # Validar que el rol exista
    if rol not in ('tesoreria', 'admin_tesoreria'):
        return jsonify(success=False, msg='Rol inválido'), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que el username no exista
        cur.execute("SELECT id FROM users WHERE username = %s", (username,))
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify(success=False, msg='El nombre de usuario ya existe'), 400

        # Obtener role_id
        cur.execute("SELECT id FROM roles WHERE name = %s LIMIT 1", (rol,))
        role_row = cur.fetchone()
        if not role_row:
            cur.close()
            conn.close()
            return jsonify(success=False, msg=f'No se encontró el rol: {rol}'), 404

        role_id = role_row['id']

        # Crear usuario
        cur.execute("""
            INSERT INTO users (username, password, role_id, local, society, status, first_login, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
        """, (username, password, role_id, local, 'TODOS', 'active', 1))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify(success=True, msg='Usuario de tesorería creado exitosamente')

    except Exception as e:
        try:
            conn.rollback()
        except:
            pass
        try:
            cur.close()
        except:
            pass
        try:
            conn.close()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/usuarios-tesoreria/eliminar/<int:user_id>', methods=['DELETE'])
@login_required
@role_min_required(6)
def api_usuarios_tesoreria_eliminar(user_id):
    """
    Elimina un usuario de tesorería.
    Solo admin_anticipos (nivel 6) y admin_tesoreria (nivel 8) pueden eliminar.
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que el usuario existe y es de tesorería
        cur.execute("""
            SELECT u.id, r.name as role, r.level
            FROM users u
            JOIN roles r ON u.role_id = r.id
            WHERE u.id = %s
        """, (user_id,))

        user = cur.fetchone()
        if not user:
            cur.close()
            conn.close()
            return jsonify(success=False, msg='Usuario no encontrado'), 404

        if user['role'] not in ('tesoreria', 'admin_tesoreria'):
            cur.close()
            conn.close()
            return jsonify(success=False, msg='El usuario no es de tesorería'), 400

        # Eliminar usuario
        cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
        conn.commit()
        cur.close()
        conn.close()

        return jsonify(success=True, msg='Usuario eliminado exitosamente')

    except Exception as e:
        try:
            conn.rollback()
        except:
            pass
        try:
            cur.close()
        except:
            pass
        try:
            conn.close()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/usuarios-tesoreria/resetear-password/<int:user_id>', methods=['POST'])
@login_required
@role_min_required(6)
def api_usuarios_tesoreria_resetear_password(user_id):
    """
    Resetea la contraseña de un usuario de tesorería.
    """
    data = request.get_json() or {}
    new_password = data.get('password', '').strip()

    if not new_password:
        return jsonify(success=False, msg='La contraseña es requerida'), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Verificar que el usuario existe y es de tesorería
        cur.execute("""
            SELECT u.id, r.name as role
            FROM users u
            JOIN roles r ON u.role_id = r.id
            WHERE u.id = %s
        """, (user_id,))

        user = cur.fetchone()
        if not user:
            cur.close()
            conn.close()
            return jsonify(success=False, msg='Usuario no encontrado'), 404

        if user['role'] not in ('tesoreria', 'admin_tesoreria'):
            cur.close()
            conn.close()
            return jsonify(success=False, msg='El usuario no es de tesorería'), 400

        # Actualizar password y forzar cambio en primer login
        cur.execute("""
            UPDATE users
            SET password = %s, first_login = 1
            WHERE id = %s
        """, (new_password, user_id))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify(success=True, msg='Contraseña reseteada. El usuario deberá cambiarla en el primer login.')

    except Exception as e:
        try:
            conn.rollback()
        except:
            pass
        try:
            cur.close()
        except:
            pass
        try:
            conn.close()
        except:
            pass
        return jsonify(success=False, msg=str(e)), 500



# ===========================
# REMESAS RETIRADAS - VISTA HISTÓRICA
# ===========================

@app.route('/remesas-retiradas')
@login_required
@role_min_required(2)  # Encargados (nivel 2+) y auditores (nivel 3+)
def remesas_retiradas_page():
    """
    Página de visualización de remesas retiradas (historial).
    - Encargados: ven solo su local, ordenadas descendentemente
    - Auditores: ven todos los locales con filtros
    """
    return render_template('remesas_retiradas.html')


@app.route('/api/remesas-retiradas/listar', methods=['GET'])
@login_required
@role_min_required(2)
def listar_remesas_retiradas():
    """
    Lista remesas que ya fueron retiradas (retirada = 1 o 'Sí').
    - Encargados: solo ven su local
    - Auditores: ven todos los locales (con filtro opcional)

    Query params:
    - local (opcional): filtrar por local (solo auditores)
    - fecha_desde (opcional): filtro de fecha de caja desde
    - fecha_hasta (opcional): filtro de fecha de caja hasta
    - fecha_retiro_desde (opcional): filtro de fecha de retiro desde
    - fecha_retiro_hasta (opcional): filtro de fecha de retiro hasta
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Obtener nivel del usuario
        user_id = session.get('user_id')
        cur.execute("SELECT role_id FROM users WHERE id = %s", (user_id,))
        user_row = cur.fetchone()
        if not user_row:
            cur.close()
            conn.close()
            return jsonify(success=False, msg="Usuario no encontrado"), 404

        cur.execute("SELECT level FROM roles WHERE id = %s", (user_row['role_id'],))
        role_row = cur.fetchone()
        user_level = role_row['level'] if role_row else 0

        # Obtener filtros de query params
        filtro_local = request.args.get('local', '').strip()
        filtro_fecha_desde = request.args.get('fecha_desde', '').strip()
        filtro_fecha_hasta = request.args.get('fecha_hasta', '').strip()
        filtro_retiro_desde = request.args.get('fecha_retiro_desde', '').strip()
        filtro_retiro_hasta = request.args.get('fecha_retiro_hasta', '').strip()

        # Construir query base
        query_parts = ["""
            SELECT
                id,
                DATE_FORMAT(fecha, '%Y-%m-%d') as fecha,
                local,
                caja,
                turno,
                nro_remesa,
                precinto,
                monto,
                DATE_FORMAT(fecha_retirada, '%Y-%m-%d') as fecha_retirada,
                retirada_por,
                estado_contable
            FROM remesas_trns
            WHERE retirada IN (1, 'Si', 'Sí', 'sí', 'si', 'SI', 'SÍ')
              AND estado_contable IN ('TRAN', 'Contabilizada')
        """]

        params = []

        # Filtrar por local según nivel de usuario
        if user_level < 3:
            # Encargado: solo su local
            user_local = session.get('local')
            query_parts.append("AND local = %s")
            params.append(user_local)
        else:
            # Auditor: filtrar por local si se especifica
            if filtro_local:
                query_parts.append("AND local = %s")
                params.append(filtro_local)

        # Filtros de fecha de caja
        if filtro_fecha_desde:
            query_parts.append("AND fecha >= %s")
            params.append(filtro_fecha_desde)

        if filtro_fecha_hasta:
            query_parts.append("AND fecha <= %s")
            params.append(filtro_fecha_hasta)

        # Filtros de fecha de retiro
        if filtro_retiro_desde:
            query_parts.append("AND fecha_retirada >= %s")
            params.append(filtro_retiro_desde)

        if filtro_retiro_hasta:
            query_parts.append("AND fecha_retirada <= %s")
            params.append(filtro_retiro_hasta)

        # Ordenar descendentemente por fecha_retirada (más recientes primero)
        query_parts.append("ORDER BY fecha_retirada DESC, fecha DESC")

        # Ejecutar query
        query = " ".join(query_parts)
        print(f"🔍 Query remesas retiradas: {query}")
        print(f"📊 Params: {params}")

        cur.execute(query, tuple(params))
        remesas = cur.fetchall()

        cur.close()
        conn.close()

        print(f"✅ {len(remesas)} remesas retiradas encontradas")

        return jsonify(success=True, remesas=remesas)

    except Exception as e:
        print("❌ ERROR listar_remesas_retiradas:", e)
        return jsonify(success=False, msg=str(e)), 500


@app.route('/api/locales-lista', methods=['GET'])
@login_required
def listar_locales():
    """
    Devuelve lista de locales únicos para filtros.
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("SELECT DISTINCT local FROM remesas_trns WHERE local IS NOT NULL AND local != '' ORDER BY local")
        rows = cur.fetchall()

        cur.close()
        conn.close()

        locales = [row[0] for row in rows]

        return jsonify(success=True, locales=locales)

    except Exception as e:
        print("❌ ERROR listar_locales:", e)
        return jsonify(success=False, msg=str(e)), 500
