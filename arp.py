import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Datos de facturación
datos = {
    'TIPO': ['BEBIDAS', 'PLANTA'],
    'ABRIL': [46457544.01, 138102496.87],
    'MAYO': [30837753.58, 111656064.29],
    'JUNIO': [9899456.00, 104292456.50]
}

# Calcular promedios
datos['PROMEDIO'] = [(datos['ABRIL'][0] + datos['MAYO'][0] + datos['JUNIO'][0])/3,
                    (datos['ABRIL'][1] + datos['MAYO'][1] + datos['JUNIO'][1])/3]

# Crear DataFrame
df = pd.DataFrame(datos)

# Exportar a Excel
archivo_excel = 'facturacion_trimestral.xlsx'
df.to_excel(archivo_excel, index=False, sheet_name='Facturación')

# Aplicar formato con openpyxl
wb = load_workbook(archivo_excel)
ws = wb.active

# Estilos
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
money_format = "$#,##0.00"  # Formato correcto para moneda

# Aplicar formato a encabezados
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Aplicar formato a celdas de datos
for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=5):
    for cell in row:
        cell.border = border
        if cell.column > 1:  # Solo para columnas numéricas
            cell.number_format = money_format

# Ajustar ancho de columnas
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

# Guardar el archivo con formato
wb.save(archivo_excel)

print(f"Archivo Excel '{archivo_excel}' generado exitosamente.")